from flask import (
    Flask,
    render_template,
    render_template_string,
    request,
    redirect,
    url_for,
    session,
    flash,
    jsonify,
    send_file,
)
from werkzeug.security import generate_password_hash, check_password_hash
from datetime import (
    datetime,
    timedelta,
    date,
    time,
    date as date_type,
    time as time_type,
)
from dateutil.relativedelta import relativedelta
from collections import defaultdict
import sqlite3
import os
import re
import json
import csv
import io
from functools import wraps
from dotenv import load_dotenv
import pytz
import zoneinfo

# For Excel Exports

import tempfile
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, NamedStyle, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd

# Email System Implementation for Riverside Equestrian
import smtplib
import ssl
import os
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
from datetime import datetime
import threading
from functools import wraps


load_dotenv()
# ===================================
# APPLICATION SETUP
# ===================================

app = Flask(__name__)
app.secret_key = "riverside-equestrian-secret-key-change-in-production"
app.config["DATABASE"] = "cancellation_system.db"
app.config["TIMEZONE"] = "America/Los_Angeles"

# Set environment timezone for the application
os.environ["TZ"] = "America/Los_Angeles"


def init_db():
    """Initialize the database with required tables and sample data"""
    conn = sqlite3.connect(app.config["DATABASE"])
    conn.row_factory = sqlite3.Row

    try:
        print("Creating database tables...")

        # Enable foreign key constraints
        conn.execute("PRAGMA foreign_keys = ON")

        # Students table (unchanged)
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS students (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                first_name TEXT NOT NULL,
                last_name TEXT NOT NULL,
                parent_first TEXT,
                parent_last TEXT,
                email TEXT UNIQUE NOT NULL,
                phone TEXT,
                membership_level TEXT NOT NULL DEFAULT 'Bronze',
                welcome_free_used BOOLEAN DEFAULT 0,
                welcome_package_date_started TIMESTAMP NULL,
                welcome_package_upgrade_date TIMESTAMP NULL,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """
        )
        print("✓ Students table created")

        # Admin users table (unchanged)
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS admin_users (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                email TEXT UNIQUE NOT NULL,
                password_hash TEXT NOT NULL,
                role TEXT NOT NULL DEFAULT 'manager',
                first_name TEXT,
                last_name TEXT,
                active BOOLEAN DEFAULT 1,
                last_login TIMESTAMP,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """
        )
        print("✓ Admin users table created")

        # Membership tiers table (unchanged)
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS membership_tiers (
                level TEXT PRIMARY KEY,
                free_notices INTEGER NOT NULL,
                deadline_hours INTEGER NOT NULL,
                deadline_display TEXT NOT NULL,
                active BOOLEAN DEFAULT 1,
                sort_order INTEGER DEFAULT 1,
                color TEXT DEFAULT '#007bff',
                description TEXT,
                welcome_message TEXT,
                policy_message TEXT,
                allow_sequential BOOLEAN DEFAULT 1,
                allow_rescheduling BOOLEAN DEFAULT 1,
                require_approval BOOLEAN DEFAULT 0,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """
        )
        print("✓ Membership tiers table created")

        # UPDATED Cancellations table with new columns
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS cancellations (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                student_id INTEGER NOT NULL,
                lesson_date DATE NOT NULL,
                lesson_time TIME NOT NULL,
                sequential_lessons TEXT,
                reschedule_requested BOOLEAN DEFAULT 0,
                reschedule_preferences TEXT,
                error_report TEXT,
                cancellation_note TEXT,                    -- NEW COLUMN
                charged BOOLEAN DEFAULT 0,
                excluded BOOLEAN DEFAULT 0,
                approved_by TEXT,
                exclusion_reason TEXT,
                manager_notes TEXT,
                status TEXT DEFAULT 'pending',
                deadline_passed BOOLEAN DEFAULT 0,        -- NEW COLUMN
                is_override BOOLEAN DEFAULT 0,            -- NEW COLUMN
                submitted_by TEXT,                        -- 'student' or 'manager'
                submission_method TEXT,                   -- 'client_portal', 'manager_portal'
                actual_submission_timestamp TIMESTAMP,    -- When submission was actually made
                manual_submission_date TIMESTAMP,         -- Policy submission date (for manager submissions)
                manager_submitted_by INTEGER,             -- User ID of manager who submitted
                is_manager_submission BOOLEAN DEFAULT 0,  -- 1 if submitted by manager, 0 if student
                suppress_notifications BOOLEAN DEFAULT 0, -- Don't send notifications if 1
                status_details TEXT,                      -- e.g., "1 of 2 used this month"
                excluded_notification_suppressed BOOLEAN DEFAULT 0,  -- Suppress exclusion notifications
                override_notification_suppressed BOOLEAN DEFAULT 0,  -- Suppress override notifications
                same_day_after_time BOOLEAN DEFAULT 0,   -- Cancelled after lesson time, same day
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (student_id) REFERENCES students (id) ON DELETE CASCADE
            )
        """
        )
        print("✓ Cancellations table created")

        # Package history table for tracking Welcome Package upgrades
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS package_history (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                student_id INTEGER NOT NULL,
                old_package TEXT,
                new_package TEXT,
                change_date TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                welcome_free_used_at_upgrade BOOLEAN DEFAULT 0,
                notes TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (student_id) REFERENCES students (id) ON DELETE CASCADE
            )
        """
        )
        print("✓ Package history table created")
        
        # Create indexes for package_history
        conn.execute(
            "CREATE INDEX IF NOT EXISTS idx_package_history_student ON package_history(student_id)"
        )
        conn.execute(
            "CREATE INDEX IF NOT EXISTS idx_package_history_date ON package_history(change_date)"
        )

        # System settings table (unchanged)
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS system_settings (
                key TEXT PRIMARY KEY,
                value TEXT NOT NULL,
                description TEXT,
                category TEXT DEFAULT 'general',
                data_type TEXT DEFAULT 'string',
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """
        )
        print("✓ System settings table created")

        # Email templates table (unchanged)
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS email_templates (
                id TEXT PRIMARY KEY,
                name TEXT NOT NULL,
                subject TEXT NOT NULL,
                body TEXT NOT NULL,
                type TEXT NOT NULL DEFAULT 'client',
                active BOOLEAN DEFAULT 1,
                auto_send BOOLEAN DEFAULT 1,
                priority TEXT DEFAULT 'normal',
                delay_minutes INTEGER DEFAULT 0,
                include_attachment BOOLEAN DEFAULT 0,
                variables_used TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """
        )
        print("✓ Email templates table created")

        # Email triggers table - NEW
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS email_triggers (
                id TEXT PRIMARY KEY,
                name TEXT NOT NULL,
                description TEXT,
                category TEXT DEFAULT 'cancellation',
                event_condition TEXT,
                active BOOLEAN DEFAULT 1,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """
        )
        print("✓ Email triggers table created")

        # Email trigger mappings table - NEW
        # IMPORTANT: Unique constraint is (trigger_id, recipient_type) only
        # This means: ONE template per trigger per recipient
        # Prevents multiple emails to same recipient from same trigger
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS email_trigger_mappings (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                trigger_id TEXT NOT NULL,
                template_id TEXT NOT NULL,
                recipient_type TEXT NOT NULL,
                enabled BOOLEAN DEFAULT 1,
                priority INTEGER DEFAULT 0,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (trigger_id) REFERENCES email_triggers(id),
                FOREIGN KEY (template_id) REFERENCES email_templates(id),
                UNIQUE(trigger_id, recipient_type)
            )
        """
        )
        print("✓ Email trigger mappings table created")
        
        # Create indexes for trigger mappings
        conn.execute(
            "CREATE INDEX IF NOT EXISTS idx_trigger_mappings_trigger ON email_trigger_mappings(trigger_id)"
        )
        conn.execute(
            "CREATE INDEX IF NOT EXISTS idx_trigger_mappings_template ON email_trigger_mappings(template_id)"
        )
        conn.execute(
            "CREATE INDEX IF NOT EXISTS idx_trigger_mappings_recipient ON email_trigger_mappings(recipient_type)"
        )

        # System logs table (unchanged)
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS system_logs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER,
                user_type TEXT DEFAULT 'unknown',
                action TEXT NOT NULL,
                details TEXT,
                ip_address TEXT,
                user_agent TEXT,
                session_id TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """
        )
        print("✓ System logs table created")

        # Create indexes for better performance (UPDATED with new columns)
        conn.execute("CREATE INDEX IF NOT EXISTS idx_students_email ON students(email)")
        conn.execute(
            "CREATE INDEX IF NOT EXISTS idx_students_membership ON students(membership_level)"
        )
        conn.execute(
            "CREATE INDEX IF NOT EXISTS idx_cancellations_student ON cancellations(student_id)"
        )
        conn.execute(
            "CREATE INDEX IF NOT EXISTS idx_cancellations_date ON cancellations(lesson_date)"
        )
        conn.execute(
            "CREATE INDEX IF NOT EXISTS idx_cancellations_created ON cancellations(created_at)"
        )
        conn.execute(
            "CREATE INDEX IF NOT EXISTS idx_cancellations_note ON cancellations(cancellation_note)"  # NEW INDEX
        )
        conn.execute(
            "CREATE INDEX IF NOT EXISTS idx_cancellations_deadline ON cancellations(deadline_passed)"  # NEW INDEX
        )
        conn.execute(
            "CREATE INDEX IF NOT EXISTS idx_cancellations_override ON cancellations(is_override)"  # NEW INDEX
        )
        conn.execute(
            "CREATE INDEX IF NOT EXISTS idx_system_logs_user ON system_logs(user_id)"
        )
        conn.execute(
            "CREATE INDEX IF NOT EXISTS idx_system_logs_action ON system_logs(action)"
        )
        print("✓ Database indexes created")

        # Insert default membership tiers (unchanged)
        membership_tiers = [
            (
                "Bronze",
                1,
                18,
                "6pm previous day",
                1,
                1,
                "#cd7f32",
                "Basic membership with standard cancellation policy",
                "Welcome to Bronze membership!",
                "Please review our standard cancellation policy.",
                1,
                1,
                0,
            ),
            (
                "Silver",
                2,
                18,
                "6pm previous day",
                1,
                2,
                "#c0c0c0",
                "Standard membership with enhanced benefits",
                "Welcome to Silver membership!",
                "You have 2 free cancellations per month.",
                1,
                1,
                0,
            ),
            (
                "Gold",
                4,
                2,
                "2 hours before lesson",
                1,
                3,
                "#ffd700",
                "Premium membership with flexible cancellation policy",
                "Welcome to Gold membership!",
                "Enjoy flexible cancellation up to 2 hours before your lesson.",
                1,
                1,
                0,
            ),
            (
                "Intro Package",
                1,
                18,
                "6pm previous day",
                1,
                4,
                "#17a2b8",
                "Introductory package for new students",
                "Welcome! We hope you enjoy your intro lessons.",
                "As an intro student, you have standard cancellation terms.",
                1,
                1,
                0,
            ),
            (
                "Legacy",
                1,
                18,
                "6pm previous day",
                0,
                5,
                "#6c757d",
                "Legacy membership tier (deprecated)",
                "Thank you for your continued loyalty.",
                "Your legacy membership terms apply.",
                1,
                1,
                0,
            ),
            (
                "Guest",
                1,
                18,
                "6pm previous day",
                1,
                6,
                "#28a745",
                "Guest lesson package",
                "Welcome as our guest!",
                "Guest cancellation policy applies.",
                1,
                1,
                0,
            ),
            (
                "Welcome Package",
                1,
                18,
                "6pm previous day",
                1,
                7,
                "#fd7e14",
                "Welcome package for first-time students",
                "Welcome to Riverside Equestrian!",
                "We are excited to have you join us.",
                1,
                1,
                0,
            ),
        ]

        for tier in membership_tiers:
            conn.execute(
                """
                INSERT OR REPLACE INTO membership_tiers 
                (level, free_notices, deadline_hours, deadline_display, active, sort_order, color, description, welcome_message, policy_message, allow_sequential, allow_rescheduling, require_approval)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
                tier,
            )
        print("✓ Default membership tiers inserted")

        # Insert default system settings (unchanged)
        system_settings = [
            (
                "system_name",
                "Riverside Equestrian Cancellation System",
                "System display name",
                "general",
                "string",
            ),
            (
                "company_email",
                "managers@riversideequestrian.ca",
                "Main company email",
                "general",
                "email",
            ),
            (
                "website_url",
                "https://www.riversideequestrian.ca",
                "Company website",
                "general",
                "url",
            ),
            (
                "policy_url",
                "https://www.riversideequestrian.ca/cancellations",
                "Cancellation policy URL",
                "general",
                "url",
            ),
            ("timezone", "America/Los_Angeles", "System timezone", "general", "string"),
            (
                "data_retention_months",
                "12",
                "How long to keep cancellation data",
                "business",
                "integer",
            ),
            (
                "max_sequential_lessons",
                "10",
                "Maximum sequential lessons per notice",
                "business",
                "integer",
            ),
            (
                "illness_documentation_days",
                "14",
                "Days to submit illness documentation",
                "business",
                "integer",
            ),
            (
                "advance_booking_days",
                "7",
                "Days before month start for cancellations",
                "business",
                "integer",
            ),
            (
                "session_timeout_minutes",
                "60",
                "User session timeout",
                "security",
                "integer",
            ),
            (
                "max_login_attempts",
                "5",
                "Maximum failed login attempts",
                "security",
                "integer",
            ),
            (
                "lockout_duration_minutes",
                "15",
                "Account lockout duration",
                "security",
                "integer",
            ),
            (
                "email_from_address",
                "noreply@riversideequestrian.ca",
                "Email from address",
                "email",
                "email",
            ),
            (
                "email_from_name",
                "Riverside Equestrian",
                "Email from name",
                "email",
                "string",
            ),
            (
                "manager_emails",
                "managers@riversideequestrian.ca, stav@riversideequestrian.ca",
                "Manager email addresses (comma-separated)",
                "email",
                "string",
            ),
            (
                "client_email_notifications",
                "true",
                "Send email notifications to clients",
                "email",
                "boolean",
            ),
            (
                "manager_email_notifications",
                "true",
                "Send email notifications to managers",
                "email",
                "boolean",
            ),
            (
                "backup_retention_days",
                "30",
                "How long to keep backup files",
                "maintenance",
                "integer",
            ),
        ]

        for setting in system_settings:
            conn.execute(
                """
                INSERT OR REPLACE INTO system_settings (key, value, description, category, data_type)
                VALUES (?, ?, ?, ?, ?)
            """,
                setting,
            )
        print("✓ Default system settings inserted")

        # Insert default email templates (unchanged)
        email_templates = [
            (
                "client_confirmation",
                "Client Confirmation",
                "Cancellation Confirmation - {{client_name}}",
                '<p>Dear {{client_name}},</p><p>This email confirms that we have received your lesson cancellation request.</p><p><strong>Cancellation Details:</strong></p><ul><li>Lesson Date: {{lesson_date}}</li><li>Lesson Time: {{lesson_time}}</li><li>Membership Level: {{membership_tier}}</li><li>Cancellation Status: {{cancellation_status}}</li></ul><p>{{status_message}}</p><p>Current cancellation usage for this month:</p><ul><li>Free cancellations used: {{used_cancellations}} of {{allowed_cancellations}}</li><li>Remaining free cancellations: {{remaining_cancellations}}</li></ul><p>If you have any questions about this cancellation or our policy, please contact us.</p><p>Best regards,<br>The Riverside Equestrian Team</p><hr><p class="small text-muted">For more information about our cancellation policy, visit: <a href="{{policy_url}}">{{policy_url}}</a></p>',
                "client",
                1,
                1,
                "normal",
                0,
                0,
                "client_name,lesson_date,lesson_time,membership_tier,cancellation_status,status_message,used_cancellations,allowed_cancellations,remaining_cancellations,policy_url",
            ),
            (
                "manager_notification",
                "Manager Notification",
                "New Cancellation - {{client_name}} - {{lesson_date}}",
                '<p>A new cancellation has been submitted:</p><p><strong>Client:</strong> {{client_name}} ({{membership_tier}})</p><p><strong>Parent:</strong> {{parent_name}}</p><p><strong>Email:</strong> {{client_email}}</p><p><strong>Phone:</strong> {{client_phone}}</p><p><strong>Lesson:</strong> {{lesson_date}} at {{lesson_time}}</p><p><strong>Status:</strong> {{cancellation_status}}</p><p><strong>Will be charged:</strong> {{will_be_charged}}</p><p><strong>Reason:</strong> {{charge_reason}}</p><p><strong>Sequential lessons:</strong> {{sequential_lessons}}</p><p><strong>Reschedule requested:</strong> {{reschedule_requested}}</p><p><strong>Preferences:</strong> {{reschedule_preferences}}</p><p><strong>Error report:</strong> {{error_report}}</p><p><strong>Submission time:</strong> {{submission_time}}</p><hr><p><a href="{{dashboard_url}}">View in Dashboard</a></p>',
                "manager",
                1,
                1,
                "normal",
                0,
                0,
                "client_name,membership_tier,parent_name,client_email,client_phone,lesson_date,lesson_time,cancellation_status,will_be_charged,charge_reason,sequential_lessons,reschedule_requested,reschedule_preferences,error_report,submission_time,dashboard_url",
            ),
            (
                "cancellation_charged",
                "Cancellation Charged",
                "Cancellation Notice - Charge Applied - {{client_name}}",
                "<p>Dear {{client_name}},</p><p>Your cancellation has been processed, and a charge has been applied to your account.</p><p><strong>Lesson Date:</strong> {{lesson_date}}</p><p><strong>Lesson Time:</strong> {{lesson_time}}</p><p><strong>Reason for charge:</strong> {{charge_reason}}</p><p><strong>Charge amount:</strong> As per your membership agreement</p><p>For questions about this charge, please contact us at {{contact_email}}.</p><p>Best regards,<br>The Riverside Equestrian Team</p>",
                "client",
                1,
                1,
                "normal",
                0,
                0,
                "client_name,lesson_date,lesson_time,charge_reason,contact_email",
            ),
            (
                "free_cancellation",
                "Free Cancellation",
                "Free Cancellation Confirmed - {{client_name}}",
                "<p>Dear {{client_name}},</p><p>Your cancellation has been processed at no charge.</p><p><strong>Lesson Date:</strong> {{lesson_date}}</p><p><strong>Lesson Time:</strong> {{lesson_time}}</p><p><strong>Remaining free cancellations this month:</strong> {{remaining_cancellations}}</p><p>Thank you for giving us advance notice.</p><p>Best regards,<br>The Riverside Equestrian Team</p>",
                "client",
                1,
                1,
                "normal",
                0,
                0,
                "client_name,lesson_date,lesson_time,remaining_cancellations",
            ),
        ]

        for template in email_templates:
            conn.execute(
                """
                INSERT OR REPLACE INTO email_templates 
                (id, name, subject, body, type, active, auto_send, priority, delay_minutes, include_attachment, variables_used)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
                template,
            )
        print("✓ Default email templates inserted")

        # Insert default email triggers - NEW
        email_triggers = [
            ("free_cancellation_triggered", "Free Cancellation", "Sent when a cancellation is approved at no charge", "cancellation", "cancellation.charged == False"),
            ("charged_cancellation_triggered", "Charged Cancellation", "Sent when a cancellation incurs a charge", "cancellation", "cancellation.charged == True"),
            ("manager_new_cancellation", "New Cancellation Notification", "Sent to managers when a new cancellation is submitted", "cancellation", "cancellation.submitted == True"),
            ("client_submission_confirmation", "Client Submission Confirmation", "Sent to client confirming receipt of cancellation", "cancellation", "cancellation.submission_confirmed == True"),
            ("manager_override_applied", "Manager Override Applied", "Sent when a manager overrides a cancellation", "override", "cancellation.override == True"),
            ("lesson_excluded_student", "Lesson Excluded - Student Notification", "Sent to student when lesson is marked as excluded", "exclusion", "cancellation.excluded == True && recipient == student"),
            ("lesson_excluded_manager", "Lesson Excluded - Manager Notification", "Sent to manager when lesson is marked as excluded", "exclusion", "cancellation.excluded == True && recipient == manager"),
            ("override_to_charged_student", "Override to Charged - Student", "Sent to student when override changes status to charged", "override", "override.new_status == charged && recipient == student"),
            ("override_to_charged_manager", "Override to Charged - Manager", "Sent to manager when override changes status to charged", "override", "override.new_status == charged && recipient == manager"),
            ("override_to_free_student", "Override to Free - Student", "Sent to student when override changes status to free", "override", "override.new_status == free && recipient == student"),
            ("override_to_free_manager", "Override to Free - Manager", "Sent to manager when override changes status to free", "override", "override.new_status == free && recipient == manager"),
            ("manager_submits_cancellation_student", "Manager Submits Cancellation - Student", "Sent to student when manager submits cancellation on their behalf", "submission", "submission.submitted_by == manager && recipient == student"),
            ("manager_submits_cancellation_manager", "Manager Submits Cancellation - Manager", "Notification to managers when another manager submits cancellation", "submission", "submission.submitted_by == manager && recipient == manager"),
        ]

        for trigger in email_triggers:
            conn.execute(
                """
                INSERT OR REPLACE INTO email_triggers 
                (id, name, description, category, event_condition)
                VALUES (?, ?, ?, ?, ?)
            """,
                trigger,
            )
        print("✓ Default email triggers inserted")

        # Insert default trigger-to-template mappings - NEW
        # CORRECTED: recipient_type is 'student' or 'manager' (not 'client')
        # IMPORTANT: Unique constraint ensures only ONE template per (trigger, recipient_type)
        trigger_mappings = [
            ("free_cancellation_triggered", "free_cancellation", "student", 1, 0),
            ("charged_cancellation_triggered", "cancellation_charged", "student", 1, 0),
            ("manager_new_cancellation", "manager_notification", "manager", 1, 0),
            ("client_submission_confirmation", "client_confirmation", "student", 1, 0),
            ("manager_override_applied", "manager_override_notification", "manager", 1, 0),
            ("lesson_excluded_student", "lesson_excluded_student", "student", 1, 0),
            ("lesson_excluded_manager", "lesson_excluded_manager", "manager", 1, 0),
            ("override_to_charged_student", "override_excluded_to_charged_student", "student", 1, 0),
            ("override_to_charged_manager", "override_excluded_to_charged_manager", "manager", 1, 0),
            ("override_to_free_student", "override_excluded_to_free_student", "student", 1, 0),
            ("override_to_free_manager", "override_excluded_to_free_manager", "manager", 1, 0),
            ("manager_submits_cancellation_student", "manager_submits_cancellation_student", "student", 1, 0),
            ("manager_submits_cancellation_manager", "manager_submits_cancellation_manager", "manager", 1, 0),
        ]

        for mapping in trigger_mappings:
            conn.execute(
                """
                INSERT OR REPLACE INTO email_trigger_mappings 
                (trigger_id, template_id, recipient_type, enabled, priority)
                VALUES (?, ?, ?, ?, ?)
            """,
                mapping,
            )
        print("✓ Default email trigger mappings inserted")
        from werkzeug.security import generate_password_hash

        admin_password_hash = generate_password_hash("admin123")

        admin_users = [
            (
                "admin@riversideequestrian.ca",
                admin_password_hash,
                "senior_manager",
                "Senior",
                "Admin",
                1,
            ),
            (
                "manager@riversideequestrian.ca",
                admin_password_hash,
                "manager",
                "Manager",
                "User",
                1,
            ),
        ]

        for user in admin_users:
            conn.execute(
                """
                INSERT OR REPLACE INTO admin_users (email, password_hash, role, first_name, last_name, active)
                VALUES (?, ?, ?, ?, ?, ?)
            """,
                user,
            )
        print("✓ Sample admin users inserted")

        # Insert sample students (unchanged)
        sample_students = [
            (
                "Chloe",
                "Chow",
                "Josephine",
                "Tang",
                "chloechow2016@gmail.com",
                "604-505-3333",
                "Silver",
            ),
            (
                "Emma",
                "Johnson",
                "Sarah",
                "Johnson",
                "emma.johnson@email.com",
                "604-123-4567",
                "Gold",
            ),
            (
                "Alex",
                "Smith",
                "Mike",
                "Smith",
                "alex.smith@email.com",
                "604-987-6543",
                "Bronze",
            ),
            (
                "Sofia",
                "Garcia",
                "Maria",
                "Garcia",
                "sofia.garcia@email.com",
                "604-555-0123",
                "Silver",
            ),
            (
                "Liam",
                "Brown",
                "Jennifer",
                "Brown",
                "liam.brown@email.com",
                "604-111-2222",
                "Bronze",
            ),
            (
                "Olivia",
                "Davis",
                "Robert",
                "Davis",
                "olivia.davis@email.com",
                "604-333-4444",
                "Gold",
            ),
            (
                "Noah",
                "Wilson",
                "Lisa",
                "Wilson",
                "noah.wilson@email.com",
                "604-555-6666",
                "Silver",
            ),
            (
                "Ava",
                "Martinez",
                "Carlos",
                "Martinez",
                "ava.martinez@email.com",
                "604-777-8888",
                "Bronze",
            ),
            (
                "Ethan",
                "Anderson",
                "Nicole",
                "Anderson",
                "ethan.anderson@email.com",
                "604-999-0000",
                "Intro Package",
            ),
            (
                "Isabella",
                "Taylor",
                "David",
                "Taylor",
                "isabella.taylor@email.com",
                "604-222-3333",
                "Welcome Package",
            ),
            (
                "Mason",
                "Moore",
                "Amanda",
                "Moore",
                "mason.moore@email.com",
                "604-444-5555",
                "Bronze",
            ),
            (
                "Sophia",
                "Jackson",
                "Kevin",
                "Jackson",
                "sophia.jackson@email.com",
                "604-666-7777",
                "Silver",
            ),
            (
                "Lucas",
                "White",
                "Rachel",
                "White",
                "lucas.white@email.com",
                "604-888-9999",
                "Gold",
            ),
            (
                "Mia",
                "Harris",
                "Jason",
                "Harris",
                "mia.harris@email.com",
                "604-111-0000",
                "Guest",
            ),
            (
                "Aiden",
                "Clark",
                "Melissa",
                "Clark",
                "aiden.clark@email.com",
                "604-222-1111",
                "Legacy",
            ),
        ]

        for student in sample_students:
            conn.execute(
                """
                INSERT OR REPLACE INTO students (first_name, last_name, parent_first, parent_last, email, phone, membership_level)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            """,
                student,
            )
        print("✓ Sample students inserted")

        # UPDATED sample cancellations with new columns
        sample_cancellations = [
            (
                1,
                "2024-08-20",
                "14:00:00",
                "Student was sick",
                0,
                "approved",
                0,
                0,
                "2024-08-15 10:30:00",
            ),
            (
                2,
                "2024-08-22",
                "16:00:00",
                "Need to reschedule",
                1,
                "charged",
                1,
                0,
                "2024-08-16 09:15:00",
            ),
            (
                3,
                "2024-08-25",
                "15:30:00",
                None,
                0,
                "approved",
                0,
                0,
                "2024-08-17 11:45:00",
            ),
            (
                1,
                "2024-08-28",
                "14:00:00",
                "Family emergency",
                0,
                "approved",
                0,
                1,
                "2024-08-18 08:20:00",
            ),
            (
                4,
                "2024-08-30",
                "13:00:00",
                None,
                1,
                "charged",
                1,
                0,
                "2024-08-19 14:10:00",
            ),
            (
                2,
                "2024-09-02",
                "16:00:00",
                "Weather concerns",
                0,
                "pending",
                0,
                0,
                "2024-08-20 15:30:00",
            ),
            (
                5,
                "2024-09-05",
                "11:00:00",
                None,
                0,
                "approved",
                0,
                0,
                "2024-08-21 12:00:00",
            ),
            (
                6,
                "2024-09-08",
                "17:00:00",
                "Schedule conflict",
                1,
                "charged",
                1,
                0,
                "2024-08-22 18:45:00",
            ),
            (
                7,
                "2024-09-10",
                "10:00:00",
                "Car trouble",
                0,
                "pending",
                0,
                0,
                "2024-08-23 09:30:00",
            ),
            (
                8,
                "2024-09-12",
                "18:00:00",
                None,
                0,
                "approved",
                0,
                0,
                "2024-08-24 16:20:00",
            ),
            (
                3,
                "2024-09-15",
                "15:30:00",
                "Work commitment",
                1,
                "charged",
                1,
                0,
                "2024-08-25 12:15:00",
            ),
            (
                9,
                "2024-09-18",
                "09:00:00",
                "Doctor appointment",
                0,
                "approved",
                0,
                1,
                "2024-08-26 14:45:00",
            ),
            (
                10,
                "2024-09-20",
                "13:30:00",
                None,
                0,
                "pending",
                0,
                0,
                "2024-08-27 11:30:00",
            ),
        ]

        for cancellation in sample_cancellations:
            conn.execute(
                """
                INSERT OR REPLACE INTO cancellations (student_id, lesson_date, lesson_time, cancellation_note, charged, status, deadline_passed, is_override, created_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
                cancellation,
            )
        print("✓ Sample cancellations inserted")

        # Insert sample system logs (unchanged)
        sample_logs = [
            (
                1,
                "admin",
                "login",
                "Senior manager login",
                "127.0.0.1",
                "Mozilla/5.0...",
                "session123",
                "2024-08-15 08:00:00",
            ),
            (
                2,
                "admin",
                "login",
                "Manager login",
                "127.0.0.1",
                "Mozilla/5.0...",
                "session124",
                "2024-08-15 08:30:00",
            ),
            (
                1,
                "client",
                "login",
                "Client login: chloechow2016@gmail.com",
                "127.0.0.1",
                "Mozilla/5.0...",
                "session125",
                "2024-08-15 10:00:00",
            ),
            (
                1,
                "client",
                "cancellation_submitted",
                "Lesson: 2024-08-20 14:00:00, Charged: False",
                "127.0.0.1",
                "Mozilla/5.0...",
                "session125",
                "2024-08-15 10:30:00",
            ),
            (
                1,
                "admin",
                "student_updated",
                "Student ID: 5",
                "127.0.0.1",
                "Mozilla/5.0...",
                "session123",
                "2024-08-16 09:00:00",
            ),
            (
                2,
                "client",
                "login",
                "Client login: emma.johnson@email.com",
                "127.0.0.1",
                "Mozilla/5.0...",
                "session126",
                "2024-08-16 09:00:00",
            ),
            (
                2,
                "client",
                "cancellation_submitted",
                "Lesson: 2024-08-22 16:00:00, Charged: True",
                "127.0.0.1",
                "Mozilla/5.0...",
                "session126",
                "2024-08-16 09:15:00",
            ),
        ]

        for log in sample_logs:
            conn.execute(
                """
                INSERT OR REPLACE INTO system_logs (user_id, user_type, action, details, ip_address, user_agent, session_id, created_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """,
                log,
            )
        print("✓ Sample system logs inserted")

        # Commit all changes
        conn.commit()
        print("✓ Database initialization completed successfully!")

        # Verify table creation
        cursor = conn.execute("SELECT name FROM sqlite_master WHERE type='table'")
        tables = [row[0] for row in cursor.fetchall()]
        print(f"✓ Created {len(tables)} tables: {', '.join(tables)}")

        # Verify data insertion
        cursor = conn.execute("SELECT COUNT(*) FROM students")
        student_count = cursor.fetchone()[0]
        print(f"✓ Inserted {student_count} students")

        cursor = conn.execute("SELECT COUNT(*) FROM cancellations")
        cancellation_count = cursor.fetchone()[0]
        print(f"✓ Inserted {cancellation_count} cancellations")

        cursor = conn.execute("SELECT COUNT(*) FROM membership_tiers")
        tier_count = cursor.fetchone()[0]
        print(f"✓ Inserted {tier_count} membership tiers")

        return True

    except Exception as e:
        print(f"❌ Error initializing database: {str(e)}")
        conn.rollback()
        return False
    finally:
        conn.close()


def reset_database():
    """Delete database file and recreate - Safest approach"""
    try:
        db_path = app.config["DATABASE"]

        # Close any existing connections first
        try:
            conn = sqlite3.connect(db_path)
            conn.close()
        except:
            pass

        # Remove the database file if it exists
        if os.path.exists(db_path):
            try:
                os.remove(db_path)
                print(f"Deleted existing database file: {db_path}")
            except OSError as e:
                print(f"Could not delete database file: {e}")
                # If we can't delete the file, try the table-by-table approach
                return reset_database_tables()

        # Create fresh database
        print("Creating fresh database...")
        return init_db()

    except Exception as e:
        print(f"Error in reset_database: {str(e)}")
        import traceback

        traceback.print_exc()
        return False


def reset_database_tables():
    """Fallback method: Drop all tables and recreate database"""
    try:
        conn = sqlite3.connect(app.config["DATABASE"])

        # Get all table names (excluding system tables)
        cursor = conn.execute(
            """SELECT name FROM sqlite_master 
               WHERE type='table' 
               AND name NOT LIKE 'sqlite_%'"""
        )
        tables = [row[0] for row in cursor.fetchall()]

        # Disable foreign key constraints temporarily
        conn.execute("PRAGMA foreign_keys = OFF")

        # Drop all user tables
        for table in tables:
            try:
                conn.execute(f"DROP TABLE IF EXISTS `{table}`")
                print(f"Dropped table: {table}")
            except Exception as e:
                print(f"Warning: Could not drop table {table}: {e}")

        # Drop user-created indexes
        cursor = conn.execute(
            """SELECT name FROM sqlite_master 
               WHERE type='index' 
               AND name NOT LIKE 'sqlite_%'
               AND sql IS NOT NULL"""
        )
        indexes = [row[0] for row in cursor.fetchall()]

        for index in indexes:
            try:
                conn.execute(f"DROP INDEX IF EXISTS `{index}`")
                print(f"Dropped index: {index}")
            except Exception as e:
                print(f"Warning: Could not drop index {index}: {e}")

        # Reset auto-increment sequences
        try:
            conn.execute("DELETE FROM sqlite_sequence")
            print("Reset auto-increment sequences")
        except:
            pass

        # Re-enable foreign key constraints
        conn.execute("PRAGMA foreign_keys = ON")

        conn.commit()
        conn.close()

        # Reinitialize with new schema
        print("Reinitializing database with new schema...")
        return init_db()

    except Exception as e:
        print(f"Error in reset_database_tables: {str(e)}")
        import traceback

        traceback.print_exc()
        return False


def verify_database():
    """Verify database structure and data"""
    try:
        conn = sqlite3.connect(app.config["DATABASE"])
        conn.row_factory = sqlite3.Row

        # Check if all required tables exist
        required_tables = [
            "students",
            "admin_users",
            "membership_tiers",
            "cancellations",
            "system_settings",
            "email_templates",
            "system_logs",
        ]

        cursor = conn.execute("SELECT name FROM sqlite_master WHERE type='table'")
        existing_tables = [row[0] for row in cursor.fetchall()]

        missing_tables = [
            table for table in required_tables if table not in existing_tables
        ]

        if missing_tables:
            print(f"Missing tables: {missing_tables}")
            return False

        # Check if tables have data
        for table in required_tables:
            cursor = conn.execute(f"SELECT COUNT(*) FROM {table}")
            count = cursor.fetchone()[0]
            print(f"Table {table}: {count} records")

        return True

    except Exception as e:
        print(f"Error verifying database: {str(e)}")
        return False
    finally:
        conn.close()


def get_db():
    """Get database connection"""
    conn = sqlite3.connect(app.config["DATABASE"])
    conn.row_factory = sqlite3.Row
    return conn


# ===================================
# UTILITY FUNCTIONS
# ===================================


def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if "user_id" not in session:
            return redirect(url_for("login"))
        return f(*args, **kwargs)

    return decorated_function


def admin_required(f):
    """Decorator to require admin access"""

    @wraps(f)
    def decorated_function(*args, **kwargs):
        if "user_role" not in session or session["user_role"] not in [
            "manager",
            "senior_manager",
        ]:
            flash("Admin access required", "error")
            return redirect(url_for("dashboard"))
        return f(*args, **kwargs)

    return decorated_function


def senior_admin_required(f):
    """Decorator to require senior admin access"""

    @wraps(f)
    def decorated_function(*args, **kwargs):
        if "user_role" not in session or session["user_role"] != "senior_manager":
            flash("Senior manager access required", "error")
            return redirect(url_for("dashboard"))
        return f(*args, **kwargs)

    return decorated_function


def log_action(action, details=None):
    """Log user action"""
    try:
        conn = get_db()
        conn.execute(
            "INSERT INTO system_logs (user_id, action, details, ip_address) VALUES (?, ?, ?, ?)",
            (session.get("user_id"), action, details, request.remote_addr),
        )
        conn.commit()
        conn.close()
    except:
        pass  # Don't fail if logging fails


def get_app_timezone():
    """Get the application timezone - always use Pacific (PST/PDT)"""
    return pytz.timezone("America/Los_Angeles")


def toronto_now():
    """Get current datetime in Pacific timezone (PST/PDT) - simple and direct"""
    pacific_tz = pytz.timezone("America/Los_Angeles")
    utc_now = datetime.utcnow()
    utc_dt = pytz.UTC.localize(utc_now)
    return utc_dt.astimezone(pacific_tz)


def now_in_app_timezone():
    """Get current datetime in Pacific timezone (PST/PDT)"""
    return toronto_now()


def localize_datetime(dt, from_tz=None):
    """
    Convert a datetime to Pacific timezone (PST/PDT).

    Args:
        dt: datetime object, string, or None
        from_tz: Optional source timezone (if None and dt is naive, assumes Pacific)

    Returns:
        Timezone-aware datetime in Pacific timezone, or None
    """
    pacific_tz = pytz.timezone("America/Los_Angeles")

    if dt is None:
        return None

    # If it's a string, parse it first
    if isinstance(dt, str):
        try:
            if "T" in dt:
                # ISO format with possible timezone
                dt = datetime.fromisoformat(dt.replace("Z", "+00:00"))
            else:
                # Assume "%Y-%m-%d %H:%M:%S" format from database (Pacific time)
                dt = datetime.strptime(dt, "%Y-%m-%d %H:%M:%S")
                # Explicitly localize to Pacific since database stores Pacific times
                return pacific_tz.localize(dt)
        except (ValueError, TypeError):
            # If parsing fails, return None instead of the string
            return None

    # If datetime is naive, assume it's in Pacific timezone (our database default)
    if dt.tzinfo is None:
        dt = pacific_tz.localize(dt)
        return dt

    # Convert to Pacific timezone if already timezone-aware
    return dt.astimezone(pacific_tz)


def format_datetime_for_display(dt):
    """Format datetime for display in Pacific timezone (PST/PDT)"""
    if dt is None:
        return "Unknown"

    pacific_tz = pytz.timezone("America/Los_Angeles")

    # If it's a string, parse it first
    if isinstance(dt, str):
        try:
            if "T" in dt:
                dt = datetime.fromisoformat(dt.replace("Z", "+00:00"))
            else:
                dt = datetime.strptime(dt, "%Y-%m-%d %H:%M:%S")
                # Assume parsed datetime from database is in Pacific timezone
                dt = pacific_tz.localize(dt)
        except (ValueError, TypeError):
            return str(dt)

    # If it's a naive datetime (from database), assume it's Pacific time
    if isinstance(dt, datetime) and dt.tzinfo is None:
        dt = pacific_tz.localize(dt)

    # Convert to Pacific timezone if needed
    if hasattr(dt, "tzinfo") and dt.tzinfo is not None:
        dt = dt.astimezone(pacific_tz)

    if isinstance(dt, datetime):
        return dt.strftime("%B %d, %Y at %I:%M %p %Z")

    return str(dt)


def format_date_for_display(dt):
    """Format date for display"""
    if dt is None:
        return "Unknown"

    # Convert to app timezone if it's a datetime
    if isinstance(dt, datetime):
        local_dt = localize_datetime(dt)
        return local_dt.strftime("%B %d, %Y")
    elif isinstance(dt, date):
        return dt.strftime("%B %d, %Y")
    elif isinstance(dt, str):
        try:
            if "T" in dt:
                dt_obj = datetime.fromisoformat(dt.replace("Z", "+00:00"))
            else:
                dt_obj = datetime.strptime(dt, "%Y-%m-%d")
            return format_date_for_display(dt_obj)
        except (ValueError, TypeError):
            return str(dt)

    return str(dt)


def get_student_by_email(email):
    """Get student information by email"""
    conn = get_db()
    student = conn.execute(
        "SELECT * FROM students WHERE email = ?", (email,)
    ).fetchone()
    conn.close()
    return student


def get_membership_tier(level):
    """Get membership tier information"""
    conn = get_db()
    tier = conn.execute(
        "SELECT * FROM membership_tiers WHERE level = ?", (level,)
    ).fetchone()
    conn.close()
    return tier


def get_monthly_cancellation_count(student_id, month=None, year=None, exclude_cancellation_id=None):
    """Get cancellation count for a student in a specific month (based on lesson date)
    
    Returns count of free (non-charged) cancellations for the month.
    Counts based on the lesson_date month, not submission date.
    This ensures quota resets based on lesson month.
    
    Args:
        student_id: The student to count for
        month: Optional month (defaults to current Pacific month)
        year: Optional year (defaults to current Pacific year)
        exclude_cancellation_id: If provided, this cancellation ID is excluded from the
            count. Use this when re-evaluating a specific cancellation during processing,
            so the cancellation being processed doesn't count itself toward its own limit.
    """
    current_time = toronto_now()
    if not month:
        month = current_time.month
    if not year:
        year = current_time.year

    conn = get_db()
    # Count free cancellations based on LESSON DATE month, not submission date
    query = """
        SELECT COUNT(*) as count FROM cancellations 
        WHERE student_id = ? 
        AND strftime('%m', lesson_date) = ? 
        AND strftime('%Y', lesson_date) = ?
        AND excluded = 0
        AND charged = 0
    """
    params = [student_id, f"{month:02d}", str(year)]

    if exclude_cancellation_id is not None:
        query += " AND id != ?"
        params.append(exclude_cancellation_id)

    count = conn.execute(query, params).fetchone()
    conn.close()
    return count["count"] if count else 0


def has_used_welcome_free(student_id):
    """Check if student has used their lifetime Welcome Package free cancellation"""
    conn = get_db()
    result = conn.execute(
        "SELECT welcome_free_used FROM students WHERE id = ?",
        (student_id,)
    ).fetchone()
    conn.close()
    return bool(result["welcome_free_used"]) if result else False


def mark_welcome_free_used(student_id):
    """Mark that a student has used their Welcome Package free cancellation"""
    conn = get_db()
    conn.execute(
        "UPDATE students SET welcome_free_used = 1 WHERE id = ?",
        (student_id,)
    )
    conn.commit()
    conn.close()


def track_package_upgrade(student_id, old_package, new_package, welcome_free_used_at_upgrade, conn=None):
    """Record when a student upgrades their package"""
    close_conn = False
    if conn is None:
        conn = get_db()
        close_conn = True
    
    try:
        current_time = toronto_now().strftime("%Y-%m-%d %H:%M:%S")
        
        conn.execute(
            """
            INSERT INTO package_history 
            (student_id, old_package, new_package, change_date, welcome_free_used_at_upgrade)
            VALUES (?, ?, ?, ?, ?)
            """,
            (student_id, old_package, new_package, current_time, int(welcome_free_used_at_upgrade))
        )
        
        # Only commit if we opened the connection
        if close_conn:
            conn.commit()
    finally:
        if close_conn and conn:
            conn.close()


def record_welcome_package_start(student_id, conn=None):
    """Record when a student starts on Welcome Package"""
    close_conn = False
    if conn is None:
        conn = get_db()
        close_conn = True
    
    try:
        current_time = toronto_now().strftime("%Y-%m-%d %H:%M:%S")
        
        conn.execute(
            """
            UPDATE students 
            SET welcome_package_date_started = ?
            WHERE id = ? AND welcome_package_date_started IS NULL
            """,
            (current_time, student_id)
        )
        
        # Only commit if we opened the connection
        if close_conn:
            conn.commit()
    finally:
        if close_conn and conn:
            conn.close()


def get_welcome_upgrade_month_carryover(student_id):
    """
    Get how many free cancellations from Welcome Package should count toward 
    the new package's monthly limit when upgrading mid-month.
    
    Returns the count of free cancellations in the month of upgrade that were 
    granted under the Welcome Package.
    Uses lesson_date to determine the month.
    """
    conn = get_db()
    current_time = toronto_now()
    
    # Get the most recent package upgrade from Welcome Package
    upgrade_record = conn.execute(
        """
        SELECT change_date, welcome_free_used_at_upgrade
        FROM package_history
        WHERE student_id = ? AND old_package = 'Welcome Package'
        ORDER BY change_date DESC
        LIMIT 1
        """,
        (student_id,)
    ).fetchone()
    
    if not upgrade_record:
        conn.close()
        return 0
    
    upgrade_time = upgrade_record["change_date"]
    
    # Parse the upgrade date
    try:
        if isinstance(upgrade_time, str):
            upgrade_datetime = datetime.strptime(upgrade_time, "%Y-%m-%d %H:%M:%S")
        else:
            upgrade_datetime = upgrade_time
    except:
        conn.close()
        return 0
    
    # Check if upgrade happened this month and if free was used
    if (upgrade_datetime.month == current_time.month and 
        upgrade_datetime.year == current_time.year and 
        upgrade_record["welcome_free_used_at_upgrade"]):
        
        # Count free cancellations in this month while on Welcome Package
        # Use lesson_date to determine the month
        count = conn.execute(
            """
            SELECT COUNT(*) as count FROM cancellations
            WHERE student_id = ?
            AND DATE(created_at) <= DATE(?)
            AND charged = 0
            AND excluded = 0
            AND strftime('%m', lesson_date) = ?
            AND strftime('%Y', lesson_date) = ?
            """,
            (student_id, upgrade_time, f"{current_time.month:02d}", str(current_time.year))
        ).fetchone()
        
        conn.close()
        return count["count"] if count else 0
    
    conn.close()
    return 0


def get_free_cancellation_count_with_welcome_logic(student_id, month=None, year=None):
    """
    Get free cancellation count for the month, accounting for Welcome Package logic.
    Counts are based on lesson_date month to ensure quota resets by lesson month.
    
    - If student is on Welcome Package: only count if they haven't used their 1 lifetime free
    - If student upgraded from Welcome Package this month: adjust the count to include carryover
    """
    conn = get_db()
    current_time = toronto_now()
    
    if not month:
        month = current_time.month
    if not year:
        year = current_time.year
    
    student = conn.execute(
        "SELECT membership_level, welcome_free_used FROM students WHERE id = ?",
        (student_id,)
    ).fetchone()
    
    if not student:
        conn.close()
        return 0
    
    # If still on Welcome Package, return 1 if not used, 0 if used
    if student["membership_level"] == "Welcome Package":
        conn.close()
        return 0 if student["welcome_free_used"] else 1
    
    # For other packages, get normal monthly count based on lesson_date
    count = conn.execute(
        """
        SELECT COUNT(*) as count FROM cancellations 
        WHERE student_id = ? 
        AND strftime('%m', lesson_date) = ? 
        AND strftime('%Y', lesson_date) = ?
        AND excluded = 0
        AND charged = 0
    """,
        (student_id, f"{month:02d}", str(year)),
    ).fetchone()
    
    conn.close()
    return count["count"] if count else 0


def calculate_cancellation_status(student):
    """Calculate current cancellation status for a student"""
    tier = get_membership_tier(student["membership_level"])
    used = get_monthly_cancellation_count(student["id"])

    return {
        "limit": tier["free_notices"],
        "used": used,
        "remaining": max(0, tier["free_notices"] - used),
    }


def calculate_deadline_datetime(tier, lesson_datetime):
    """
    Calculate the actual deadline datetime for a cancellation based on tier policy.

    Supports two policy types based on the tier's `deadline_display` text:
      1. "X hours before lesson" (e.g. Gold "2 hours before lesson"):
         deadline = lesson_datetime - deadline_hours
      2. "<time> previous day" (e.g. Bronze/Silver "6pm previous day"):
         deadline = absolute time on the day BEFORE the lesson date.

    Falls back to the relative "deadline_hours before lesson" interpretation
    when no "previous day" wording is detected, preserving Gold-tier behavior.

    Args:
        tier: Membership tier dict (must have `deadline_display` and `deadline_hours`)
        lesson_datetime: Timezone-aware lesson datetime (Pacific) or naive (assumed Pacific)

    Returns:
        Timezone-aware datetime in Pacific timezone representing the deadline.
    """
    pacific_tz = pytz.timezone("America/Los_Angeles")

    # Ensure lesson_datetime is timezone-aware
    if lesson_datetime.tzinfo is None:
        lesson_datetime = pacific_tz.localize(lesson_datetime)

    # Convert tier to dict if needed
    if hasattr(tier, "keys"):
        tier_dict = dict(tier)
    else:
        tier_dict = tier or {}

    deadline_display = (tier_dict.get("deadline_display") or "").strip().lower()
    deadline_hours = tier_dict.get("deadline_hours", 18) or 18

    # Detect "<time> previous day" pattern, e.g. "6pm previous day", "5:30pm previous day"
    prev_day_match = re.search(
        r"(\d{1,2})(?::(\d{2}))?\s*(am|pm)?\s*previous\s*day",
        deadline_display,
    )

    if prev_day_match:
        hour = int(prev_day_match.group(1))
        minute = int(prev_day_match.group(2)) if prev_day_match.group(2) else 0
        ampm = prev_day_match.group(3)

        # Convert to 24-hour format if am/pm was provided
        if ampm == "pm" and hour != 12:
            hour += 12
        elif ampm == "am" and hour == 12:
            hour = 0
        # If no am/pm, assume the value is already 24-hour

        # Build deadline: previous day at the specified clock time, Pacific tz
        prev_day = lesson_datetime.date() - timedelta(days=1)
        naive_deadline = datetime.combine(prev_day, time(hour, minute))
        return pacific_tz.localize(naive_deadline)

    # Fallback: relative "X hours before lesson"
    return lesson_datetime - timedelta(hours=deadline_hours)


def is_after_deadline(tier, lesson_datetime, submission_datetime):
    """
    Return True iff `submission_datetime` is strictly after the tier's deadline
    for the given lesson (i.e., the submission was LATE).

    Submitting exactly at the deadline is considered on time.
    """
    pacific_tz = pytz.timezone("America/Los_Angeles")
    if submission_datetime is None:
        return False
    if submission_datetime.tzinfo is None:
        submission_datetime = pacific_tz.localize(submission_datetime)
    deadline_dt = calculate_deadline_datetime(tier, lesson_datetime)
    return submission_datetime > deadline_dt


def will_be_charged(student, lesson_datetime, submission_datetime=None, exclude_cancellation_id=None):
    """
    Check if a cancellation will be charged based on:
    1. Same-day after-time (always charged)
    2. Welcome Package status (1 lifetime free if not used)
    3. Deadline (per-tier policy: relative hours OR "<time> previous day")
    4. Monthly limit (free_notices per month, based on LESSON month)
    5. Welcome Package upgrade carryover (if upgraded this month, use both old and new allowances)
    
    Args:
        student: Student record dictionary
        lesson_datetime: Lesson datetime (timezone-aware, Pacific)
        submission_datetime: When the cancellation was submitted (timezone-aware, Pacific)
                            If None, uses current time
        exclude_cancellation_id: If provided, this cancellation ID is excluded from any
            policy counts (monthly limit, Welcome Package lifetime check). Use when
            re-evaluating a specific cancellation during manager processing so the
            cancellation being processed doesn't count itself.
    
    Returns:
        tuple: (will_charge: bool, reason: str)
    """
    pacific_tz = pytz.timezone("America/Los_Angeles")
    
    # Get submission time (default to now if not provided)
    if submission_datetime is None:
        submission_datetime = toronto_now()
    
    # Ensure datetimes are timezone-aware
    if lesson_datetime.tzinfo is None:
        lesson_datetime = pacific_tz.localize(lesson_datetime)
    if submission_datetime.tzinfo is None:
        submission_datetime = pacific_tz.localize(submission_datetime)
    
    # ===== SAME-DAY AFTER-TIME CHECK =====
    # If lesson time has passed on the same day, always charge
    if submission_datetime.date() == lesson_datetime.date() and submission_datetime >= lesson_datetime:
        return True, "Cancelled on same day after lesson time (same-day surcharge applies)"
    
    # Get membership tier
    tier = get_membership_tier(student["membership_level"])
    if not tier:
        return False, "Student membership not found"
    
    # ===== WELCOME PACKAGE SPECIAL LOGIC =====
    # Welcome Package has a lifetime limit of 1 free cancellation.
    # We determine "already used" by dynamically counting prior non-charged cancellations
    # for this student (excluding the one being processed, if any). This is more reliable
    # than the `welcome_free_used` flag alone, which can become stale when a pending
    # cancellation has been recorded but not yet manager-approved.
    if student["membership_level"] == "Welcome Package":
        conn = get_db()
        other_free_query = """
            SELECT COUNT(*) as cnt FROM cancellations
            WHERE student_id = ? AND charged = 0 AND excluded = 0
        """
        other_free_params = [student["id"]]
        if exclude_cancellation_id is not None:
            other_free_query += " AND id != ?"
            other_free_params.append(exclude_cancellation_id)

        other_free_row = conn.execute(other_free_query, other_free_params).fetchone()
        conn.close()
        other_free_count = other_free_row["cnt"] if other_free_row else 0

        if other_free_count >= 1:
            # Some other (or earlier) free cancellation has already used the lifetime free
            return True, "Welcome Package free cancellation already used (lifetime limit)"
        else:
            # This is the first/only free cancellation
            return False, "Welcome Package - first free cancellation"
    
    # ===== REGULAR MEMBERSHIP LOGIC =====
    
    # 1. CHECK DEADLINE
    # Use tier-aware deadline: supports both "X hours before lesson" (e.g. Gold)
    # and absolute "<time> previous day" deadlines (e.g. Bronze "6pm previous day").
    # The earlier behavior of subtracting `deadline_hours` from lesson time was
    # incorrect for tiers like Bronze/Silver, which use an absolute clock time
    # on the prior day rather than a fixed offset.
    if is_after_deadline(tier, lesson_datetime, submission_datetime):
        # Submitted after deadline → CHARGE
        return True, f"Notice submitted after deadline"
    
    # 2. CHECK MONTHLY LIMIT (based on lesson month, not submission month)
    lesson_month = lesson_datetime.month
    lesson_year = lesson_datetime.year
    
    # Count free cancellations used this month (for the lesson month),
    # excluding the cancellation currently being processed so it doesn't count itself.
    monthly_count = get_monthly_cancellation_count(
        student["id"], lesson_month, lesson_year,
        exclude_cancellation_id=exclude_cancellation_id,
    )
    free_notices = tier["free_notices"]
    
    # 3. HANDLE WELCOME PACKAGE UPGRADE CARRYOVER
    # If student upgraded from Welcome Package in the same month as the lesson
    # and used their Welcome Package free, they get an extra free cancellation
    welcome_carryover = 0
    
    conn = get_db()
    # Check if student upgraded from Welcome Package in the LESSON month
    upgrade_record = conn.execute(
        """
        SELECT change_date, welcome_free_used_at_upgrade
        FROM package_history
        WHERE student_id = ? AND old_package = 'Welcome Package'
        AND strftime('%m', change_date) = ?
        AND strftime('%Y', change_date) = ?
        ORDER BY change_date DESC
        LIMIT 1
        """,
        (student["id"], f"{lesson_month:02d}", str(lesson_year))
    ).fetchone()
    conn.close()
    
    if upgrade_record and upgrade_record["welcome_free_used_at_upgrade"]:
        # Student used their Welcome Package free before upgrading in the lesson month
        # They get 1 carryover free from the Welcome Package
        welcome_carryover = 1
    
    # Adjust the limit to account for Welcome Package carryover
    adjusted_free_limit = free_notices + welcome_carryover
    
    if monthly_count >= adjusted_free_limit:
        # Exceeded monthly limit → CHARGE
        return True, f"Monthly free cancellation limit exceeded ({monthly_count}/{adjusted_free_limit} used)"
    
    # 4. ALL CHECKS PASSED → FREE
    return False, f"Free cancellation ({monthly_count + 1}/{adjusted_free_limit} used this month)"


def parse_lesson_datetime(lesson_date_str, lesson_time_str):
    """
    Parse lesson date and time strings into a timezone-aware datetime object in Pacific timezone.
    Handles both HH:MM and HH:MM:SS time formats.
    Returns timezone-aware datetime in America/Los_Angeles timezone.
    """
    pacific_tz = pytz.timezone("America/Los_Angeles")

    try:
        # Parse date
        lesson_date = datetime.strptime(str(lesson_date_str), "%Y-%m-%d").date()
    except (ValueError, TypeError):
        # Fallback to today's date in Pacific timezone if parsing fails
        lesson_date = toronto_now().date()

    try:
        # Parse time - handle both HH:MM and HH:MM:SS formats
        time_str = str(lesson_time_str)
        if len(time_str.split(":")) == 2:  # HH:MM format
            lesson_time = datetime.strptime(time_str, "%H:%M").time()
        elif len(time_str.split(":")) == 3:  # HH:MM:SS format
            lesson_time = datetime.strptime(time_str, "%H:%M:%S").time()
        else:
            # Default to noon if format is unexpected
            lesson_time = time(12, 0)
    except (ValueError, TypeError):
        # Default to noon if parsing fails
        lesson_time = time(12, 0)

    # Combine date and time, then localize to Pacific timezone
    naive_dt = datetime.combine(lesson_date, lesson_time)
    return pacific_tz.localize(naive_dt)


def get_dashboard_stats():
    """Calculate dashboard statistics with new 5-box metrics and debugging"""
    conn = get_db()

    # Get current Pacific time (timezone-aware)
    pacific_now = toronto_now()
    today = pacific_now.date()
    yesterday = today - timedelta(days=1)
    current_month = pacific_now.strftime("%Y-%m")

    try:
        print(f"DEBUG: Today is {today}, current month is {current_month}")

        # Today's cancellations
        today_cancellations_result = conn.execute(
            "SELECT COUNT(*) as count FROM cancellations WHERE DATE(created_at) = ?",
            (today.strftime("%Y-%m-%d"),),
        ).fetchone()
        today_cancellations = (
            today_cancellations_result["count"] if today_cancellations_result else 0
        )
        print(f"DEBUG: Today's cancellations: {today_cancellations}")

        # Yesterday's cancellations for comparison
        yesterday_cancellations_result = conn.execute(
            "SELECT COUNT(*) as count FROM cancellations WHERE DATE(created_at) = ?",
            (yesterday.strftime("%Y-%m-%d"),),
        ).fetchone()
        yesterday_cancellations = (
            yesterday_cancellations_result["count"]
            if yesterday_cancellations_result
            else 0
        )
        print(f"DEBUG: Yesterday's cancellations: {yesterday_cancellations}")

        # Calculate today vs yesterday
        if yesterday_cancellations > 0:
            change = (
                (today_cancellations - yesterday_cancellations)
                / yesterday_cancellations
            ) * 100
            if change > 0:
                today_vs_yesterday = f"+{change:.0f}% from yesterday"
            elif change < 0:
                today_vs_yesterday = f"{change:.0f}% from yesterday"
            else:
                today_vs_yesterday = "Same as yesterday"
        else:
            if today_cancellations > 0:
                today_vs_yesterday = "New activity"
            else:
                today_vs_yesterday = "No activity"

        # This month's cancellations - using date range method
        print(f"DEBUG: Checking month cancellations for {current_month}")

        month_start = f"{current_month}-01"
        next_month = (
            (pacific_now.replace(day=1) + timedelta(days=32))
            .replace(day=1)
            .strftime("%Y-%m-01")
        )

        month_cancellations_result = conn.execute(
            "SELECT COUNT(*) as count FROM cancellations WHERE created_at >= ? AND created_at < ?",
            (month_start, next_month),
        ).fetchone()
        month_cancellations = (
            month_cancellations_result["count"] if month_cancellations_result else 0
        )
        print(f"DEBUG: Month cancellations: {month_cancellations}")

        # This month's detailed stats - UPDATED for 5-box layout
        month_stats = conn.execute(
            """SELECT 
                SUM(CASE WHEN charged = 0 AND excluded = 0 AND status = 'approved' THEN 1 ELSE 0 END) as free,
                SUM(CASE WHEN charged = 1 AND excluded = 0 THEN 1 ELSE 0 END) as charged,
                SUM(CASE WHEN excluded = 1 THEN 1 ELSE 0 END) as excluded,
                SUM(CASE WHEN is_override = 1 THEN 1 ELSE 0 END) as override_count,
                SUM(CASE WHEN deadline_passed = 1 THEN 1 ELSE 0 END) as deadline_passed,
                SUM(CASE WHEN 
                    (sequential_lessons IS NOT NULL AND sequential_lessons != '' AND sequential_lessons != '[]') 
                    OR 
                    (error_report IS NOT NULL AND error_report != '')
                    OR
                    (reschedule_requested = 1)
                THEN 1 ELSE 0 END) as with_notes
               FROM cancellations 
               WHERE created_at >= ? AND created_at < ?""",
            (month_start, next_month),
        ).fetchone()

        free_cancellations = int(month_stats["free"] or 0)
        charged_cancellations = int(month_stats["charged"] or 0)
        excluded_cancellations = int(month_stats["excluded"] or 0)
        override_this_month = int(month_stats["override_count"] or 0)
        deadline_passed = int(month_stats["deadline_passed"] or 0)
        with_notes = int(month_stats["with_notes"] or 0)

        print(f"DEBUG: Free: {free_cancellations}, Charged: {charged_cancellations}")
        print(
            f"DEBUG: Excluded: {excluded_cancellations}, Override: {override_this_month}"
        )
        print(f"DEBUG: Deadline passed: {deadline_passed}, With notes: {with_notes}")

        # Active students
        active_students_result = conn.execute(
            "SELECT COUNT(*) as count FROM students"
        ).fetchone()
        active_students = (
            active_students_result["count"] if active_students_result else 0
        )
        print(f"DEBUG: Active students: {active_students}")

        # Sample students for debugging
        students_sample = conn.execute(
            "SELECT first_name, last_name FROM students LIMIT 5"
        ).fetchall()
        student_names = [
            f"{row['first_name']} {row['last_name']}" for row in students_sample
        ]
        print(f"DEBUG: Sample students: {student_names}")

        # New students this month
        try:
            new_students_month_result = conn.execute(
                "SELECT COUNT(*) as count FROM students WHERE created_at >= ? AND created_at < ?",
                (month_start, next_month),
            ).fetchone()
            new_students_month = (
                new_students_month_result["count"] if new_students_month_result else 0
            )
        except Exception as e:
            print(f"DEBUG: Error getting new students: {e}")
            new_students_month = 0

        # Pending reviews - count only items truly awaiting manager action
        # (status = 'pending' and not overridden). The previous query also counted
        # any charged-not-excluded record whose status wasn't literally 'processed',
        # which inflated the count with already-handled charged cancellations.
        print("DEBUG: Checking pending reviews...")

        pending_reviews_result = conn.execute(
            """SELECT COUNT(*) as count FROM cancellations
               WHERE status = 'pending'
               AND is_override = 0"""
        ).fetchone()
        pending_reviews = (
            pending_reviews_result["count"] if pending_reviews_result else 0
        )
        print(f"DEBUG: Total pending reviews: {pending_reviews}")

        # Check what statuses we have for debugging
        statuses = conn.execute(
            "SELECT status, COUNT(*) as count FROM cancellations GROUP BY status"
        ).fetchall()
        print(
            f"DEBUG: All statuses: {[(row['status'], row['count']) for row in statuses]}"
        )

        # Active membership tiers
        try:
            active_tiers_result = conn.execute(
                "SELECT COUNT(*) as count FROM membership_tiers WHERE active = 1"
            ).fetchone()
            active_tiers = active_tiers_result["count"] if active_tiers_result else 7
            print(f"DEBUG: Active tiers: {active_tiers}")
        except Exception as e:
            print(f"DEBUG: Error getting active tiers: {e}")
            active_tiers = 7

        # System alerts
        system_alerts = 0
        try:
            old_pending = conn.execute(
                "SELECT COUNT(*) as count FROM cancellations WHERE status = 'pending' AND created_at < date('now', '-3 days')"
            ).fetchone()
            old_pending_count = old_pending["count"] if old_pending else 0
            if old_pending_count > 0:
                system_alerts += 1
        except:
            pass

        # Recent cancellation dates for debugging
        all_dates = conn.execute(
            "SELECT created_at FROM cancellations ORDER BY created_at DESC LIMIT 10"
        ).fetchall()
        print(
            f"DEBUG: Recent cancellation dates: {[row['created_at'] for row in all_dates]}"
        )

        conn.close()

        # Final stats object - UPDATED with new 5-box metrics
        stats = {
            # 5-Box Layout Metrics (Primary)
            "today_cancellations": int(today_cancellations),
            "free_cancellations": int(free_cancellations),
            "charged_cancellations": int(charged_cancellations),
            "excluded_cancellations": int(excluded_cancellations),
            "override_this_month": int(override_this_month),
            # Additional Stats (for other parts of the system)
            "today_vs_yesterday": today_vs_yesterday,
            "month_cancellations": int(month_cancellations),
            "active_students": int(active_students),
            "new_students_month": int(new_students_month),
            "pending_reviews": int(pending_reviews),
            "total_students": int(active_students),
            "monthly_cancellations": int(month_cancellations),
            "active_tiers": int(active_tiers),
            "system_alerts": int(system_alerts),
            # Filter Tab Metrics
            "deadline_passed": int(deadline_passed),
            "with_notes": int(with_notes),
            "total_cancellations": int(month_cancellations),
            # Legacy compatibility
            "today_submissions": int(today_cancellations),
            "urgent_cancellations": int(pending_reviews),
        }

        print(f"DEBUG: Final stats: {stats}")
        return stats

    except Exception as e:
        print(f"DEBUG: Error in get_dashboard_stats: {e}")
        import traceback

        traceback.print_exc()
        conn.close()

        # Return safe defaults if there's an error
        return {
            # 5-Box Layout Metrics
            "today_cancellations": 0,
            "free_cancellations": 0,
            "charged_cancellations": 0,
            "excluded_cancellations": 0,
            "override_this_month": 0,
            # Additional Stats
            "today_vs_yesterday": "Error",
            "month_cancellations": 0,
            "active_students": 0,
            "new_students_month": 0,
            "pending_reviews": 0,
            "total_students": 0,
            "monthly_cancellations": 0,
            "active_tiers": 7,
            "system_alerts": 0,
            # Filter Tab Metrics
            "deadline_passed": 0,
            "with_notes": 0,
            "total_cancellations": 0,
            # Legacy compatibility
            "today_submissions": 0,
            "urgent_cancellations": 0,
        }


# ===================================
# NEW: STATUS DETAILS & EXCLUSION/OVERRIDE HELPER FUNCTIONS
# ===================================

def get_cancellation_status_details(student_id, month=None, year=None):
    """
    Get cancellation usage details for a student in a month
    Returns a string like "1 of 2 used this month" or "None"
    """
    current_time = toronto_now()
    if not month:
        month = current_time.month
    if not year:
        year = current_time.year
    
    conn = get_db()
    
    # Get membership info
    student = conn.execute(
        "SELECT membership_level FROM students WHERE id = ?", (student_id,)
    ).fetchone()
    
    if not student:
        conn.close()
        return "None"
    
    # Get membership tier to find free notices allowed
    tier = conn.execute(
        "SELECT free_notices FROM membership_tiers WHERE level = ?",
        (student["membership_level"],)
    ).fetchone()
    
    free_notices_allowed = tier["free_notices"] if tier else 1
    
    # Get count of free cancellations this month (not charged, not excluded)
    count = conn.execute(
        """
        SELECT COUNT(*) as count FROM cancellations 
        WHERE student_id = ? 
        AND strftime('%m', 
            CASE 
                WHEN is_manager_submission = 1 AND manual_submission_date IS NOT NULL
                THEN manual_submission_date
                ELSE created_at
            END
        ) = ? 
        AND strftime('%Y', 
            CASE 
                WHEN is_manager_submission = 1 AND manual_submission_date IS NOT NULL
                THEN manual_submission_date
                ELSE created_at
            END
        ) = ?
        AND excluded = 0
        AND charged = 0
    """,
        (student_id, f"{month:02d}", str(year)),
    ).fetchone()
    
    count_value = count["count"] if count else 0
    conn.close()
    
    if free_notices_allowed == 0:
        return "None"
    
    return f"{count_value} of {free_notices_allowed} used this month"


def send_exclusion_notification(cancellation_id, recipient_type="student"):
    """
    Send notification when a cancellation is marked as excluded
    recipient_type: 'student' or 'manager'
    USES TRIGGER SYSTEM - respects trigger-to-template mappings
    """
    try:
        conn = get_db()
        cancellation = conn.execute(
            """
            SELECT c.*, s.id, s.first_name, s.last_name, s.email, s.parent_first, s.parent_last, 
                   s.phone, s.membership_level
            FROM cancellations c
            JOIN students s ON c.student_id = s.id
            WHERE c.id = ?
            """,
            (cancellation_id,)
        ).fetchone()
        
        if not cancellation:
            print(f"ERROR: Cancellation {cancellation_id} not found")
            conn.close()
            return False
        
        # Convert to dict
        if hasattr(cancellation, "keys"):
            cancellation_dict = dict(cancellation)
        else:
            cancellation_dict = cancellation
        
        # Build student dict for get_template_variables - MUST include id!
        student_dict = {
            "id": cancellation_dict.get("student_id"),
            "first_name": cancellation_dict.get("first_name"),
            "last_name": cancellation_dict.get("last_name"),
            "email": cancellation_dict.get("email"),
            "phone": cancellation_dict.get("phone", ""),
            "parent_first": cancellation_dict.get("parent_first", ""),
            "parent_last": cancellation_dict.get("parent_last", ""),
            "membership_level": cancellation_dict.get("membership_level"),
        }
        
        print(f"DEBUG: Student dict prepared with id={student_dict.get('id')}")
        
        # Calculate status details for variables
        status = calculate_cancellation_status(student_dict)
        tier = get_membership_tier(student_dict.get("membership_level"))
        allowed = 1
        if tier:
            tier_dict = dict(tier) if hasattr(tier, "keys") else tier
            allowed = tier_dict.get("free_notices", 1)
        
        exclusion_info = {
            "exclusion_reason": cancellation_dict.get("exclusion_reason", "Policy exclusion"),
            "status_details": f"{status['used']} free cancellation(s) used this month",
            "used_cancellations": str(status["used"]),
            "allowed_cancellations": str(allowed),
        }
        
        # DETERMINE THE CORRECT TRIGGER BASED ON RECIPIENT TYPE
        if recipient_type == "student":
            trigger_id = "lesson_excluded_student"
        else:  # manager
            trigger_id = "lesson_excluded_manager"
        
        print(f"DEBUG: Using trigger: {trigger_id}")
        
        # USE TRIGGER SYSTEM to send email
        conn.close()
        result = send_email_by_trigger(
            trigger_id,
            student_dict,
            cancellation_dict,
            recipient_type=recipient_type,
            extra_vars=exclusion_info
        )
        
        print(f"DEBUG: Exclusion email result: {result}")
        return result.get("success", False) if isinstance(result, dict) else result
        
    except Exception as e:
        print(f"ERROR: Exception in send_exclusion_notification: {str(e)}")
        import traceback
        traceback.print_exc()
        return False


def send_override_notification(cancellation_id, new_status, recipient_type="student"):
    """
    DEPRECATED: Use send_override_notification_emails() instead.
    This function is kept for backward compatibility but is no longer used internally.
    
    Send notification when an excluded cancellation is overridden
    new_status: 'free' or 'charged'
    recipient_type: 'student' or 'manager'
    """
    print("⚠️ WARNING: send_override_notification() is deprecated. Use send_override_notification_emails() instead.")
    try:
        conn = get_db()
        cancellation = conn.execute(
            """
            SELECT c.*, s.first_name, s.last_name, s.email,
                   (SELECT email FROM admin_users WHERE id = c.manager_submitted_by LIMIT 1) as manager_email
            FROM cancellations c
            JOIN students s ON c.student_id = s.id
            WHERE c.id = ?
            """,
            (cancellation_id,)
        ).fetchone()
        
        conn.close()
        
        if not cancellation:
            return False
        
        # Get template
        template_id = f"override_excluded_to_{new_status}_{recipient_type}"
        template = get_email_template(template_id)
        if not template:
            return False
        
        # Prepare variables
        variables = {
            "STUDENT_NAME": f"{cancellation['first_name']} {cancellation['last_name']}",
            "LESSON_DATE": cancellation["lesson_date"],
            "LESSON_TIME": format_time_for_email(cancellation["lesson_time"]),
            "OVERRIDE_FROM_STATUS": "Excluded",
            "OVERRIDE_TO_STATUS": "Free" if new_status == "free" else "Charged",
            "CHARGE_AMOUNT": "$0" if new_status == "free" else "Full lesson cost",
        }
        
        # Render template
        subject = render_template_string(template["subject"], **variables)
        body = render_template_string(template["body"], **variables)
        
        # Send email
        to_email = cancellation["email"] if recipient_type == "student" else (cancellation["manager_email"] if cancellation["manager_email"] else None)
        if to_email:
            return send_email(to_email, subject, body, template_type=recipient_type)
        
        return False
        
    except Exception as e:
        print(f"Error sending override notification: {str(e)}")
        return False


def format_time_for_email(time_str):
    """Format time string to readable format for emails"""
    try:
        if isinstance(time_str, str):
            parts = time_str.split(":")
            hour = int(parts[0])
            minute = parts[1]
            ampm = "PM" if hour >= 12 else "AM"
            display_hour = hour % 12 if hour % 12 != 0 else 12
            return f"{display_hour}:{minute} {ampm}"
        return str(time_str)
    except:
        return str(time_str)



# ===================================
# AUTHENTICATION ROUTES
# ===================================


@app.route("/")
def index():
    """Home page - redirect to appropriate dashboard"""
    if "user_id" in session:
        return redirect(url_for("dashboard"))
    return redirect(url_for("login"))


@app.route("/login", methods=["GET", "POST"])
def login():
    """Login page for all user types"""
    if request.method == "POST":
        email = request.form["email"].lower().strip()
        password = request.form.get("password", "")

        # Try admin login first
        conn = get_db()
        admin = conn.execute(
            "SELECT * FROM admin_users WHERE email = ?", (email,)
        ).fetchone()

        if admin and check_password_hash(admin["password_hash"], password):
            session["user_id"] = admin["id"]
            session["user_email"] = admin["email"]
            session["user_role"] = admin["role"]
            session["user_name"] = admin["email"]
            conn.close()
            log_action("login", f"Admin login: {admin['role']}")

            # Redirect managers to cancellations page
            return redirect(url_for("manager_cancellations"))

        # Try student login (email only, no password required)
        student = conn.execute(
            "SELECT * FROM students WHERE email = ?", (email,)
        ).fetchone()
        conn.close()

        if student:
            session["user_id"] = student["id"]
            session["user_email"] = student["email"]
            session["user_role"] = "client"
            session["user_name"] = f"{student['first_name']} {student['last_name']}"
            log_action("login", f"Client login: {student['email']}")

            # Redirect clients to cancel lesson page
            return redirect(url_for("client_cancel"))

        log_action("login_failed", f"Failed login attempt: {email}")
        flash("Invalid email or password", "error")

    return render_template("login.html")


@app.route("/logout")
def logout():
    """Logout user"""
    user_name = session.get("user_name", "User")
    log_action("logout", f"User logged out: {user_name}")
    session.clear()

    return redirect(url_for("login"))


# ===================================
# DASHBOARD ROUTES
# ===================================


@app.route("/dashboard")
@login_required
def dashboard():
    """Main dashboard - route to appropriate user dashboard"""
    if session["user_role"] == "client":
        return redirect(url_for("client_dashboard"))
    elif session["user_role"] == "manager":
        return redirect(url_for("manager_dashboard"))
    elif session["user_role"] == "senior_manager":
        return redirect(url_for("senior_dashboard"))
    else:
        return redirect(url_for("login"))


# ===================================
# CLIENT ROUTES
# ===================================


@app.route("/client/dashboard")
@login_required
def client_dashboard():
    """Client dashboard"""
    if session["user_role"] != "client":
        return redirect(url_for("dashboard"))

    conn = get_db()
    client = conn.execute(
        "SELECT * FROM students WHERE id = ?", (session["user_id"],)
    ).fetchone()

    # Get recent cancellations
    recent_cancellations_raw = conn.execute(
        """
        SELECT * FROM cancellations 
        WHERE student_id = ? 
        ORDER BY created_at DESC 
        LIMIT 5
    """,
        (session["user_id"],),
    ).fetchall()
    conn.close()

    if not client:
        flash("Student record not found", "error")
        return redirect(url_for("logout"))

    # Convert string dates to datetime objects for template
    recent_cancellations = []
    for cancellation in recent_cancellations_raw:
        cancellation_dict = dict(cancellation)
        # Convert string dates to datetime objects
        try:
            cancellation_dict["lesson_date"] = datetime.strptime(
                cancellation["lesson_date"], "%Y-%m-%d"
            ).date()

            # Handle both HH:MM and HH:MM:SS time formats
            time_str = str(cancellation["lesson_time"])
            if len(time_str.split(":")) == 2:  # HH:MM format
                cancellation_dict["lesson_time"] = datetime.strptime(
                    time_str, "%H:%M"
                ).time()
            else:  # HH:MM:SS format
                cancellation_dict["lesson_time"] = datetime.strptime(
                    time_str, "%H:%M:%S"
                ).time()

            # Use localize_datetime to handle both old and new timestamp formats
            cancellation_dict["created_at"] = localize_datetime(
                cancellation["created_at"]
            )
            if cancellation_dict["created_at"] is None:
                # Skip if parsing fails
                continue
        except (ValueError, TypeError):
            # Skip problematic records
            continue
        recent_cancellations.append(cancellation_dict)

    # Calculate status
    cancellation_status = calculate_cancellation_status(client)
    cancellation_policy = get_membership_tier(client["membership_level"])

    return render_template(
        "client_dashboard.html",
        client=client,
        cancellation_status=cancellation_status,
        cancellation_policy=cancellation_policy,
        current_month=toronto_now().strftime("%B %Y"),
        recent_cancellations=recent_cancellations,
    )


@app.route("/client/cancel", methods=["GET", "POST"])
@login_required
def client_cancel():
    """Client cancellation form - UPDATED with cancellation note support and date range validation"""
    if session["user_role"] != "client":
        return redirect(url_for("dashboard"))

    conn = get_db()
    client = conn.execute(
        "SELECT * FROM students WHERE id = ?", (session["user_id"],)
    ).fetchone()
    conn.close()

    if not client:
        flash("Student record not found", "error")
        return redirect(url_for("logout"))

    if request.method == "POST":
        lesson_date = request.form["lesson_date"]
        lesson_time = request.form["lesson_time"]
        sequential_dates = request.form.getlist("sequential_dates[]")
        sequential_times = request.form.getlist("sequential_times[]")
        wants_reschedule = "wants_reschedule" in request.form
        reschedule_preferences = request.form.get("reschedule_preferences", "")
        error_report = request.form.get("error_report", "")
        cancellation_note = request.form.get("cancellation_note", "")  # NEW FIELD

        # Validate main lesson
        try:
            # Parse and create timezone-aware datetime
            lesson_datetime = parse_lesson_datetime(lesson_date, lesson_time)
        except ValueError:
            flash("Invalid date or time format", "error")
            return redirect(url_for("client_cancel"))

        # Get current time in Pacific timezone (timezone-aware)
        current_time = toronto_now()

        # ========== NEW: Validate date is within allowed range ==========
        lesson_date_obj = datetime.strptime(lesson_date, "%Y-%m-%d").date()
        max_allowed_date = calculate_max_cancellation_date()

        if lesson_date_obj > max_allowed_date:
            pacific_now = toronto_now()
            today = pacific_now.date()
            next_month_start = (
                date(today.year, today.month + 1, 1)
                if today.month < 12
                else date(today.year + 1, 1, 1)
            )
            days_until_next_month = (next_month_start - today).days

            if days_until_next_month <= 7:
                error_msg = "Cancellations can only be submitted for the current month and next month (within 7 days of next month)."
            else:
                error_msg = "Cancellations can only be submitted for the current month."

            flash(error_msg, "error")
            return redirect(url_for("client_cancel"))
        # ========== END NEW VALIDATION ==========

# Check if lesson has already occurred
        if lesson_datetime <= current_time:
            # Lesson time has passed - allow ONLY if it's the same day
            lesson_date_obj_compare = lesson_datetime.date()
            current_date_obj = current_time.date()
            
            if lesson_date_obj_compare != current_date_obj:
                # Lesson is from a past day (not today) - reject it
                flash("Cannot cancel lessons that have already occurred", "error")
                return redirect(url_for("client_cancel"))
            
            # If we reach here: lesson has passed BUT it's still the same day
            # These must be charged (no free cancellations for same-day past lessons)
            will_charge = True
            charge_reason = "Same-day cancellation after lesson time (charged)"
        else:
            # Lesson hasn't occurred yet - use normal charge logic
            will_charge, charge_reason = will_be_charged(client, lesson_datetime)

        # Determine deadline status for new database fields
        tier = get_membership_tier(client["membership_level"])
        # Use tier-aware deadline calculation so "<time> previous day" tiers
        # (Bronze, Silver, Intro, Legacy, Guest, Welcome) are evaluated against
        # the absolute previous-day clock time rather than a fixed hour offset.
        deadline_passed = (
            is_after_deadline(tier, lesson_datetime, current_time) if tier else False
        )

        # Prepare sequential lessons data
        sequential_lessons = []
        if sequential_dates and sequential_times:
            for i, (seq_date, seq_time) in enumerate(
                zip(sequential_dates, sequential_times)
            ):
                if seq_date and seq_time:
                    sequential_lessons.append({"date": seq_date, "time": seq_time})

        sequential_lessons_json = (
            json.dumps(sequential_lessons) if sequential_lessons else None
        )

        # All cancellations start as 'pending' - manager must review and approve
        # The 'charged' field indicates if it will be charged, but status is always pending initially
        initial_status = "pending"

        # FIXED: Store created_at in Pacific timezone as a plain datetime string
        # This prevents timezone confusion and ensures the submission time is fixed
        pacific_now = toronto_now()
        created_at_string = pacific_now.strftime("%Y-%m-%d %H:%M:%S")

        # Insert cancellation - UPDATED with new fields
        conn = get_db()
        cursor = conn.execute(
            """
            INSERT INTO cancellations
            (student_id, lesson_date, lesson_time, sequential_lessons,
             reschedule_requested, reschedule_preferences, error_report, 
             cancellation_note, charged, deadline_passed, is_override, 
             status, created_at, updated_at, submitted_by, submission_method,
             actual_submission_timestamp, is_manager_submission, suppress_notifications)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'student', 'client_portal',
                    CURRENT_TIMESTAMP, 0, 0)
        """,
            (
                session["user_id"],
                lesson_date,
                lesson_time,
                sequential_lessons_json,
                wants_reschedule,
                reschedule_preferences,
                error_report,
                cancellation_note,  # NEW
                will_charge,
                deadline_passed,  # NEW
                False,  # is_override starts as False
                initial_status,  # Auto-approve free, or set as charged
                created_at_string,  # FIXED: Store as Pacific time string - this NEVER changes
                created_at_string,  # updated_at can change, but created_at stays fixed
            ),
        )
        cancellation_id = cursor.lastrowid
        
        # Note: We deliberately do NOT mark welcome_free_used here at submission time.
        # The flag is only set once a manager confirms the cancellation as free
        # (in process_cancellation / process_all_pending). Until then, the lifetime
        # check in will_be_charged() relies on a dynamic count of charged=0
        # cancellations, which already accounts for this still-pending submission.
        # This prevents the bug where re-evaluating a pending Welcome cancellation
        # during manager processing would see the flag set by its own submission and
        # incorrectly classify it as "already used".
        
        conn.commit()
        conn.close()

        # Prepare cancellation data for email - UPDATED with new fields
        cancellation_data = {
            "id": cancellation_id,
            "lesson_date": lesson_date,
            "lesson_time": lesson_time,
            "sequential_lessons": sequential_lessons_json,
            "reschedule_requested": wants_reschedule,
            "reschedule_preferences": reschedule_preferences,
            "error_report": error_report,
            "cancellation_note": cancellation_note,  # NEW
            "charged": will_charge,
            "deadline_passed": deadline_passed,  # NEW
            "manager_notes": charge_reason,
        }

        # Send confirmation email
        try:
            email_result = send_cancellation_confirmation(client, cancellation_data)

            if email_result["success"]:
                log_action(
                    "email_sent",
                    f"Confirmation email sent for cancellation {cancellation_id}",
                )
            else:
                log_action(
                    "email_failed",
                    f"Failed to send confirmation email: {email_result['message']}",
                )
                print(f"Email sending failed: {email_result['message']}")
        except Exception as e:
            log_action(
                "email_error",
                f"Email system error for cancellation {cancellation_id}: {str(e)}",
            )
            print(f"Email system error: {str(e)}")

        # Send notification email to managers
        try:
            manager_result = send_manager_notification(client, cancellation_data)

            if manager_result["success"]:
                log_action(
                    "email_sent",
                    f"Manager notification sent for cancellation {cancellation_id}: {manager_result['message']}",
                )
                print(f"✅ Manager notification sent: {manager_result['message']}")
            else:
                log_action(
                    "email_failed",
                    f"Failed to send manager notification: {manager_result['message']}",
                )
                print(f"❌ Manager email failed: {manager_result['message']}")
        except Exception as e:
            log_action(
                "email_error",
                f"Manager email system error for cancellation {cancellation_id}: {str(e)}",
            )
            print(f"❌ Manager email error: {str(e)}")

        log_action(
            "cancellation_submitted",
            f"Lesson: {lesson_date} {lesson_time}, Charged: {will_charge}, Note: {'Yes' if cancellation_note else 'No'}",
        )

        # Flash appropriate message - ENHANCED with note acknowledgment
        if will_charge:
            if cancellation_note:
                flash(f"Cancellation submitted with note. {charge_reason}", "warning")
            else:
                flash(f"Cancellation submitted. {charge_reason}", "warning")
        else:
            if cancellation_note:
                flash(
                    "Cancellation with note submitted successfully! This was processed as a free cancellation.",
                    "success",
                )
            else:
                flash(
                    "Cancellation submitted successfully! This was processed as a free cancellation.",
                    "success",
                )

        return redirect(url_for("client_dashboard"))

    # GET request - show form
    cancellation_status = calculate_cancellation_status(client)
    cancellation_policy = get_membership_tier(client["membership_level"])
    pacific_today = toronto_now().date()
    max_cancellation_date = calculate_max_cancellation_date()  # NEW LINE

    return render_template(
        "client_cancel.html",
        client=client,
        cancellation_status=cancellation_status,
        cancellation_policy=cancellation_policy,
        min_date=pacific_today.isoformat(),
        max_date=max_cancellation_date.isoformat(),  # NEW LINE
    )


# Backend Route Fix
@app.route("/client/history")
@login_required
def client_history():
    """Client cancellation history"""
    if session["user_role"] != "client":
        return redirect(url_for("dashboard"))
    conn = get_db()
    client = conn.execute(
        "SELECT * FROM students WHERE id = ?", (session["user_id"],)
    ).fetchone()
    # Get all cancellations for this client
    cancellations = conn.execute(
        """
        SELECT * FROM cancellations
        WHERE student_id = ?
        ORDER BY created_at DESC
    """,
        (session["user_id"],),
    ).fetchall()
    conn.close()

    # Group by month
    cancellations_by_month = defaultdict(list)
    for cancellation in cancellations:
        # Use localize_datetime to handle both old and new timestamp formats
        created_date = localize_datetime(cancellation["created_at"])
        if created_date is None:
            # Fallback if parsing fails
            created_date = toronto_now()
        month_key = created_date.strftime("%Y-%m")

        # Parse sequential lessons if any
        sequential_lessons = []
        if cancellation["sequential_lessons"]:
            try:
                sequential_lessons = eval(cancellation["sequential_lessons"])
                # Convert string dates to datetime objects for template
                for lesson in sequential_lessons:
                    if isinstance(lesson.get("date"), str):
                        lesson["date"] = datetime.strptime(
                            lesson["date"], "%Y-%m-%d"
                        ).date()
                    if isinstance(lesson.get("time"), str):
                        try:
                            lesson["time"] = datetime.strptime(
                                lesson["time"], "%H:%M:%S"
                            ).time()
                        except ValueError:
                            lesson["time"] = datetime.strptime(
                                lesson["time"], "%H:%M"
                            ).time()
            except:
                sequential_lessons = []

        cancellation_dict = dict(cancellation)
        cancellation_dict["sequential_lessons"] = sequential_lessons
        cancellation_dict["created_at"] = created_date
        cancellation_dict["lesson_date"] = datetime.strptime(
            cancellation["lesson_date"], "%Y-%m-%d"
        ).date()

        # Add formatted sequential lessons message
        if sequential_lessons:
            cancellation_dict["sequential_lessons_formatted"] = (
                format_sequential_lessons(
                    sequential_lessons, cancellation["lesson_date"]
                )
            )
        else:
            cancellation_dict["sequential_lessons_formatted"] = (
                "None"
            )

        # Fix: Handle both HH:MM and HH:MM:SS time formats
        lesson_time_str = cancellation["lesson_time"]
        try:
            # Try parsing with seconds first
            cancellation_dict["lesson_time"] = datetime.strptime(
                lesson_time_str, "%H:%M:%S"
            ).time()
        except ValueError:
            # If that fails, try parsing without seconds
            try:
                cancellation_dict["lesson_time"] = datetime.strptime(
                    lesson_time_str, "%H:%M"
                ).time()
            except ValueError:
                # If both fail, set a default time
                cancellation_dict["lesson_time"] = datetime.strptime(
                    "00:00", "%H:%M"
                ).time()

        cancellations_by_month[month_key].append(cancellation_dict)

    # Format for template
    formatted_months = []
    for month_key in sorted(cancellations_by_month.keys(), reverse=True):
        month_date = datetime.strptime(month_key, "%Y-%m")
        formatted_months.append(
            {
                "month": month_key,
                "month_display": month_date.strftime("%B %Y"),
                "count": len(cancellations_by_month[month_key]),
                "cancellations": cancellations_by_month[month_key],
            }
        )

    # Current month stats
    current_month_stats = calculate_cancellation_status(client)

    return render_template(
        "client_history.html",
        client=client,
        cancellations=cancellations,
        cancellations_by_month=formatted_months,
        current_month=toronto_now().strftime("%B %Y"),
        current_month_stats=current_month_stats,
        has_more_history=False,
    )


# ===================================
# MANAGER ROUTES
# ===================================


@app.route("/manager/dashboard")
@login_required
@admin_required
def manager_dashboard():
    """Manager dashboard"""
    # Add current_time for template - keep timezone-aware to match created_at
    current_time = toronto_now()

    # Get properly calculated stats
    stats = get_dashboard_stats()
    conn = get_db()

    # Recent cancellations with student names
    recent_cancellations_raw = conn.execute(
        """
        SELECT c.*, s.first_name, s.last_name, s.membership_level
        FROM cancellations c
        JOIN students s ON c.student_id = s.id
        ORDER BY c.created_at DESC
        LIMIT 10
    """
    ).fetchall()

    # Convert to proper format for template
    recent_cancellations = []
    for cancellation in recent_cancellations_raw:
        cancellation_dict = dict(cancellation)

        # Use localize_datetime to handle both old and new timestamp formats
        cancellation_dict["created_at"] = localize_datetime(cancellation["created_at"])
        if cancellation_dict["created_at"] is None:
            # Fallback to current time if parsing fails
            cancellation_dict["created_at"] = toronto_now()

        # Convert lesson_date string to date object
        cancellation_dict["lesson_date"] = datetime.strptime(
            cancellation["lesson_date"], "%Y-%m-%d"
        ).date()

        # Convert lesson_time string to time object (handle both HH:MM and HH:MM:SS)
        lesson_time_str = cancellation["lesson_time"]
        try:
            cancellation_dict["lesson_time"] = datetime.strptime(
                lesson_time_str, "%H:%M:%S"
            ).time()
        except ValueError:
            try:
                cancellation_dict["lesson_time"] = datetime.strptime(
                    lesson_time_str, "%H:%M"
                ).time()
            except ValueError:
                # Default fallback
                cancellation_dict["lesson_time"] = datetime.strptime(
                    "00:00", "%H:%M"
                ).time()

        # Create student_name for template
        cancellation_dict["student_name"] = (
            f"{cancellation['first_name']} {cancellation['last_name']}"
        )

        recent_cancellations.append(cancellation_dict)

    # Pending actions (charged cancellations, exclusions needed, etc.)
    pending_actions_raw = conn.execute(
        """
        SELECT c.*, s.first_name, s.last_name, s.membership_level
        FROM cancellations c
        JOIN students s ON c.student_id = s.id
        WHERE c.charged = 1 AND c.excluded = 0
        ORDER BY c.created_at DESC
        LIMIT 5
    """
    ).fetchall()

    # Convert pending actions to proper format
    pending_actions = []
    for action in pending_actions_raw:
        action_dict = dict(action)
        # Use localize_datetime to handle both old and new timestamp formats
        action_dict["created_at"] = localize_datetime(action["created_at"])
        if action_dict["created_at"] is None:
            action_dict["created_at"] = toronto_now()

        action_dict["lesson_date"] = datetime.strptime(
            action["lesson_date"], "%Y-%m-%d"
        ).date()

        lesson_time_str = action["lesson_time"]
        try:
            action_dict["lesson_time"] = datetime.strptime(
                lesson_time_str, "%H:%M:%S"
            ).time()
        except ValueError:
            try:
                action_dict["lesson_time"] = datetime.strptime(
                    lesson_time_str, "%H:%M"
                ).time()
            except ValueError:
                action_dict["lesson_time"] = datetime.strptime("00:00", "%H:%M").time()

        action_dict["student_name"] = f"{action['first_name']} {action['last_name']}"
        pending_actions.append(action_dict)

    # Monthly stats for chart
    monthly_stats_raw = conn.execute(
        """
        SELECT
            strftime('%Y-%m', created_at) as month,
            COUNT(*) as total,
            SUM(CASE WHEN charged = 0 THEN 1 ELSE 0 END) as free,
            SUM(CASE WHEN charged = 1 THEN 1 ELSE 0 END) as charged
        FROM cancellations
        WHERE created_at >= date('now', '-6 months')
        GROUP BY strftime('%Y-%m', created_at)
        ORDER BY month
    """
    ).fetchall()

    # Prepare chart data - ensure we have data
    if monthly_stats_raw:
        monthly_labels = []
        monthly_totals = []
        monthly_charged = []
        monthly_free = []

        for stat in monthly_stats_raw:
            month_date = datetime.strptime(stat["month"], "%Y-%m")
            monthly_labels.append(month_date.strftime("%B %Y"))
            monthly_totals.append(stat["total"] or 0)
            monthly_charged.append(stat["charged"] or 0)
            monthly_free.append(stat["free"] or 0)
    else:
        # Default empty data if no cancellations
        monthly_labels = ["No Data"]
        monthly_totals = [0]
        monthly_charged = [0]
        monthly_free = [0]

    # Create recent activity from actual recent cancellations
    recent_activity = []
    for cancellation in recent_cancellations[:5]:  # Take first 5
        activity_type = (
            "excluded"
            if cancellation.get("excluded")
            else ("charged" if cancellation.get("charged") else "new")
        )
        activity_title = {
            "new": "New Cancellation",
            "charged": "Charged Cancellation",
            "excluded": "Excluded Cancellation",
        }.get(activity_type, "Cancellation")

        # Calculate time ago - use PST time for comparison
        created_at = localize_datetime(cancellation["created_at"])
        if created_at is None:
            created_at = toronto_now()  # Fallback
        time_diff = toronto_now() - created_at
        if time_diff.days > 0:
            time_ago = f"{time_diff.days} day{'s' if time_diff.days != 1 else ''} ago"
        elif time_diff.seconds > 3600:
            hours = time_diff.seconds // 3600
            time_ago = f"{hours} hour{'s' if hours != 1 else ''} ago"
        else:
            minutes = max(1, time_diff.seconds // 60)
            time_ago = f"{minutes} minute{'s' if minutes != 1 else ''} ago"

        recent_activity.append(
            {
                "type": activity_type,
                "title": activity_title,
                "description": f"Lesson on {cancellation['lesson_date'].strftime('%m/%d/%Y')}",
                "student_name": cancellation["student_name"],
                "time_ago": time_ago,
            }
        )

    conn.close()

    return render_template(
        "manager_dashboard.html",
        current_time=current_time,
        stats=stats,
        recent_cancellations=recent_cancellations,
        pending_actions=pending_actions,
        monthly_stats=monthly_stats_raw,
        monthly_labels=monthly_labels,
        monthly_totals=monthly_totals,
        monthly_charged=monthly_charged,
        monthly_free=monthly_free,
        recent_activity=recent_activity,
        alerts=[],  # Add any system alerts here
    )


@app.route("/manager/students", methods=["GET", "POST"])
@login_required
@admin_required
def manager_students():
    """Manager students page"""
    if request.method == "POST":
        # Handle POST requests (add student, bulk import, etc.)
        action = request.form.get("action", "add_student")

        if action == "add_student":
            # Add single student
            conn = get_db()
            try:
                current_time = toronto_now().strftime("%Y-%m-%d %H:%M:%S")
                membership_level = request.form["membership_level"]
                welcome_start_date = current_time if membership_level == "Welcome Package" else None
                
                conn.execute(
                    """INSERT INTO students 
                       (first_name, last_name, parent_first, parent_last, email, phone, membership_level, welcome_package_date_started, created_at)
                       VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                    (
                        request.form["first_name"],
                        request.form["last_name"],
                        request.form.get("parent_first", ""),
                        request.form.get("parent_last", ""),
                        request.form["email"],
                        request.form.get("phone", ""),
                        membership_level,
                        welcome_start_date,
                        current_time,
                    ),
                )
                
                # Record in package history if Welcome Package
                if membership_level == "Welcome Package":
                    student_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
                    track_package_upgrade(student_id, "None", "Welcome Package", False, conn)
                
                conn.commit()
                conn.close()
                flash("Student added successfully!", "success")
            except Exception as e:
                conn.close()
                flash(f"Error adding student: {str(e)}", "error")

        elif action == "bulk_import":
            # Handle bulk import from CSV file or pasted data
            csv_data = None
            default_membership = request.form.get("default_membership", "Bronze")

            # Check if file was uploaded
            if "csv_file" in request.files and request.files["csv_file"].filename:
                file = request.files["csv_file"]
                csv_data = file.read().decode("utf-8")
            # Otherwise check for pasted data
            elif request.form.get("paste_data"):
                csv_data = request.form.get("paste_data")

            if not csv_data:
                flash("No data provided for import", "error")
            else:
                # Get database connection
                conn = get_db()

                try:
                    # Parse CSV data
                    csv_reader = csv.reader(io.StringIO(csv_data))
                    imported = 0
                    errors = []

                    for row_num, row in enumerate(csv_reader, 1):
                        # Skip empty rows
                        if not row or all(not cell.strip() for cell in row):
                            continue

                        # Check minimum required columns (at least email is needed)
                        if len(row) < 5:
                            errors.append(
                                f"Row {row_num}: Insufficient data (need at least: First Name, Last Name, Parent First, Parent Last, Email)"
                            )
                            continue

                        try:
                            # Get values with defaults for optional fields
                            first_name = row[0].strip() if len(row) > 0 else ""
                            last_name = row[1].strip() if len(row) > 1 else ""
                            parent_first = row[2].strip() if len(row) > 2 else ""
                            parent_last = row[3].strip() if len(row) > 3 else ""
                            email = row[4].strip() if len(row) > 4 else ""
                            phone = row[5].strip() if len(row) > 5 else ""
                            membership = (
                                row[6].strip() if len(row) > 6 else default_membership
                            )

                            # Validate required fields
                            if not email:
                                errors.append(f"Row {row_num}: Email is required")
                                continue
                            if not first_name:
                                errors.append(f"Row {row_num}: First name is required")
                                continue
                            if not last_name:
                                errors.append(f"Row {row_num}: Last name is required")
                                continue

                            # Insert into database
                            current_time = toronto_now().strftime("%Y-%m-%d %H:%M:%S")
                            welcome_start_date = current_time if membership == "Welcome Package" else None
                            
                            conn.execute(
                                """INSERT INTO students 
                                   (first_name, last_name, parent_first, parent_last, email, phone, membership_level, welcome_package_date_started, created_at)
                                   VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                                (
                                    first_name,
                                    last_name,
                                    parent_first,
                                    parent_last,
                                    email,
                                    phone,
                                    membership,
                                    welcome_start_date,
                                    current_time,
                                ),
                            )
                            imported += 1

                        except sqlite3.IntegrityError as e:
                            if "UNIQUE constraint failed: students.email" in str(e):
                                errors.append(
                                    f"Row {row_num}: Email '{email}' already exists"
                                )
                            else:
                                errors.append(
                                    f"Row {row_num}: Database error - {str(e)}"
                                )
                        except Exception as e:
                            errors.append(f"Row {row_num}: {str(e)}")

                    # Commit all successful imports
                    conn.commit()

                    # Show results
                    if imported > 0:
                        flash(
                            f"Successfully imported {imported} student(s)!", "success"
                        )
                    if errors:
                        error_message = (
                            f"{len(errors)} error(s) occurred:<br>"
                            + "<br>".join(errors[:10])
                        )
                        if len(errors) > 10:
                            error_message += (
                                f"<br>... and {len(errors) - 10} more errors"
                            )
                        flash(error_message, "warning")

                    log_action(
                        "bulk_import",
                        f"Imported {imported} students, {len(errors)} errors",
                    )

                except Exception as e:
                    flash(f"Import failed: {str(e)}", "error")
                    log_action("bulk_import_error", str(e))
                finally:
                    conn.close()

        return redirect(url_for("manager_students"))

    # GET request - display students
    conn = get_db()

    # Get search and filter parameters
    search = request.args.get("search", "")
    membership_filter = request.args.get("membership", "")
    sort_by = request.args.get("sort", "name")
    page = request.args.get("page", 1, type=int)
    per_page = 12  # Students per page

    # Build the query with filters
    where_conditions = []
    params = []

    if search:
        where_conditions.append(
            "(s.first_name LIKE ? OR s.last_name LIKE ? OR s.email LIKE ? OR s.phone LIKE ?)"
        )
        search_param = f"%{search}%"
        params.extend([search_param, search_param, search_param, search_param])

    if membership_filter:
        where_conditions.append("s.membership_level = ?")
        params.append(membership_filter)

    where_clause = " AND ".join(where_conditions)
    if where_clause:
        where_clause = "WHERE " + where_clause

    # Sort order - Updated with new options
    sort_orders = {
        "name": "s.last_name, s.first_name",
        "email": "s.email",
        "membership": "s.membership_level",
        "recent": "s.created_at DESC",
        "recent_activity": "s.created_at DESC",  # For now, same as recent - can be enhanced later
        "status": "s.membership_level DESC",  # For now, sort by membership - can be enhanced later
    }
    order_by = sort_orders.get(sort_by, "s.last_name, s.first_name")

    # Get students with cancellation statistics
    students_query = f"""
        SELECT s.*,
               COUNT(c.id) as total_cancellations,
               SUM(CASE WHEN strftime('%Y-%m', 
                   CASE 
                       WHEN c.is_manager_submission = 1 AND c.manual_submission_date IS NOT NULL
                       THEN c.manual_submission_date
                       ELSE c.created_at
                   END
               ) = strftime('%Y-%m', 'now') THEN 1 ELSE 0 END) as cancellations_this_month,
               SUM(CASE WHEN c.charged = 0 AND c.excluded = 0 AND strftime('%Y-%m', 
                   CASE 
                       WHEN c.is_manager_submission = 1 AND c.manual_submission_date IS NOT NULL
                       THEN c.manual_submission_date
                       ELSE c.created_at
                   END
               ) = strftime('%Y-%m', 'now') THEN 1 ELSE 0 END) as free_used_this_month,
               MAX(c.created_at) as last_cancellation_date
        FROM students s
        LEFT JOIN cancellations c ON s.id = c.student_id
        {where_clause}
        GROUP BY s.id
        ORDER BY {order_by}
    """

    students_raw = conn.execute(students_query, params).fetchall()

    # Process students data for template
    students = []
    for student_raw in students_raw:
        student = dict(student_raw)

        # Convert created_at to datetime object if it exists and is a string
        if student.get("created_at") and isinstance(student["created_at"], str):
            try:
                student["created_at"] = datetime.strptime(
                    student["created_at"], "%Y-%m-%d %H:%M:%S"
                )
                # Localize to Pacific timezone since database stores in Pacific time
                pacific_tz = pytz.timezone("America/Los_Angeles")
                student["created_at"] = pacific_tz.localize(student["created_at"])
            except ValueError:
                # If parsing fails, set to current time in Pacific
                student["created_at"] = toronto_now()
        elif not student.get("created_at"):
            student["created_at"] = toronto_now()

        # Calculate statistics
        student["total_cancellations"] = student.get("total_cancellations") or 0
        student["cancellations_this_month"] = (
            student.get("cancellations_this_month") or 0
        )
        student["free_used_this_month"] = student.get("free_used_this_month") or 0

        # Get membership tier limits
        membership_limits = {
            "Bronze": 1,
            "Silver": 2,
            "Gold": 4,
            "Intro Package": 1,
            "Legacy": 1,
            "Guest": 1,
            "Welcome Package": 1,
        }

        limit = membership_limits.get(student["membership_level"], 1)
        student["free_remaining"] = max(0, limit - student["free_used_this_month"])

        students.append(student)

    # Apply custom sorting for specific sort types
    if sort_by == "recent_activity":
        # Sort by most recent cancellation date, then by created date
        students.sort(
            key=lambda s: (
                (
                    datetime.strptime(
                        s.get("last_cancellation_date", "1900-01-01 00:00:00"),
                        "%Y-%m-%d %H:%M:%S",
                    )
                    if s.get("last_cancellation_date")
                    else datetime.min
                ),
                s["created_at"],
            ),
            reverse=True,
        )
    elif sort_by == "status":
        # Sort by free_remaining (descending - most free cancellations first)
        students.sort(key=lambda s: s["free_remaining"], reverse=True)

    # Pagination
    total_students = len(students)
    start = (page - 1) * per_page
    end = start + per_page
    paginated_students = students[start:end]

    # Create pagination object
    class Pagination:
        def __init__(self, page, per_page, total):
            self.page = page
            self.per_page = per_page
            self.total = total
            self.pages = (total + per_page - 1) // per_page
            self.has_prev = page > 1
            self.has_next = page < self.pages
            self.prev_num = page - 1 if self.has_prev else None
            self.next_num = page + 1 if self.has_next else None

        def iter_pages(
            self, left_edge=2, left_current=2, right_current=3, right_edge=2
        ):
            last = self.pages
            for num in range(1, last + 1):
                if (
                    num <= left_edge
                    or (self.page - left_current - 1 < num < self.page + right_current)
                    or num > last - right_edge
                ):
                    yield num

    pagination = Pagination(page, per_page, total_students)

    # Calculate statistics for the template - Updated with new stats
    current_month = toronto_now().strftime("%Y-%m")

    # Total students
    total_students_count = len(students)

    # Students active this month (had cancellations)
    active_this_month = len([s for s in students if s["cancellations_this_month"] > 0])

    # New students this month
    new_this_month = len(
        [s for s in students if s["created_at"].strftime("%Y-%m") == current_month]
    )

    # Membership breakdown - Updated with all membership types
    gold_members = len([s for s in students if s["membership_level"] == "Gold"])
    silver_members = len([s for s in students if s["membership_level"] == "Silver"])
    bronze_members = len([s for s in students if s["membership_level"] == "Bronze"])
    welcome_members = len(
        [s for s in students if s["membership_level"] == "Welcome Package"]
    )
    guest_members = len([s for s in students if s["membership_level"] == "Guest"])

    stats = {
        "total_students": total_students_count,
        "active_this_month": active_this_month,
        "new_this_month": new_this_month,
        "gold_members": gold_members,
        "silver_members": silver_members,
        "bronze_members": bronze_members,  # NEW
        "welcome_members": welcome_members,  # NEW
        "guest_members": guest_members,  # NEW
    }

    # Get membership tiers for dropdown (use static list if table doesn't exist)
    try:
        membership_tiers_raw = conn.execute(
            "SELECT level FROM membership_tiers WHERE active = 1 ORDER BY sort_order"
        ).fetchall()
        membership_tiers = [tier["level"] for tier in membership_tiers_raw]
    except:
        # Fallback to static list if table doesn't exist
        membership_tiers = [
            "Bronze",
            "Silver",
            "Gold",
            "Intro Package",
            "Legacy",
            "Guest",
            "Welcome Package",
        ]

    conn.close()

    return render_template(
        "manager_students.html",
        students=paginated_students,
        membership_tiers=membership_tiers,
        stats=stats,
        pagination=pagination,
        search=search,
        current_membership_filter=membership_filter,
        current_sort=sort_by,
    )


def process_cancellation_dates(cancellation):
    """Process cancellation dates to ensure consistent formatting for templates"""
    cancellation_dict = (
        dict(cancellation) if hasattr(cancellation, "keys") else cancellation
    )

    # Convert created_at string to datetime - FIXED: Always use Pacific timezone
    if isinstance(cancellation_dict.get("created_at"), str):
        try:
            # Parse the datetime string
            naive_dt = datetime.strptime(
                cancellation_dict["created_at"], "%Y-%m-%d %H:%M:%S"
            )
            # Localize to Pacific timezone since database stores in Pacific time
            pacific_tz = pytz.timezone("America/Los_Angeles")
            cancellation_dict["created_at"] = pacific_tz.localize(naive_dt)
        except (ValueError, TypeError):
            cancellation_dict["created_at"] = toronto_now()
    elif isinstance(cancellation_dict.get("created_at"), datetime):
        # If it's already a datetime object but naive, localize it
        if cancellation_dict["created_at"].tzinfo is None:
            pacific_tz = pytz.timezone("America/Los_Angeles")
            cancellation_dict["created_at"] = pacific_tz.localize(
                cancellation_dict["created_at"]
            )

    # Convert lesson_date string to date object
    if isinstance(cancellation_dict.get("lesson_date"), str):
        try:
            cancellation_dict["lesson_date"] = datetime.strptime(
                cancellation_dict["lesson_date"], "%Y-%m-%d"
            ).date()
        except (ValueError, TypeError):
            cancellation_dict["lesson_date"] = toronto_now().date()

    # Convert lesson_time string to time object (handle both HH:MM and HH:MM:SS)
    if isinstance(cancellation_dict.get("lesson_time"), str):
        lesson_time_str = cancellation_dict["lesson_time"]
        try:
            cancellation_dict["lesson_time"] = datetime.strptime(
                lesson_time_str, "%H:%M:%S"
            ).time()
        except ValueError:
            try:
                cancellation_dict["lesson_time"] = datetime.strptime(
                    lesson_time_str, "%H:%M"
                ).time()
            except ValueError:
                cancellation_dict["lesson_time"] = datetime.strptime(
                    "00:00", "%H:%M"
                ).time()

    # Convert actual_submission_timestamp string to datetime (for manager submissions)
    if isinstance(cancellation_dict.get("actual_submission_timestamp"), str):
        try:
            naive_dt = datetime.strptime(
                cancellation_dict["actual_submission_timestamp"], "%Y-%m-%d %H:%M:%S"
            )
            pacific_tz = pytz.timezone("America/Los_Angeles")
            cancellation_dict["actual_submission_timestamp"] = pacific_tz.localize(naive_dt)
        except (ValueError, TypeError):
            cancellation_dict["actual_submission_timestamp"] = cancellation_dict.get("created_at")

    # Convert manual_submission_date string to datetime (for manager submissions)
    if isinstance(cancellation_dict.get("manual_submission_date"), str):
        try:
            # Try format without seconds first (format from manager form: "2026-04-20 14:30")
            try:
                naive_dt = datetime.strptime(
                    cancellation_dict["manual_submission_date"], "%Y-%m-%d %H:%M"
                )
            except ValueError:
                # Fallback to format with seconds if first fails
                naive_dt = datetime.strptime(
                    cancellation_dict["manual_submission_date"], "%Y-%m-%d %H:%M:%S"
                )
            pacific_tz = pytz.timezone("America/Los_Angeles")
            cancellation_dict["manual_submission_date"] = pacific_tz.localize(naive_dt)
        except (ValueError, TypeError):
            cancellation_dict["manual_submission_date"] = None

    # Process submitted date and time - CRITICAL FIX: Always set submitted_date, preferring policy date for managers
    # Determine which date to use based on submission type
    
    # Check if this is a manager submission
    is_manager_sub = cancellation_dict.get("is_manager_submission") == 1 or cancellation_dict.get("is_manager_submission") == True
    
    # Initialize defaults
    cancellation_dict["submitted_date"] = None
    cancellation_dict["submitted_time"] = None
    cancellation_dict["is_policy_submitted_date"] = False
    reference_datetime = None
    
    if is_manager_sub:
        # This IS a manager submission - try to use policy date FIRST
        policy_date = cancellation_dict.get("manual_submission_date")
        
        if policy_date and isinstance(policy_date, datetime):
            # Policy date exists and is a proper datetime object - USE IT
            cancellation_dict["submitted_date"] = policy_date.date()
            cancellation_dict["submitted_time"] = policy_date.strftime("%I:%M %p PST")
            reference_datetime = policy_date
            cancellation_dict["is_policy_submitted_date"] = True
        elif cancellation_dict.get("created_at"):
            # Fallback to created_at if policy date is missing or invalid
            created_at = cancellation_dict["created_at"]
            cancellation_dict["submitted_date"] = created_at.date()
            cancellation_dict["submitted_time"] = created_at.strftime("%I:%M %p PST")
            reference_datetime = created_at
            cancellation_dict["is_policy_submitted_date"] = False
        else:
            # Emergency fallback - use current time
            cancellation_dict["submitted_date"] = toronto_now().date()
            cancellation_dict["submitted_time"] = toronto_now().strftime("%I:%M %p PST")
            reference_datetime = toronto_now()
            cancellation_dict["is_policy_submitted_date"] = False
            
    elif cancellation_dict.get("created_at"):
        # For student submissions, use actual submission timestamp
        created_at = cancellation_dict["created_at"]
        cancellation_dict["submitted_date"] = created_at.date()
        cancellation_dict["submitted_time"] = created_at.strftime("%I:%M %p PST")
        reference_datetime = created_at
        cancellation_dict["is_policy_submitted_date"] = False
    else:
        # Emergency fallback for student
        cancellation_dict["submitted_date"] = toronto_now().date()
        cancellation_dict["submitted_time"] = toronto_now().strftime("%I:%M %p PST")
        reference_datetime = toronto_now()
        cancellation_dict["is_policy_submitted_date"] = False

    # Calculate time ago using the appropriate reference date
    if reference_datetime:
        now = toronto_now()  # Use PST time, not system time
        time_diff = now - reference_datetime
        if time_diff.days > 0:
            cancellation_dict["time_ago"] = (
                f"{time_diff.days} day{'s' if time_diff.days != 1 else ''} ago"
            )
        elif time_diff.seconds > 3600:
            hours = time_diff.seconds // 3600
            cancellation_dict["time_ago"] = (
                f"{hours} hour{'s' if hours != 1 else ''} ago"
            )
        else:
            minutes = max(1, time_diff.seconds // 60)
            cancellation_dict["time_ago"] = (
                f"{minutes} minute{'s' if minutes != 1 else ''} ago"
            )

    # Process sequential lessons
    if cancellation_dict.get("sequential_lessons"):
        try:
            if isinstance(cancellation_dict["sequential_lessons"], str):
                sequential_lessons = eval(cancellation_dict["sequential_lessons"])
            else:
                sequential_lessons = cancellation_dict["sequential_lessons"]

            # Convert string dates/times to objects for template consistency
            for lesson in sequential_lessons:
                if isinstance(lesson.get("date"), str):
                    lesson["date"] = datetime.strptime(
                        lesson["date"], "%Y-%m-%d"
                    ).date()
                if isinstance(lesson.get("time"), str):
                    try:
                        lesson["time"] = datetime.strptime(
                            lesson["time"], "%H:%M:%S"
                        ).time()
                    except ValueError:
                        lesson["time"] = datetime.strptime(
                            lesson["time"], "%H:%M"
                        ).time()

            cancellation_dict["sequential_lessons"] = sequential_lessons
        except:
            cancellation_dict["sequential_lessons"] = []

    # Add computed status properties for analytics
    cancellation_dict["is_free"] = (
        not cancellation_dict.get("charged", False)
        and not cancellation_dict.get("excluded", False)
        and cancellation_dict.get("status") == "approved"
    )

    cancellation_dict["is_charged"] = cancellation_dict.get(
        "charged", False
    ) and not cancellation_dict.get("excluded", False)

    # Ensure boolean fields are properly set
    cancellation_dict["is_override"] = bool(cancellation_dict.get("is_override", False))
    cancellation_dict["deadline_passed"] = bool(
        cancellation_dict.get("deadline_passed", False)
    )
    cancellation_dict["excluded"] = bool(cancellation_dict.get("excluded", False))

    return cancellation_dict


@app.route("/manager/api/cancellation/process", methods=["POST"])
@login_required
@admin_required
def process_cancellation():
    """Process individual cancellation - UPDATED with new approve as policy functionality"""
    data = request.json
    action = data.get("action")  # 'approve_policy', 'force_free', 'force_charge'
    cancellation_id = data.get("cancellation_id")
    reason = data.get("reason", "")

    # DEBUG LOGGING
    print(f"\n=== PROCESS CANCELLATION DEBUG ===")
    print(f"Action: {action}")
    print(f"Cancellation ID: {cancellation_id}")
    print(f"Reason: {reason}")
    
    # Get email suppression flags
    suppress_student_email = data.get("suppress_student_email", False)
    suppress_manager_email = data.get("suppress_manager_email", False)
    print(f"Suppress student email: {suppress_student_email}")
    print(f"Suppress manager email: {suppress_manager_email}")

    if not action or not cancellation_id:
        return jsonify({"success": False, "message": "Missing required fields"})

    # Only force actions require a reason
    if action in ["force_free", "force_charge"] and not reason.strip():
        return jsonify(
            {
                "success": False,
                "message": "Override reason is required for force actions",
            }
        )

    conn = get_db()

    try:
        # Get cancellation and student data BEFORE updating
        cancellation_data = conn.execute(
            """
            SELECT c.*, s.first_name, s.last_name, s.email, s.phone, s.membership_level,
                   s.parent_first, s.parent_last
            FROM cancellations c
            JOIN students s ON c.student_id = s.id
            WHERE c.id = ?
            """,
            (cancellation_id,),
        ).fetchone()

        if not cancellation_data:
            return jsonify({"success": False, "message": "Cancellation not found"})

        # Convert Row object to dict for easier access
        cancellation_dict = dict(cancellation_data)

        # Prepare student and cancellation data
        student_dict = {
            "id": cancellation_dict["student_id"],
            "first_name": cancellation_dict["first_name"],
            "last_name": cancellation_dict["last_name"],
            "email": cancellation_dict["email"],
            "phone": cancellation_dict["phone"],
            "membership_level": cancellation_dict["membership_level"],
            "parent_first": cancellation_dict["parent_first"],
            "parent_last": cancellation_dict["parent_last"],
        }

        # Get lesson datetime for policy check
        lesson_datetime = parse_lesson_datetime(
            cancellation_dict["lesson_date"], cancellation_dict["lesson_time"]
        )

        # Update database based on action
        if action == "approve_policy":
            # Use the unified will_be_charged function for both student and manager submissions
            # The key difference: managers specify submission_datetime via manual_submission_date
            
            pacific_tz = pytz.timezone("America/Los_Angeles")
            
            # Determine which submission datetime to use
            if cancellation_dict["is_manager_submission"] and cancellation_dict["manual_submission_date"]:
                # Manager submission: use the policy submission date they specified
                try:
                    manual_date_str = cancellation_dict["manual_submission_date"].strip()
                    # Handle both "YYYY-MM-DD HH:MM" and "YYYY-MM-DD HH:MM:SS" formats
                    if manual_date_str.count(':') == 1:  # HH:MM format (missing seconds)
                        submission_datetime = datetime.strptime(manual_date_str + ":00", "%Y-%m-%d %H:%M:%S")
                    else:  # HH:MM:SS format
                        submission_datetime = datetime.strptime(manual_date_str, "%Y-%m-%d %H:%M:%S")
                    
                    if submission_datetime.tzinfo is None:
                        submission_datetime = pacific_tz.localize(submission_datetime)
                    print(f"DEBUG: Successfully parsed manual_submission_date: {manual_date_str} -> {submission_datetime}")
                except Exception as parse_error:
                    print(f"DEBUG: Failed to parse manual_submission_date '{manual_date_str}': {parse_error}")
                    submission_datetime = None
            else:
                # Student submission: use when they actually submitted
                try:
                    submission_datetime = datetime.strptime(
                        cancellation_dict["created_at"], "%Y-%m-%d %H:%M:%S"
                    )
                    if submission_datetime.tzinfo is None:
                        submission_datetime = pacific_tz.localize(submission_datetime)
                except:
                    submission_datetime = None
            
            # Call the unified will_be_charged function.
            # We pass exclude_cancellation_id so this cancellation does NOT count
            # itself toward the student's monthly free limit or Welcome Package
            # lifetime check. Without this, a Bronze student's only pending free
            # cancellation would count itself (1 >= 1) and be flipped to charged
            # on Process.
            should_be_charged, charge_reason = will_be_charged(
                student_dict, 
                lesson_datetime,
                submission_datetime,
                exclude_cancellation_id=cancellation_id,
            )
            
            print(f"Should be charged: {should_be_charged}")
            print(f"Charge reason: {charge_reason}")
            print(f"Manager email: {session.get('user_email', 'Unknown Manager')}")

            if should_be_charged:
                # Process as charged according to policy
                print(f"Executing UPDATE for CHARGED cancellation...")
                conn.execute(
                    """UPDATE cancellations 
                       SET charged = 1, status = 'charged', manager_notes = ?, 
                           approved_by = ?, updated_at = ? 
                       WHERE id = ?""",
                    (
                        f"Processed according to policy: {charge_reason}",
                        session.get("user_email", "Unknown Manager"),
                        toronto_now().strftime("%Y-%m-%d %H:%M:%S"),
                        cancellation_id,
                    ),
                )
                print(f"UPDATE executed for CHARGED")
                updated_cancellation = dict(cancellation_dict)
                updated_cancellation.update(
                    {
                        "charged": 1,
                        "status": "charged",
                        "manager_notes": f"Processed according to policy: {charge_reason}",
                        "approved_by": session.get("user_email", "Unknown Manager"),
                    }
                )
                log_message = f"Cancellation {cancellation_id} charged according to policy: {charge_reason}"
            else:
                # Process as free according to policy
                print(f"Executing UPDATE for FREE cancellation...")
                
                # If Welcome Package, mark the free as used
                if cancellation_dict["membership_level"] == "Welcome Package":
                    mark_welcome_free_used(cancellation_dict["student_id"])
                
                conn.execute(
                    """UPDATE cancellations 
                       SET status = 'approved', charged = 0, manager_notes = ?, 
                           approved_by = ?, updated_at = ? 
                       WHERE id = ?""",
                    (
                        f"Processed according to policy: {charge_reason}",
                        session.get("user_email", "Unknown Manager"),
                        toronto_now().strftime("%Y-%m-%d %H:%M:%S"),
                        cancellation_id,
                    ),
                )
                print(f"UPDATE executed for FREE")
                updated_cancellation = dict(cancellation_dict)
                updated_cancellation.update(
                    {
                        "charged": 0,
                        "status": "approved",
                        "manager_notes": f"Processed according to policy: {charge_reason}",
                        "approved_by": session.get("user_email", "Unknown Manager"),
                    }
                )
                log_message = f"Cancellation {cancellation_id} approved as free according to policy: {charge_reason}"

        elif action == "force_free":
            # Force as free (override policy)
            conn.execute(
                """UPDATE cancellations 
                   SET excluded = 0, status = 'approved', charged = 0, manager_notes = ?, 
                       is_override = 1, approved_by = ?, updated_at = ? 
                   WHERE id = ?""",
                (
                    f"Manager Override (Force Free): {reason}",
                    session.get("user_email", "Unknown Manager"),
                    toronto_now().strftime("%Y-%m-%d %H:%M:%S"),
                    cancellation_id,
                ),
            )
            updated_cancellation = dict(cancellation_dict)
            updated_cancellation.update(
                {
                    "charged": 0,
                    "status": "approved",
                    "is_override": 1,
                    "manager_notes": f"Manager Override (Force Free): {reason}",
                    "approved_by": session.get("user_email", "Unknown Manager"),
                }
            )
            log_message = (
                f"Cancellation {cancellation_id} forced to free (Override: {reason})"
            )

        elif action == "force_charge":
            # Force as charged (override policy)
            conn.execute(
                """UPDATE cancellations 
                   SET excluded = 0, charged = 1, status = 'charged', manager_notes = ?, 
                       is_override = 1, approved_by = ?, updated_at = ? 
                   WHERE id = ?""",
                (
                    f"Manager Override (Force Charge): {reason}",
                    session.get("user_email", "Unknown Manager"),
                    toronto_now().strftime("%Y-%m-%d %H:%M:%S"),
                    cancellation_id,
                ),
            )
            updated_cancellation = dict(cancellation_dict)
            updated_cancellation.update(
                {
                    "charged": 1,
                    "status": "charged",
                    "is_override": 1,
                    "manager_notes": f"Manager Override (Force Charge): {reason}",
                    "approved_by": session.get("user_email", "Unknown Manager"),
                }
            )
            log_message = (
                f"Cancellation {cancellation_id} forced to charge (Override: {reason})"
            )

        else:
            return jsonify({"success": False, "message": "Invalid action"})

        conn.commit()

        # SEND EMAIL NOTIFICATIONS - Only for force actions (overrides)
        if action in ["force_free", "force_charge"]:
            # Only send emails if not suppressed
            if not suppress_student_email or not suppress_manager_email:
                email_results = send_override_notification_emails(
                    student_dict,
                    updated_cancellation,
                    action,
                    reason,
                    session.get("user_email", "Unknown Manager"),
                    suppress_student_email=suppress_student_email,
                    suppress_manager_email=suppress_manager_email
                )
            else:
                # Both emails suppressed
                email_results = {
                    "client_email": {
                        "success": True,
                        "message": "Suppressed by manager",
                    },
                    "manager_notification": {
                        "success": True,
                        "message": "Suppressed by manager",
                    },
                }
        else:
            # For approve_policy, no override emails needed (normal workflow)
            email_results = {
                "client_email": {
                    "success": True,
                    "message": "No override email needed",
                },
                "manager_notification": {
                    "success": True,
                    "message": "No override email needed",
                },
            }

        # CRITICAL: Commit the database changes before closing
        print(f"Committing changes to database...")
        conn.commit()
        print(f"Changes committed successfully!")
        conn.close()

        # Log the action
        log_action("cancellation_processed", log_message)

        print(f"Returning success response")
        print(f"===================================\n")

        return jsonify(
            {
                "success": True,
                "message": f"Cancellation processed successfully",
                "email_summary": email_results,
            }
        )

    except Exception as e:
        conn.close()
        return jsonify({"success": False, "message": str(e)})

        # Log email results
        if email_results["client_email"].get("success"):
            log_action(
                "override_client_email",
                f"Override confirmation sent to {student_dict['email']}",
            )
        else:
            log_action(
                "override_client_email_failed",
                f"Failed to send override confirmation: {email_results['client_email'].get('message', 'Unknown error')}",
            )

        if email_results["manager_notification"].get("success"):
            log_action(
                "override_manager_email", f"Override notification sent to managers"
            )
        else:
            log_action(
                "override_manager_email_failed",
                f"Failed to send override notification: {email_results['manager_notification'].get('message', 'Unknown error')}",
            )

        # Return success with email status
        return jsonify(
            {
                "success": True,
                "message": f"Cancellation {action}d successfully",
                "email_status": {
                    "client_notified": email_results["client_email"].get(
                        "success", False
                    ),
                    "managers_notified": email_results["manager_notification"].get(
                        "success", False
                    ),
                },
            }
        )

    except Exception as e:
        conn.close()
        return jsonify({"success": False, "message": str(e)})


@app.route("/manager/api/cancellation/batch", methods=["POST"])
@login_required
@admin_required
def batch_process_cancellations():
    """Batch process multiple cancellations - UPDATED with new approve as policy functionality"""
    data = request.json
    action = data.get(
        "action"
    )  # 'approve_policy', 'force_free', 'charge', or 'exclude'
    cancellation_ids = data.get("cancellation_ids", [])
    reason = data.get("reason", "Batch processing")

    if not action or not cancellation_ids:
        return jsonify({"success": False, "message": "Missing required fields"})

    # Force actions require a reason
    if action in ["force_free"] and not reason.strip():
        return jsonify(
            {"success": False, "message": "Reason is required for force free action"}
        )

    conn = get_db()
    email_summary = {
        "client_emails_sent": 0,
        "client_emails_failed": 0,
        "manager_notifications_sent": 0,
    }

    try:
        print(f"\n=== BATCH PROCESSING DEBUG ===")
        print(f"Action: {action}")
        print(f"Number of cancellations: {len(cancellation_ids)}")

        processed_count = 0

        for cancellation_id in cancellation_ids:
            try:
                # Get cancellation and student data
                cancellation_data = conn.execute(
                    """
                    SELECT c.*, s.first_name, s.last_name, s.email, s.phone, s.membership_level,
                           s.parent_first, s.parent_last
                    FROM cancellations c
                    JOIN students s ON c.student_id = s.id
                    WHERE c.id = ?
                    """,
                    (cancellation_id,),
                ).fetchone()

                if not cancellation_data:
                    print(f"Cancellation {cancellation_id} not found, skipping...")
                    continue

                # Convert Row to dict for easier access
                cancellation_dict = dict(cancellation_data)

                print(f"Processing cancellation {cancellation_id}...")

                # Prepare data
                student_dict = {
                    "id": cancellation_dict["student_id"],
                    "first_name": cancellation_dict["first_name"],
                    "last_name": cancellation_dict["last_name"],
                    "email": cancellation_dict["email"],
                    "phone": cancellation_dict["phone"],
                    "membership_level": cancellation_dict["membership_level"],
                    "parent_first": cancellation_dict["parent_first"],
                    "parent_last": cancellation_dict["parent_last"],
                }

                # Get lesson datetime for policy check
                lesson_datetime = parse_lesson_datetime(
                    cancellation_dict["lesson_date"], cancellation_dict["lesson_time"]
                )

                # Update database based on action
                if action == "approve_policy":
                    # CRITICAL FIX: Calculate charge status based on actual policy, not stored value
                    # Get membership tier info
                    tier = get_membership_tier(cancellation_dict["membership_level"])
                    
                    # Calculate hours notice using manual_submission_date for manager submissions
                    try:
                        if cancellation_dict["is_manager_submission"] and cancellation_dict["manual_submission_date"]:
                            created_datetime = datetime.strptime(
                                cancellation_dict["manual_submission_date"], "%Y-%m-%d %H:%M:%S"
                            )
                        else:
                            created_datetime = datetime.strptime(
                                cancellation_dict["created_at"], "%Y-%m-%d %H:%M:%S"
                            )
                        
                        # Make datetimes timezone-aware
                        pacific_tz = pytz.timezone("America/Los_Angeles")
                        if lesson_datetime.tzinfo is None:
                            lesson_datetime = pacific_tz.localize(lesson_datetime)
                        if created_datetime.tzinfo is None:
                            created_datetime = pacific_tz.localize(created_datetime)
                        
                        # Use tier-aware deadline calculation (supports both
                        # relative "X hours before lesson" and absolute
                        # "<time> previous day" deadlines).
                        deadline_passed = is_after_deadline(
                            tier, lesson_datetime, created_datetime
                        )
                    except:
                        deadline_passed = False
                    
                    # Determine if should be charged based on policy
                    should_be_charged = False
                    charge_reason = "Processed according to policy"
                    
                    if deadline_passed:
                        should_be_charged = True
                        charge_reason = "Submitted after deadline"
                    elif cancellation_dict["membership_level"] == "Welcome Package":
                        # Welcome Package: count other free cancellations dynamically,
                        # excluding this one (same fix as the single-process path).
                        other_free_row = conn.execute(
                            """SELECT COUNT(*) as cnt FROM cancellations
                               WHERE student_id = ? AND charged = 0 AND excluded = 0 AND id != ?""",
                            (cancellation_dict["student_id"], cancellation_id),
                        ).fetchone()
                        other_free_count = other_free_row["cnt"] if other_free_row else 0
                        if other_free_count >= 1:
                            should_be_charged = True
                            charge_reason = "Welcome Package free cancellation already used"
                    else:
                        # Exclude the cancellation being processed so it doesn't
                        # count itself toward the student's monthly free limit.
                        monthly_count = get_monthly_cancellation_count(
                            cancellation_dict["student_id"],
                            exclude_cancellation_id=cancellation_id,
                        )
                        free_notices = tier["free_notices"] if tier else 1
                        if monthly_count >= free_notices:
                            should_be_charged = True
                            charge_reason = f"Monthly free cancellation limit exceeded ({monthly_count}/{free_notices} used)"

                    print(f"  FIXED should_be_charged={should_be_charged}")

                    if should_be_charged:
                        # Process as charged according to policy
                        conn.execute(
                            """UPDATE cancellations 
                               SET charged = 1, status = 'charged', manager_notes = ?, 
                                   approved_by = ?, updated_at = ? 
                               WHERE id = ?""",
                            (
                                f"Batch processed according to policy: {charge_reason}",
                                session.get("user_email", "Unknown Manager"),
                                toronto_now().strftime("%Y-%m-%d %H:%M:%S"),
                                cancellation_id,
                            ),
                        )
                        print(f"  -> Set to CHARGED")
                        updated_cancellation = dict(cancellation_dict)
                        updated_cancellation.update(
                            {
                                "charged": 1,
                                "status": "charged",
                                "approved_by": session.get(
                                    "user_email", "Unknown Manager"
                                ),
                            }
                        )
                    else:
                        # Process as free according to policy
                        
                        # If Welcome Package, mark the free as used
                        if cancellation_dict["membership_level"] == "Welcome Package":
                            mark_welcome_free_used(cancellation_dict["student_id"])
                        
                        conn.execute(
                            """UPDATE cancellations 
                               SET status = 'approved', charged = 0, manager_notes = ?, 
                                   approved_by = ?, updated_at = ? 
                               WHERE id = ?""",
                            (
                                f"Batch processed according to policy: {charge_reason}",
                                session.get("user_email", "Unknown Manager"),
                                toronto_now().strftime("%Y-%m-%d %H:%M:%S"),
                                cancellation_id,
                            ),
                        )
                        print(f"  -> Set to APPROVED (free)")
                        updated_cancellation = dict(cancellation_dict)
                        updated_cancellation.update(
                            {
                                "charged": 0,
                                "status": "approved",
                                "approved_by": session.get(
                                    "user_email", "Unknown Manager"
                                ),
                            }
                        )

                elif action == "force_free":
                    conn.execute(
                        """UPDATE cancellations 
                           SET status = 'approved', charged = 0, manager_notes = ?, 
                               is_override = 1, approved_by = ?, updated_at = ? 
                           WHERE id = ?""",
                        (
                            f"Batch force free: {reason}",
                            session.get("user_email", "Unknown Manager"),
                            toronto_now().strftime("%Y-%m-%d %H:%M:%S"),
                            cancellation_id,
                        ),
                    )
                    print(f"  -> Forced to FREE (override)")
                    updated_cancellation = dict(cancellation_dict)
                    updated_cancellation.update(
                        {
                            "charged": 0,
                            "status": "approved",
                            "is_override": 1,
                            "approved_by": session.get("user_email", "Unknown Manager"),
                        }
                    )

                elif action == "charge":
                    conn.execute(
                        """UPDATE cancellations 
                           SET charged = 1, status = 'charged', manager_notes = ?, 
                               is_override = 1, approved_by = ?, updated_at = ? 
                           WHERE id = ?""",
                        (
                            f"Batch charge: {reason}",
                            session.get("user_email", "Unknown Manager"),
                            toronto_now().strftime("%Y-%m-%d %H:%M:%S"),
                            cancellation_id,
                        ),
                    )
                    print(f"  -> Forced to CHARGED (override)")
                    updated_cancellation = dict(cancellation_dict)
                    updated_cancellation.update(
                        {
                            "charged": 1,
                            "status": "charged",
                            "is_override": 1,
                            "approved_by": session.get("user_email", "Unknown Manager"),
                        }
                    )

                elif action == "exclude":
                    conn.execute(
                        """UPDATE cancellations 
                           SET excluded = 1, exclusion_reason = ?, approved_by = ?, 
                               is_override = 1, updated_at = ? 
                           WHERE id = ?""",
                        (
                            reason,
                            session["user_email"],
                            toronto_now().strftime("%Y-%m-%d %H:%M:%S"),
                            cancellation_id,
                        ),
                    )
                    print(f"  -> EXCLUDED")
                    updated_cancellation = dict(cancellation_dict)
                    updated_cancellation.update({"excluded": 1, "is_override": 1})

                # Send email notifications for each processed cancellation
                # Only send override emails for force actions, not for approve_policy
                if action in [
                    "force_free",
                    "charge",
                ]:  # Don't send emails for approve_policy or exclusions
                    email_results = send_override_notification_emails(
                        student_dict,
                        updated_cancellation,
                        action,
                        f"Batch {action}: {reason}",
                        session.get("user_email", "Unknown Manager"),
                    )

                    # Track email results
                    if email_results["client_email"].get("success"):
                        email_summary["client_emails_sent"] += 1
                    else:
                        email_summary["client_emails_failed"] += 1

                processed_count += 1

            except Exception as e:
                print(f"ERROR processing cancellation {cancellation_id}: {str(e)}")
                import traceback

                traceback.print_exc()
                # Continue with next cancellation even if one fails
                continue

        # Commit all changes at once
        print(f"Committing {processed_count} updates...")
        conn.commit()
        print(f"Commit successful!")

        # Send single manager notification about batch operation
        email_summary["manager_notifications_sent"] = 1  # Simplified for batch

        log_action(
            "batch_processing",
            f"Batch {action}: {processed_count} cancellations ({reason})",
        )

        print(f"===================================\n")

        return jsonify(
            {
                "success": True,
                "processed": processed_count,
                "email_summary": email_summary,
            }
        )

    except Exception as e:
        print(f"CRITICAL ERROR in batch_process: {str(e)}")
        import traceback

        traceback.print_exc()
        conn.rollback()  # Roll back on error
        return jsonify({"success": False, "message": str(e)})
    finally:
        conn.close()


@app.route("/manager/api/cancellation/process-all-pending", methods=["POST"])
@login_required
@admin_required
def process_all_pending_cancellations():
    """Process all pending cancellations according to policy"""
    conn = get_db()

    try:
        # Get all pending cancellations
        pending_cancellations = conn.execute(
            """
            SELECT c.*, s.first_name, s.last_name, s.email, s.phone, s.membership_level,
                   s.parent_first, s.parent_last
            FROM cancellations c
            JOIN students s ON c.student_id = s.id
            WHERE c.status = 'pending' OR c.status IS NULL
            """
        ).fetchall()

        print(f"\n=== PROCESS ALL PENDING DEBUG ===")
        print(f"Found {len(pending_cancellations)} pending cancellations")

        processed_count = 0

        for cancellation_data in pending_cancellations:
            try:
                # Convert Row to dict for easier access
                cancellation_dict = dict(cancellation_data)

                # Prepare student data
                student_dict = {
                    "id": cancellation_dict["student_id"],
                    "first_name": cancellation_dict["first_name"],
                    "last_name": cancellation_dict["last_name"],
                    "email": cancellation_dict["email"],
                    "phone": cancellation_dict["phone"],
                    "membership_level": cancellation_dict["membership_level"],
                    "parent_first": cancellation_dict["parent_first"],
                    "parent_last": cancellation_dict["parent_last"],
                }

                # CRITICAL FIX: Use the STORED charge status that was calculated at submission time
                # DO NOT recalculate based on current time - that would be incorrect!
                # The 'charged' field was set correctly when the cancellation was submitted
                should_be_charged = bool(cancellation_dict["charged"])

                # Get charge reason from stored data or create default message
                charge_reason = (
                    cancellation_dict.get("manager_notes")
                    or "Processed according to membership policy"
                )

                print(
                    f"Processing cancellation {cancellation_dict['id']}: should_be_charged={should_be_charged}"
                )

                if should_be_charged:
                    # Process as charged according to policy
                    conn.execute(
                        """UPDATE cancellations 
                           SET charged = 1, status = 'charged', manager_notes = ?, 
                               approved_by = ?, updated_at = ? 
                           WHERE id = ?""",
                        (
                            f"Processed according to policy: {charge_reason}",
                            session.get("user_email", "Unknown Manager"),
                            toronto_now().strftime("%Y-%m-%d %H:%M:%S"),
                            cancellation_dict["id"],
                        ),
                    )
                    print(f"  -> Set to CHARGED")
                else:
                    # Process as free according to policy
                    conn.execute(
                        """UPDATE cancellations 
                           SET status = 'approved', charged = 0, manager_notes = ?, 
                               approved_by = ?, updated_at = ? 
                           WHERE id = ?""",
                        (
                            f"Processed according to policy: {charge_reason}",
                            session.get("user_email", "Unknown Manager"),
                            toronto_now().strftime("%Y-%m-%d %H:%M:%S"),
                            cancellation_dict["id"],
                        ),
                    )
                    print(f"  -> Set to APPROVED (free)")

                processed_count += 1
            except Exception as e:
                print(
                    f"ERROR processing cancellation {cancellation_data['id']}: {str(e)}"
                )
                import traceback

                traceback.print_exc()
                # Continue with next cancellation even if one fails
                continue

        # Commit all changes at once
        print(f"Committing {processed_count} updates...")
        conn.commit()
        print(f"Commit successful!")

        log_action(
            "process_all_pending",
            f"Processed {processed_count} pending cancellations according to policy",
        )

        print(f"===================================\n")

        return jsonify(
            {
                "success": True,
                "processed": processed_count,
                "message": f"Processed {processed_count} pending cancellations according to policy",
            }
        )

    except Exception as e:
        print(f"CRITICAL ERROR in process_all_pending: {str(e)}")
        import traceback

        traceback.print_exc()
        conn.rollback()  # Roll back on error
        return jsonify({"success": False, "message": str(e)})
    finally:
        conn.close()


@app.route("/manager/api/cancellation/revert", methods=["POST"])
@login_required
@admin_required
def revert_cancellation():
    """Revert cancellation back to pending - UPDATED to clear override flags"""
    data = request.json
    cancellation_id = data.get("cancellation_id")

    if not cancellation_id:
        return jsonify({"success": False, "message": "Missing cancellation ID"})

    conn = get_db()

    try:
        conn.execute(
            """UPDATE cancellations 
               SET status = 'pending', charged = 0, excluded = 0, 
                   is_override = 0, manager_notes = ?, updated_at = ? 
               WHERE id = ?""",
            (
                f"Reverted by {session['user_email']} on {toronto_now().strftime('%Y-%m-%d %H:%M')}",
                toronto_now().strftime("%Y-%m-%d %H:%M:%S"),
                cancellation_id,
            ),
        )
        conn.commit()
        conn.close()

        log_action(
            "cancellation_reverted",
            f"Cancellation {cancellation_id} reverted to pending",
        )
        return jsonify({"success": True})

    except Exception as e:
        conn.close()
        return jsonify({"success": False, "message": str(e)})


@app.route("/manager/api/cancellation/note", methods=["POST"])
@login_required
@admin_required
def add_cancellation_note():
    """Add note to cancellation - UNCHANGED"""
    data = request.json
    cancellation_id = data.get("cancellation_id")
    note = data.get("note")

    if not cancellation_id or not note:
        return jsonify({"success": False, "message": "Missing required fields"})

    conn = get_db()

    try:
        # Get existing notes
        existing = conn.execute(
            "SELECT manager_notes FROM cancellations WHERE id = ?", (cancellation_id,)
        ).fetchone()

        existing_notes = (
            existing["manager_notes"] if existing and existing["manager_notes"] else ""
        )
        new_notes = f"{existing_notes}\n[{toronto_now().strftime('%Y-%m-%d %H:%M')} - {session['user_email']}]: {note}".strip()

        conn.execute(
            "UPDATE cancellations SET manager_notes = ?, updated_at = ? WHERE id = ?",
            (new_notes, toronto_now().strftime("%Y-%m-%d %H:%M:%S"), cancellation_id),
        )
        conn.commit()
        conn.close()

        log_action("note_added", f"Note added to cancellation {cancellation_id}")
        return jsonify({"success": True})

    except Exception as e:
        conn.close()
        return jsonify({"success": False, "message": str(e)})


@app.route("/manager/api/cancellation/process-all-pending", methods=["POST"])
@login_required
@admin_required
def process_all_pending():
    """Auto-process all pending cancellations according to policy - UPDATED"""
    conn = get_db()

    try:
        # Get all pending cancellations with student info
        pending = conn.execute(
            """
            SELECT c.*, s.membership_level 
            FROM cancellations c
            JOIN students s ON c.student_id = s.id
            WHERE c.status = 'pending' OR c.status IS NULL
        """
        ).fetchall()

        processed_count = 0

        for cancellation in pending:
            # Get membership tier info
            tier = get_membership_tier(cancellation["membership_level"])
            if not tier:
                continue

            # Check if within deadline using existing deadline_passed field or calculate it
            deadline_passed = cancellation.get("deadline_passed", 0)

            if not deadline_passed:
                # Calculate if not already set
                try:
                    lesson_datetime = datetime.strptime(
                        f"{cancellation['lesson_date']} {cancellation['lesson_time']}",
                        "%Y-%m-%d %H:%M:%S",
                    )
                    # Use localize_datetime to handle both formats
                    created_datetime = localize_datetime(cancellation["created_at"])
                    if created_datetime is None:
                        created_datetime = toronto_now()

                    # Use tier-aware deadline calculation so we honor absolute
                    # "<time> previous day" deadlines for Bronze/Silver/etc.
                    deadline_passed = is_after_deadline(
                        tier, lesson_datetime, created_datetime
                    )
                except:
                    deadline_passed = False

            # Check membership level and apply appropriate free cancellation logic
            student = conn.execute(
                "SELECT membership_level, welcome_free_used FROM students WHERE id = ?",
                (cancellation["student_id"],)
            ).fetchone()
            
            # For Welcome Package, check lifetime usage dynamically (count other free
            # cancellations) instead of relying solely on the welcome_free_used flag,
            # which may have been set by this same still-pending cancellation.
            if student and student["membership_level"] == "Welcome Package":
                other_free_row = conn.execute(
                    """SELECT COUNT(*) as cnt FROM cancellations
                       WHERE student_id = ? AND charged = 0 AND excluded = 0 AND id != ?""",
                    (cancellation["student_id"], cancellation["id"]),
                ).fetchone()
                other_free_count = other_free_row["cnt"] if other_free_row else 0
                welcome_free_allowed = (other_free_count == 0)
                monthly_count = 0  # Don't limit by month for Welcome
            else:
                # For other packages, use monthly count, excluding the cancellation
                # being processed so it doesn't count itself toward its own limit.
                monthly_count = get_monthly_cancellation_count(
                    cancellation["student_id"],
                    exclude_cancellation_id=cancellation["id"],
                )
                welcome_free_allowed = False

            # Determine if should be charged
            should_charge = False
            charge_reason = ""

            if deadline_passed:
                should_charge = True
                charge_reason = "Submitted after deadline"
            elif student and student["membership_level"] == "Welcome Package":
                # Welcome Package: charge if not within deadline AND free already used
                if not welcome_free_allowed:
                    should_charge = True
                    charge_reason = "Welcome Package free cancellation already used"
            elif monthly_count >= tier["free_notices"]:
                should_charge = True
                charge_reason = "Monthly free cancellation limit exceeded"

            # Update cancellation - NOT marked as override since this is automatic
            if should_charge:
                conn.execute(
                    """UPDATE cancellations 
                       SET charged = 1, status = 'charged', manager_notes = ?, 
                           deadline_passed = ?, updated_at = ? 
                       WHERE id = ?""",
                    (
                        f"Auto-processed: {charge_reason}",
                        (
                            1
                            if deadline_passed
                            else cancellation.get("deadline_passed", 0)
                        ),
                        toronto_now().strftime("%Y-%m-%d %H:%M:%S"),
                        cancellation["id"],
                    ),
                )
            else:
                # Mark as free - if Welcome Package, mark the free cancellation as used
                if student and student["membership_level"] == "Welcome Package":
                    mark_welcome_free_used(cancellation["student_id"])
                
                conn.execute(
                    """UPDATE cancellations 
                       SET status = 'approved', charged = 0, manager_notes = ?, 
                           deadline_passed = ?, updated_at = ? 
                       WHERE id = ?""",
                    (
                        "Auto-processed: Within policy",
                        (
                            1
                            if deadline_passed
                            else cancellation.get("deadline_passed", 0)
                        ),
                        toronto_now().strftime("%Y-%m-%d %H:%M:%S"),
                        cancellation["id"],
                    ),
                )

            processed_count += 1

        conn.commit()
        conn.close()

        log_action(
            "auto_process_all",
            f"Auto-processed {processed_count} pending cancellations",
        )
        return jsonify({"success": True, "processed": processed_count})

    except Exception as e:
        conn.close()
        return jsonify({"success": False, "message": str(e)})


@app.route("/manager/api/student/<int:student_id>/details", methods=["GET"])
@login_required
@admin_required
def get_student_details(student_id):
    """Get detailed student information including cancellation statistics"""
    conn = get_db()

    try:
        # Get student info
        student = conn.execute(
            "SELECT * FROM students WHERE id = ?", (student_id,)
        ).fetchone()

        if not student:
            return jsonify({"success": False, "message": "Student not found"})

        # Get cancellation statistics with more detail
        stats = conn.execute(
            """
            SELECT 
                COUNT(*) as total_cancellations,
                SUM(CASE WHEN strftime('%Y-%m', 
                    CASE 
                        WHEN is_manager_submission = 1 AND manual_submission_date IS NOT NULL
                        THEN manual_submission_date
                        ELSE created_at
                    END
                ) = strftime('%Y-%m', 'now') THEN 1 ELSE 0 END) as this_month,
                SUM(CASE WHEN charged = 0 AND excluded = 0 AND strftime('%Y-%m', 
                    CASE 
                        WHEN is_manager_submission = 1 AND manual_submission_date IS NOT NULL
                        THEN manual_submission_date
                        ELSE created_at
                    END
                ) = strftime('%Y-%m', 'now') THEN 1 ELSE 0 END) as free_used,
                SUM(CASE WHEN charged = 1 THEN 1 ELSE 0 END) as total_charged,
                SUM(CASE WHEN excluded = 1 THEN 1 ELSE 0 END) as total_excluded,
                SUM(CASE WHEN charged = 1 AND strftime('%Y-%m', 
                    CASE 
                        WHEN is_manager_submission = 1 AND manual_submission_date IS NOT NULL
                        THEN manual_submission_date
                        ELSE created_at
                    END
                ) = strftime('%Y-%m', 'now') THEN 1 ELSE 0 END) as charged_this_month,
                SUM(CASE WHEN excluded = 1 AND strftime('%Y-%m', 
                    CASE 
                        WHEN is_manager_submission = 1 AND manual_submission_date IS NOT NULL
                        THEN manual_submission_date
                        ELSE created_at
                    END
                ) = strftime('%Y-%m', 'now') THEN 1 ELSE 0 END) as excluded_this_month
            FROM cancellations 
            WHERE student_id = ?
        """,
            (student_id,),
        ).fetchone()

        # Get membership tier info
        tier = get_membership_tier(student["membership_level"])
        
        # Calculate limits based on membership type
        if student["membership_level"] == "Welcome Package":
            # For Welcome Package, show lifetime usage
            monthly_limit = 1
            free_used = 1 if student["welcome_free_used"] else 0
            remaining = 0 if student["welcome_free_used"] else 1
            limit_type = "lifetime"
        else:
            # For other packages, show monthly usage
            monthly_limit = tier["free_notices"] if tier else 1
            free_used = stats["free_used"] or 0
            remaining = max(0, monthly_limit - free_used)
            limit_type = "monthly"

        # Get recent cancellations
        recent_cancellations = conn.execute(
            """
            SELECT lesson_date, lesson_time, created_at, charged, excluded, status
            FROM cancellations 
            WHERE student_id = ? 
            ORDER BY created_at DESC 
            LIMIT 5
        """,
            (student_id,),
        ).fetchall()

        conn.close()

        return jsonify(
            {
                "success": True,
                "student": dict(student),
                "stats": {
                    "total_cancellations": stats["total_cancellations"] or 0,
                    "this_month": stats["this_month"] or 0,
                    "free_used": free_used,
                    "monthly_limit": monthly_limit,
                    "remaining": remaining,
                    "limit_type": limit_type,
                    "total_charged": stats["total_charged"] or 0,
                    "total_excluded": stats["total_excluded"] or 0,
                    "charged_this_month": stats["charged_this_month"] or 0,
                    "excluded_this_month": stats["excluded_this_month"] or 0,
                },
                "recent_cancellations": [dict(c) for c in recent_cancellations],
            }
        )

    except Exception as e:
        conn.close()
        return jsonify({"success": False, "message": str(e)})


@app.route("/manager/api/cancellation/<int:cancellation_id>/details", methods=["GET"])
@login_required
@admin_required
def get_cancellation_details(cancellation_id):
    """Get detailed cancellation information - UPDATED with new fields"""
    conn = get_db()

    try:
        # Get cancellation with student info - UPDATED query
        cancellation = conn.execute(
            """
            SELECT c.*, s.first_name, s.last_name, s.parent_first, s.parent_last, 
                   s.email, s.phone, s.membership_level, s.created_at as student_created_at
            FROM cancellations c
            JOIN students s ON c.student_id = s.id
            WHERE c.id = ?
        """,
            (cancellation_id,),
        ).fetchone()

        if not cancellation:
            return jsonify({"success": False, "message": "Cancellation not found"})

        # Convert Row to dict to use .get() method
        cancellation_dict = dict(cancellation)

        # Get membership tier info for policy analysis
        tier = get_membership_tier(cancellation_dict["membership_level"])

        # Calculate policy compliance
        try:
            lesson_datetime = datetime.strptime(
                f"{cancellation_dict['lesson_date']} {cancellation_dict['lesson_time']}",
                "%Y-%m-%d %H:%M:%S",
            )
        except ValueError:
            # Handle time without seconds
            lesson_datetime = datetime.strptime(
                f"{cancellation_dict['lesson_date']} {cancellation_dict['lesson_time']}:00",
                "%Y-%m-%d %H:%M:%S",
            )

        # Use localize_datetime to handle both old and new timestamp formats
        created_datetime = localize_datetime(cancellation_dict["created_at"])
        if created_datetime is None:
            created_datetime = toronto_now()

        # Ensure lesson_datetime is timezone-aware
        pacific_tz = pytz.timezone("America/Los_Angeles")
        if lesson_datetime.tzinfo is None:
            lesson_datetime = pacific_tz.localize(lesson_datetime)

        hours_notice = (lesson_datetime - created_datetime).total_seconds() / 3600

        # Get monthly usage - handle Welcome Package specially
        if cancellation_dict["membership_level"] == "Welcome Package":
            # For Welcome Package, show if the free cancellation has been used
            student_record = conn.execute(
                "SELECT welcome_free_used FROM students WHERE id = ?",
                (cancellation_dict["student_id"],)
            ).fetchone()
            welcome_free_used = bool(student_record["welcome_free_used"]) if student_record else False
            monthly_count = 1 if welcome_free_used else 0
            remaining_free = 0 if welcome_free_used else 1
        else:
            # For other packages, use normal monthly count
            monthly_count = get_monthly_cancellation_count(cancellation_dict["student_id"])
            remaining_free = max(0, (tier["free_notices"] if tier else 1) - monthly_count)

        # Parse sequential lessons
        sequential_lessons = []
        if cancellation_dict["sequential_lessons"]:
            try:
                sequential_lessons = eval(cancellation_dict["sequential_lessons"])
            except:
                sequential_lessons = []

        # Get action history from logs
        action_history = conn.execute(
            """
            SELECT action, details, created_at
            FROM system_logs
            WHERE details LIKE ?
            ORDER BY created_at DESC
            LIMIT 10
        """,
            (f"%{cancellation_id}%",),
        ).fetchall()

        conn.close()

        # Use deadline_passed from database if available, otherwise calculate
        within_deadline = True
        if cancellation_dict.get("deadline_passed") is not None:
            within_deadline = not bool(cancellation_dict["deadline_passed"])
        elif tier:
            within_deadline = hours_notice >= tier["deadline_hours"]
        
        # Determine policy result for Welcome Package
        if cancellation_dict["membership_level"] == "Welcome Package":
            policy_limit = 1 if not welcome_free_used else 0
            policy_result = (
                "Within policy"
                if within_deadline and not welcome_free_used
                else "Policy violation"
            )
        else:
            policy_limit = tier["free_notices"] if tier else 1
            policy_result = (
                "Within policy"
                if within_deadline
                and monthly_count < (tier["free_notices"] if tier else 1)
                else "Policy violation"
            )

        return jsonify(
            {
                "success": True,
                "cancellation": {
                    **cancellation_dict,
                    "sequential_lessons": sequential_lessons,
                },
                "student": {
                    "first_name": cancellation_dict["first_name"],
                    "last_name": cancellation_dict["last_name"],
                    "parent_first": cancellation_dict["parent_first"] or "",
                    "parent_last": cancellation_dict["parent_last"] or "",
                    "email": cancellation_dict["email"],
                    "phone": cancellation_dict["phone"] or "",
                    "membership_level": cancellation_dict["membership_level"],
                },
                "policy": {
                    "limit_type": "lifetime" if cancellation_dict["membership_level"] == "Welcome Package" else "monthly",
                    "monthly_limit": policy_limit,
                    "used_this_month": monthly_count,
                    "remaining": remaining_free,
                    "deadline_display": (
                        tier["deadline_display"] if tier else "6pm previous day"
                    ),
                    "hours_notice": round(hours_notice, 1),
                    "within_deadline": within_deadline,
                    "policy_result": policy_result,
                },
                "action_history": [dict(a) for a in action_history],
            }
        )

    except Exception as e:
        conn.close()
        print(f"Error in get_cancellation_details: {str(e)}")
        return jsonify({"success": False, "message": str(e)})


# Update the existing manager_cancellations route to handle student and view parameters properly
# Debug version of the manager_cancellations route
@app.route("/manager/cancellations")
@login_required
@admin_required
def manager_cancellations():
    """Manager cancellations page - CONTEXT-AWARE STATS VERSION"""
    # Get filter parameters
    filter_status = request.args.get("status", "")
    search = request.args.get("search", "")
    date_range = request.args.get("date_range", "month")
    membership = request.args.get("membership", "")
    sort_by = request.args.get("sort", "submit_date")
    student_id = request.args.get("student")
    date_from = request.args.get("date_from")
    date_to = request.args.get("date_to")
    approval_status = request.args.get("approval_status", "")
    submission_type = request.args.get("submission_type", "")

    # Check if this is an AJAX request
    is_ajax = request.headers.get("X-Requested-With") == "XMLHttpRequest"

    conn = get_db()

    # Build query based on filters
    where_clauses = []
    params = []

    # Add student filter
    if student_id:
        where_clauses.append("c.student_id = ?")
        params.append(student_id)

    # Status filters
    if filter_status == "pending":
        where_clauses.append("c.status = 'pending'")
    elif filter_status == "free":
        where_clauses.append("c.charged = 0 AND c.status = 'approved'")
    elif filter_status == "charged":
        where_clauses.append("c.charged = 1 AND c.excluded = 0")
    elif filter_status == "excluded":
        where_clauses.append("c.excluded = 1")
    elif filter_status == "note":
        where_clauses.append(
            """(
                (c.sequential_lessons IS NOT NULL AND c.sequential_lessons != '' AND c.sequential_lessons != '[]') 
                OR 
                (c.error_report IS NOT NULL AND c.error_report != '')
                OR
                (c.reschedule_requested = 1)
            )"""
        )
    elif filter_status == "deadline_passed":
        where_clauses.append(
            """
        (
            (s.membership_level = 'Gold' AND 
             datetime(c.lesson_date || ' ' || c.lesson_time) <= datetime(c.created_at, '+2 hours'))
            OR
            (s.membership_level != 'Gold' AND 
             datetime(c.lesson_date || ' 18:00:00', '-1 day') <= datetime(c.created_at))
        )
    """
        )
    elif filter_status == "override":
        where_clauses.append("c.is_override = 1")

    # Approval status filter
    if approval_status == "approved":
        where_clauses.append(
            "(c.status = 'approved' OR c.excluded = 1 OR c.is_override = 1)"
        )
    elif approval_status == "pending":
        where_clauses.append(
            "c.status = 'pending' AND c.excluded = 0 AND c.is_override = 0"
        )

    # Submission type filter (student vs manager submitted)
    if submission_type == "student":
        where_clauses.append("c.is_manager_submission = 0")
    elif submission_type == "manager":
        where_clauses.append("c.is_manager_submission = 1")

    # Search filter
    if search:
        where_clauses.append(
            "(s.first_name LIKE ? OR s.last_name LIKE ? OR s.email LIKE ?)"
        )
        search_param = f"%{search}%"
        params.extend([search_param, search_param, search_param])

    # CORRECTED: Use correct date field based on submission type
    # For manager submissions, use manual_submission_date (policy date)
    # For student submissions, use created_at (actual submission timestamp)
    date_field = """
        CASE 
            WHEN c.is_manager_submission = 1 AND c.manual_submission_date IS NOT NULL
            THEN c.manual_submission_date
            ELSE c.created_at
        END
    """

    # Date range filter - use Pacific (America/Los_Angeles) dates rather than
    # SQLite's DATE('now')/strftime(..., 'now'), which return UTC. Between roughly
    # 4pm-midnight Pacific, UTC is already on the next calendar day, so the old
    # SQL would file today's submissions under tomorrow's UTC date — which is
    # what made the "Today" tab show 0 and "Yesterday" show today's submissions.
    pacific_today_date = toronto_now().date()
    pacific_today_str = pacific_today_date.strftime("%Y-%m-%d")
    pacific_yesterday_str = (pacific_today_date - timedelta(days=1)).strftime("%Y-%m-%d")
    pacific_7days_str = (pacific_today_date - timedelta(days=7)).strftime("%Y-%m-%d")
    pacific_30days_str = (pacific_today_date - timedelta(days=30)).strftime("%Y-%m-%d")

    if date_range == "today":
        where_clauses.append(f"DATE({date_field}) = ?")
        params.append(pacific_today_str)
    elif date_range == "yesterday":
        where_clauses.append(f"DATE({date_field}) = ?")
        params.append(pacific_yesterday_str)
    elif date_range == "7days":
        where_clauses.append(f"DATE({date_field}) >= ?")
        params.append(pacific_7days_str)
    elif date_range == "month":
        where_clauses.append(f"DATE({date_field}) >= ?")
        params.append(pacific_30days_str)
    elif date_range == "all":
        pass  # No date filter
    elif date_range == "custom" and date_from and date_to:
        where_clauses.append(f"DATE({date_field}) BETWEEN ? AND ?")
        params.extend([date_from, date_to])

    # Membership filter
    if membership:
        where_clauses.append("s.membership_level = ?")
        params.append(membership)

    where_sql = " AND ".join(where_clauses) if where_clauses else "1=1"

    # Sort order - CORRECTED to use correct date field for manager submissions
    if sort_by == "submit_date":
        order_by = f"{date_field} DESC"
    elif sort_by == "lesson_date":
        order_by = "c.lesson_date DESC"
    elif sort_by == "student":
        order_by = "s.last_name, s.first_name"
    elif sort_by == "status":
        order_by = "c.status, c.charged, c.excluded"
    else:
        order_by = f"{date_field} DESC"

    # Get cancellations with student info
    cancellations_raw = conn.execute(
        f"""
        SELECT c.*, s.first_name, s.last_name, s.email, s.membership_level
        FROM cancellations c
        JOIN students s ON c.student_id = s.id
        WHERE {where_sql}
        ORDER BY {order_by}
        LIMIT 100
        """,
        params,
    ).fetchall()

    # Process cancellations data
    cancellations = []
    for row in cancellations_raw:
        cancellation = process_cancellation_dates(row)

        # Add computed fields
        cancellation["student_name"] = (
            f"{cancellation['first_name']} {cancellation['last_name']}"
        )

        # Status class for CSS
        is_approved = (
            cancellation.get("approved_by") is not None
            and cancellation.get("approved_by") != ""
        )

        # Priority 1: Special markers
        if cancellation.get("cancellation_note"):
            cancellation["status_class"] = "note"
        elif cancellation.get("is_override"):
            cancellation["status_class"] = "override"
        elif cancellation["excluded"]:
            cancellation["status_class"] = "excluded"
        # Priority 2: Check if approved
        elif not is_approved:
            if cancellation.get("deadline_passed"):
                cancellation["status_class"] = "deadline-passed"
            else:
                cancellation["status_class"] = "pending"
        # Priority 3: Approved cancellations
        elif cancellation["charged"]:
            cancellation["status_class"] = "charged"
        else:
            cancellation["status_class"] = "free"

        # Calculate deadline status - CORRECTED: Use policy date for manager submissions
        tier = get_membership_tier(cancellation["membership_level"])
        if tier:
            lesson_datetime = datetime.combine(
                cancellation["lesson_date"], cancellation["lesson_time"]
            )
            pacific_tz = pytz.timezone("America/Los_Angeles")
            lesson_datetime = pacific_tz.localize(lesson_datetime)

            # CORRECTED: Use the correct submission date based on submission type
            submission_datetime = None
            if cancellation.get("is_manager_submission") and cancellation.get("manual_submission_date"):
                # For manager submissions, use the policy submission date
                submission_datetime = cancellation.get("manual_submission_date")
                if isinstance(submission_datetime, str):
                    submission_datetime = localize_datetime(submission_datetime)
            
            if submission_datetime is None:
                # Fallback to actual submission timestamp for student submissions
                submission_datetime = localize_datetime(cancellation["created_at"])
            
            if submission_datetime is None:
                submission_datetime = toronto_now()

            hours_notice = (lesson_datetime - submission_datetime).total_seconds() / 3600
            cancellation["within_deadline"] = hours_notice >= tier["deadline_hours"]
        else:
            cancellation["within_deadline"] = True

        # Add usage statistics
        cancellation["used_this_month"] = get_monthly_cancellation_count(
            cancellation.get("student_id")
        )
        cancellation["monthly_limit"] = tier["free_notices"] if tier else 1

        # Additional fields
        cancellation["reschedule_requested"] = bool(
            cancellation.get("reschedule_requested")
        )
        cancellation["reschedule_preferences"] = cancellation.get(
            "reschedule_preferences", ""
        )
        cancellation["error_report"] = cancellation.get("error_report", "")
        cancellation["approved_by"] = cancellation.get("approved_by", "")

        # Add has_notes flag for Notes column - FIXED: Better sequential lessons detection
        sequential_lessons_json = cancellation.get("sequential_lessons", "")
        has_sequential_lessons = False
        if sequential_lessons_json:
            try:
                # Handle both JSON string and already-parsed list
                if isinstance(sequential_lessons_json, str):
                    if (
                        sequential_lessons_json
                        and sequential_lessons_json.strip() not in ["", "[]", "null"]
                    ):
                        import json

                        sequential_data = json.loads(sequential_lessons_json)
                        has_sequential_lessons = (
                            isinstance(sequential_data, list)
                            and len(sequential_data) > 0
                        )
                elif isinstance(sequential_lessons_json, list):
                    has_sequential_lessons = len(sequential_lessons_json) > 0
            except Exception as e:
                print(
                    f"Error parsing sequential_lessons for cancellation {cancellation.get('id')}: {e}"
                )
                has_sequential_lessons = False

        cancellation["has_notes"] = (
            has_sequential_lessons
            or bool(cancellation.get("error_report"))
            or bool(cancellation.get("reschedule_requested"))
            or bool(cancellation.get("cancellation_note"))
        )

        # Add formatted sequential lessons message
        if has_sequential_lessons:
            try:
                if isinstance(sequential_lessons_json, str):
                    sequential_data = (
                        json.loads(sequential_lessons_json)
                        if sequential_lessons_json.strip() not in ["", "[]", "null"]
                        else []
                    )
                else:
                    sequential_data = sequential_lessons_json

                cancellation["sequential_lessons_formatted"] = (
                    format_sequential_lessons(
                        sequential_data, cancellation.get("lesson_date")
                    )
                )
            except:
                cancellation["sequential_lessons_formatted"] = (
                    "None"
                )
        else:
            cancellation["sequential_lessons_formatted"] = (
                "None"
            )

        # Add urgency flags
        if cancellation.get("created_at"):
            # Parse created_at to timezone-aware datetime
            created_at = localize_datetime(cancellation["created_at"])
            if created_at:
                hours_since = (toronto_now() - created_at).total_seconds() / 3600
                cancellation["is_recent"] = hours_since < 2
                cancellation["is_urgent"] = (
                    hours_since > 24 and cancellation.get("status") == "pending"
                )
            else:
                cancellation["is_recent"] = False
                cancellation["is_urgent"] = False
        else:
            cancellation["is_recent"] = False
            cancellation["is_urgent"] = False

        cancellations.append(cancellation)

    # ============================================================
    # CONTEXT-AWARE STATS CALCULATION
    # ============================================================

    # Determine context
    stats_context = "global"  # default

    if student_id:
        stats_context = "student"
    elif date_range == "today":
        stats_context = "today"
    elif date_range == "yesterday":
        stats_context = "yesterday"
    elif date_range in ["7days", "month", "custom"]:
        stats_context = "date_range"
    elif filter_status in ["free", "charged", "excluded", "override"]:
        stats_context = "status_filter"

    # Calculate context-appropriate stats
    if stats_context == "student":
        # Student-specific stats
        student_stats = conn.execute(
            """
            SELECT
                COUNT(*) as total,
                SUM(CASE WHEN strftime('%Y-%m', 
                    CASE 
                        WHEN c.is_manager_submission = 1 AND c.manual_submission_date IS NOT NULL
                        THEN c.manual_submission_date
                        ELSE c.created_at
                    END
                ) = strftime('%Y-%m', 'now') THEN 1 ELSE 0 END) as this_month,
                SUM(CASE WHEN c.charged = 0 AND c.excluded = 0 AND strftime('%Y-%m', 
                    CASE 
                        WHEN c.is_manager_submission = 1 AND c.manual_submission_date IS NOT NULL
                        THEN c.manual_submission_date
                        ELSE c.created_at
                    END
                ) = strftime('%Y-%m', 'now') THEN 1 ELSE 0 END) as free_used,
                SUM(CASE WHEN c.charged = 1 AND strftime('%Y-%m', 
                    CASE 
                        WHEN c.is_manager_submission = 1 AND c.manual_submission_date IS NOT NULL
                        THEN c.manual_submission_date
                        ELSE c.created_at
                    END
                ) = strftime('%Y-%m', 'now') THEN 1 ELSE 0 END) as charged,
                SUM(CASE WHEN c.excluded = 1 THEN 1 ELSE 0 END) as excluded
            FROM cancellations c
            WHERE c.student_id = ?
            """,
            (student_id,),
        ).fetchone()

        # Get student's membership tier for limit
        student_info = conn.execute(
            "SELECT membership_level FROM students WHERE id = ?", (student_id,)
        ).fetchone()
        tier = (
            get_membership_tier(student_info["membership_level"])
            if student_info
            else None
        )
        monthly_limit = tier["free_notices"] if tier else 1

        stats = {
            "box1_value": student_stats["total"] or 0,
            "box1_label": "Total Cancellations",
            "box1_icon": "calendar-check",
            "box1_color": "primary",
            "box2_value": student_stats["this_month"] or 0,
            "box2_label": "This Month",
            "box2_icon": "calendar-day",
            "box2_color": "info",
            "box3_value": f"{student_stats['free_used'] or 0}/{monthly_limit}",
            "box3_label": "Free Used",
            "box3_icon": "gift",
            "box3_color": "success",
            "box4_value": student_stats["charged"] or 0,
            "box4_label": "Charged",
            "box4_icon": "dollar-sign",
            "box4_color": "warning",
            "box5_value": max(0, monthly_limit - (student_stats["free_used"] or 0)),
            "box5_label": "Remaining Free",
            "box5_icon": "check-circle",
            "box5_color": "success",
        }

    elif stats_context == "today":
        # Today's breakdown - use Pacific date (not UTC) so this matches the
        # date_range filter the manager actually selected.
        pacific_today_str = toronto_now().date().strftime("%Y-%m-%d")
        today_stats = conn.execute(
            """
            SELECT
                COUNT(*) as total,
                SUM(CASE WHEN c.charged = 0 AND c.status = 'approved' THEN 1 ELSE 0 END) as free,
                SUM(CASE WHEN c.charged = 1 THEN 1 ELSE 0 END) as charged,
                SUM(CASE WHEN c.excluded = 1 THEN 1 ELSE 0 END) as excluded,
                SUM(CASE WHEN c.status = 'pending' THEN 1 ELSE 0 END) as pending
            FROM cancellations c
            JOIN students s ON c.student_id = s.id
            WHERE DATE(c.created_at) = ?
            """,
            (pacific_today_str,),
        ).fetchone()

        stats = {
            "box1_value": today_stats["total"] or 0,
            "box1_label": "Total Today",
            "box1_icon": "calendar-day",
            "box1_color": "primary",
            "box2_value": today_stats["free"] or 0,
            "box2_label": "Free Today",
            "box2_icon": "gift",
            "box2_color": "success",
            "box3_value": today_stats["charged"] or 0,
            "box3_label": "Charged Today",
            "box3_icon": "dollar-sign",
            "box3_color": "warning",
            "box4_value": today_stats["excluded"] or 0,
            "box4_label": "Excluded Today",
            "box4_icon": "shield-alt",
            "box4_color": "info",
            "box5_value": today_stats["pending"] or 0,
            "box5_label": "Pending Today",
            "box5_icon": "clock",
            "box5_color": "secondary",
        }

    elif stats_context == "yesterday":
        # Yesterday's breakdown - use Pacific date (not UTC).
        pacific_yesterday_str = (toronto_now().date() - timedelta(days=1)).strftime("%Y-%m-%d")
        yesterday_stats = conn.execute(
            """
            SELECT
                COUNT(*) as total,
                SUM(CASE WHEN c.charged = 0 AND c.status = 'approved' THEN 1 ELSE 0 END) as free,
                SUM(CASE WHEN c.charged = 1 THEN 1 ELSE 0 END) as charged,
                SUM(CASE WHEN c.excluded = 1 THEN 1 ELSE 0 END) as excluded,
                SUM(CASE WHEN c.status = 'pending' THEN 1 ELSE 0 END) as pending
            FROM cancellations c
            JOIN students s ON c.student_id = s.id
            WHERE DATE(c.created_at) = ?
            """,
            (pacific_yesterday_str,),
        ).fetchone()

        stats = {
            "box1_value": yesterday_stats["total"] or 0,
            "box1_label": "Total Yesterday",
            "box1_icon": "calendar-alt",
            "box1_color": "primary",
            "box2_value": yesterday_stats["free"] or 0,
            "box2_label": "Free Yesterday",
            "box2_icon": "gift",
            "box2_color": "success",
            "box3_value": yesterday_stats["charged"] or 0,
            "box3_label": "Charged Yesterday",
            "box3_icon": "dollar-sign",
            "box3_color": "warning",
            "box4_value": yesterday_stats["excluded"] or 0,
            "box4_label": "Excluded Yesterday",
            "box4_icon": "shield-alt",
            "box4_color": "info",
            "box5_value": yesterday_stats["pending"] or 0,
            "box5_label": "Pending Yesterday",
            "box5_icon": "clock",
            "box5_color": "secondary",
        }

    elif stats_context == "date_range":
        # Date range breakdown
        range_stats = conn.execute(
            f"""
            SELECT
                COUNT(*) as total,
                SUM(CASE WHEN c.charged = 0 AND c.status = 'approved' THEN 1 ELSE 0 END) as free,
                SUM(CASE WHEN c.charged = 1 THEN 1 ELSE 0 END) as charged,
                SUM(CASE WHEN c.excluded = 1 THEN 1 ELSE 0 END) as excluded,
                SUM(CASE WHEN c.status = 'pending' THEN 1 ELSE 0 END) as pending
            FROM cancellations c
            JOIN students s ON c.student_id = s.id
            WHERE {where_sql}
            """,
            params,
        ).fetchone()

        # Determine label based on date range
        range_label_map = {
            "7days": "Last 7 Days",
            "month": "Last 30 Days",
            "custom": "Date Range",
        }
        range_label = range_label_map.get(date_range, "Selected Period")

        stats = {
            "box1_value": range_stats["total"] or 0,
            "box1_label": f"Total ({range_label})",
            "box1_icon": "calendar-check",
            "box1_color": "primary",
            "box2_value": range_stats["free"] or 0,
            "box2_label": f"Free ({range_label})",
            "box2_icon": "gift",
            "box2_color": "success",
            "box3_value": range_stats["charged"] or 0,
            "box3_label": f"Charged ({range_label})",
            "box3_icon": "dollar-sign",
            "box3_color": "warning",
            "box4_value": range_stats["excluded"] or 0,
            "box4_label": f"Excluded ({range_label})",
            "box4_icon": "shield-alt",
            "box4_color": "info",
            "box5_value": range_stats["pending"] or 0,
            "box5_label": f"Pending ({range_label})",
            "box5_icon": "clock",
            "box5_color": "secondary",
        }

    else:
        # Global overview (default for "all" and status filters).
        # Use Pacific (not UTC) for today/yesterday and for the current year-month
        # so these counters match what a Pacific-time user expects to see.
        _pac_today = toronto_now().date()
        _pac_today_str = _pac_today.strftime("%Y-%m-%d")
        _pac_yesterday_str = (_pac_today - timedelta(days=1)).strftime("%Y-%m-%d")
        _pac_year_month_str = _pac_today.strftime("%Y-%m")

        global_stats = conn.execute(
            """
            SELECT
                COUNT(*) as total,
                SUM(CASE WHEN DATE(c.created_at) = ? THEN 1 ELSE 0 END) as today,
                SUM(CASE WHEN DATE(c.created_at) = ? THEN 1 ELSE 0 END) as yesterday,
                SUM(CASE WHEN c.charged = 0 AND c.status = 'approved' AND strftime('%Y-%m', 
                    CASE 
                        WHEN c.is_manager_submission = 1 AND c.manual_submission_date IS NOT NULL
                        THEN c.manual_submission_date
                        ELSE c.created_at
                    END
                ) = ? THEN 1 ELSE 0 END) as free_month,
                SUM(CASE WHEN c.charged = 1 AND strftime('%Y-%m', 
                    CASE 
                        WHEN c.is_manager_submission = 1 AND c.manual_submission_date IS NOT NULL
                        THEN c.manual_submission_date
                        ELSE c.created_at
                    END
                ) = ? THEN 1 ELSE 0 END) as charged_month,
                SUM(CASE WHEN c.excluded = 1 AND strftime('%Y-%m', 
                    CASE 
                        WHEN c.is_manager_submission = 1 AND c.manual_submission_date IS NOT NULL
                        THEN c.manual_submission_date
                        ELSE c.created_at
                    END
                ) = ? THEN 1 ELSE 0 END) as excluded_month
            FROM cancellations c
            JOIN students s ON c.student_id = s.id
            """,
            (
                _pac_today_str,
                _pac_yesterday_str,
                _pac_year_month_str,
                _pac_year_month_str,
                _pac_year_month_str,
            ),
        ).fetchone()

        stats = {
            "box1_value": global_stats["today"] or 0,
            "box1_label": "Today",
            "box1_icon": "calendar-day",
            "box1_color": "info",
            "box2_value": global_stats["yesterday"] or 0,
            "box2_label": "Yesterday",
            "box2_icon": "calendar-alt",
            "box2_color": "secondary",
            "box3_value": global_stats["free_month"] or 0,
            "box3_label": "Free This Month",
            "box3_icon": "gift",
            "box3_color": "success",
            "box4_value": global_stats["charged_month"] or 0,
            "box4_label": "Charged This Month",
            "box4_icon": "dollar-sign",
            "box4_color": "warning",
            "box5_value": global_stats["excluded_month"] or 0,
            "box5_label": "Excluded This Month",
            "box5_icon": "shield-alt",
            "box5_color": "primary",
        }

        # If filtering by status, highlight the relevant box
        if filter_status == "free":
            stats["box3_color"] = "success-highlight"
        elif filter_status == "charged":
            stats["box4_color"] = "warning-highlight"
        elif filter_status == "excluded":
            stats["box5_color"] = "primary-highlight"

    # Get filter tab stats (always show these for the tabs).
    # Use Pacific dates (not UTC) so the Today/Yesterday tab counters match the
    # date-range filter behavior and what the manager sees.
    _pac_today2 = toronto_now().date()
    _pac_today_str2 = _pac_today2.strftime("%Y-%m-%d")
    _pac_yesterday_str2 = (_pac_today2 - timedelta(days=1)).strftime("%Y-%m-%d")
    _pac_7days_str2 = (_pac_today2 - timedelta(days=7)).strftime("%Y-%m-%d")
    tab_stats_raw = conn.execute(
        """
        SELECT
            COUNT(*) as total_cancellations,
            SUM(CASE WHEN DATE(c.created_at) = ? THEN 1 ELSE 0 END) as today_cancellations,
            SUM(CASE WHEN DATE(c.created_at) = ? THEN 1 ELSE 0 END) as yesterday_cancellations,
            SUM(CASE WHEN c.deadline_passed = 1 AND DATE(c.created_at) >= ? THEN 1 ELSE 0 END) as deadline_passed,
            SUM(CASE WHEN 
                (c.sequential_lessons IS NOT NULL AND c.sequential_lessons != '' AND c.sequential_lessons != '[]') 
                OR 
                (c.error_report IS NOT NULL AND c.error_report != '')
                OR
                (c.reschedule_requested = 1)
            THEN 1 ELSE 0 END) as with_notes
        FROM cancellations c
        JOIN students s ON c.student_id = s.id
        """,
        (_pac_today_str2, _pac_yesterday_str2, _pac_7days_str2),
    ).fetchone()

    tab_stats = (
        dict(tab_stats_raw)
        if tab_stats_raw
        else {
            "total_cancellations": 0,
            "today_cancellations": 0,
            "yesterday_cancellations": 0,
            "deadline_passed": 0,
            "with_notes": 0,
        }
    )

    # Get membership tiers for filter dropdown
    membership_tiers = [
        "Bronze",
        "Silver",
        "Gold",
        "Intro Package",
        "Legacy",
        "Guest",
        "Welcome Package",
    ]

    # Handle student context for filtering
    current_student_id = None
    filtered_student_name = ""

    if student_id:
        try:
            student_info = conn.execute(
                "SELECT first_name, last_name FROM students WHERE id = ?", (student_id,)
            ).fetchone()
            if student_info:
                current_student_id = int(student_id)
                filtered_student_name = (
                    f"{student_info['first_name']} {student_info['last_name']}"
                )
        except (ValueError, TypeError):
            pass

    # Mock pagination
    pagination = type(
        "Pagination",
        (),
        {
            "pages": 1,
            "has_prev": False,
            "has_next": False,
            "prev_num": None,
            "next_num": None,
            "page": 1,
            "iter_pages": lambda: [1],
        },
    )()

    conn.close()

    return render_template(
        "manager_cancellations.html",
        cancellations=cancellations,
        stats=stats,
        tab_stats=tab_stats,
        stats_context=stats_context,
        filter_status=filter_status,
        membership_tiers=membership_tiers,
        pagination=pagination,
        current_student_id=current_student_id,
        filtered_student_name=filtered_student_name,
    )


def calculate_max_cancellation_date():
    """
    Calculate the maximum allowed date for cancellation submissions.

    Policy: Cancellations allowed for current month and next month
    only if next month is 7 or less days ahead.

    Returns:
        date: Maximum allowed cancellation date
    """
    pacific_now = toronto_now()
    today = pacific_now.date()

    # Get the first day of next month
    if today.month == 12:
        next_month_start = date(today.year + 1, 1, 1)
    else:
        next_month_start = date(today.year, today.month + 1, 1)

    # Calculate days until next month
    days_until_next_month = (next_month_start - today).days

    # If next month is 7 or fewer days away, allow submissions for next month
    if days_until_next_month <= 7:
        # Allow up to the last day of next month
        if next_month_start.month == 12:
            max_date = date(next_month_start.year + 1, 1, 1) - timedelta(days=1)
        else:
            max_date = date(
                next_month_start.year, next_month_start.month + 1, 1
            ) - timedelta(days=1)
    else:
        # Only allow current month - last day of current month
        max_date = next_month_start - timedelta(days=1)

    return max_date


# Add these helper functions to your app.py file in the UTILITY FUNCTIONS section
def prepare_cancellations_for_json(cancellations):
    """
    Prepare cancellation data specifically for JSON serialization
    """
    json_cancellations = []
    for cancellation in cancellations:
        # Create a clean dict with only serializable data
        json_cancellation = {
            "id": cancellation.get("id"),
            "student_name": cancellation.get("student_name", ""),
            "first_name": cancellation.get("first_name", ""),
            "last_name": cancellation.get("last_name", ""),
            "membership_level": cancellation.get("membership_level", ""),
            "lesson_date": str(cancellation.get("lesson_date", "")),
            "lesson_time": str(cancellation.get("lesson_time", "")),
            "created_at": str(cancellation.get("created_at", "")),
            "status": cancellation.get("status", ""),
            "charged": bool(cancellation.get("charged", False)),
            "excluded": bool(cancellation.get("excluded", False)),
            "deadline_passed": bool(cancellation.get("deadline_passed", False)),
            "is_override": bool(cancellation.get("is_override", False)),
            "cancellation_note": str(cancellation.get("cancellation_note", "") or ""),
            "manager_notes": str(cancellation.get("manager_notes", "") or ""),
            "reschedule_requested": bool(
                cancellation.get("reschedule_requested", False)
            ),
            "reschedule_preferences": str(
                cancellation.get("reschedule_preferences", "") or ""
            ),
            "error_report": str(cancellation.get("error_report", "") or ""),
            # New submission fields
            "is_manager_submission": bool(cancellation.get("is_manager_submission", False)),
            "submitted_by": str(cancellation.get("submitted_by", "") or "student"),
            "submission_method": str(cancellation.get("submission_method", "") or ""),
            "actual_submission_timestamp": str(cancellation.get("actual_submission_timestamp", "")),
            "manual_submission_date": str(cancellation.get("manual_submission_date", "") or ""),
            "manager_submitted_by": cancellation.get("manager_submitted_by"),
            "suppress_notifications": bool(cancellation.get("suppress_notifications", False)),
        }
        json_cancellations.append(json_cancellation)
    return json_cancellations


def make_json_serializable(obj):
    """
    Convert objects to JSON serializable format
    """
    from datetime import datetime, date, time

    if isinstance(obj, dict):
        return {key: make_json_serializable(value) for key, value in obj.items()}
    elif isinstance(obj, list):
        return [make_json_serializable(item) for item in obj]
    elif isinstance(obj, (date, time, datetime)):
        return str(obj)
    elif hasattr(obj, "__dict__"):
        return make_json_serializable(obj.__dict__)
    else:
        return obj


def get_analytics_summary(
    date_range="month",
    status_filter="",
    membership_filter="",
    start_date="",
    end_date="",
):
    """
    Get analytics summary for dashboard widgets
    """
    conn = get_db()

    try:
        # Build filters similar to main analytics
        date_conditions = {
            "today": "DATE(c.created_at) = DATE('now')",
            "7days": "c.created_at >= DATE('now', '-7 days')",
            "month": "c.created_at >= DATE('now', '-30 days')",
            "all": "1=1",
        }

        # Handle custom date range
        if date_range == "custom" and start_date and end_date:
            date_clause = "DATE(c.created_at) BETWEEN ? AND ?"
        else:
            date_clause = date_conditions.get(date_range, date_conditions["month"])

        where_clause = f"WHERE {date_clause}"
        params = []

        # Add custom date parameters if needed
        if date_range == "custom" and start_date and end_date:
            params.extend([start_date, end_date])

        if status_filter:
            status_conditions = {
                "free": "c.charged = 0 AND c.excluded = 0 AND c.status = 'approved'",
                "charged": "c.charged = 1 AND c.excluded = 0",
                "excluded": "c.excluded = 1",
                "deadline_passed": "c.deadline_passed = 1",
                "override": "c.is_override = 1",
            }
            if status_filter in status_conditions:
                where_clause += f" AND {status_conditions[status_filter]}"

        if membership_filter:
            where_clause += " AND s.membership_level = ?"
            params.append(membership_filter)

        summary_query = f"""
            SELECT
                COUNT(*) as total_count,
                COUNT(DISTINCT c.student_id) as unique_students,
                AVG(CASE WHEN c.charged = 1 THEN 1.0 ELSE 0.0 END) as charge_rate,
                COUNT(CASE WHEN c.is_override = 1 THEN 1 END) as override_count
            FROM cancellations c
            JOIN students s ON c.student_id = s.id
            {where_clause}
        """

        result = conn.execute(summary_query, params).fetchone()
        conn.close()

        return {
            "total_count": result["total_count"] or 0,
            "unique_students": result["unique_students"] or 0,
            "charge_rate": round((result["charge_rate"] or 0) * 100, 1),
            "override_count": result["override_count"] or 0,
        }

    except Exception as e:
        conn.close()
        print(f"Error in get_analytics_summary: {e}")
        return {
            "total_count": 0,
            "unique_students": 0,
            "charge_rate": 0.0,
            "override_count": 0,
        }


# Replace your manager_analytics route with this version that fixes the data structure issues


@app.route("/manager/analytics")
@login_required
@admin_required
def manager_analytics():
    """Manager analytics page with consistent data between stats and charts"""
    # Get filter parameters
    date_range = request.args.get("date_range", "month")
    status_filter = request.args.get("status", "")
    membership_filter = request.args.get("membership", "")
    start_date = request.args.get("start_date", "")
    end_date = request.args.get("end_date", "")
    format_type = request.args.get("format", "html")

    conn = get_db()

    try:
        print(f"\n=== ANALYTICS DEBUG START ===")
        print(
            f"Filters: date_range={date_range}, status={status_filter}, membership={membership_filter}"
        )

        # FIRST: Let's see what data we actually have
        total_in_db = conn.execute(
            "SELECT COUNT(*) as count FROM cancellations"
        ).fetchone()["count"]
        print(f"Total cancellations in database: {total_in_db}")

        # Check date range of actual data
        date_range_info = conn.execute(
            """
            SELECT 
                MIN(created_at) as earliest,
                MAX(created_at) as latest,
                COUNT(*) as total
            FROM cancellations
        """
        ).fetchone()

        if date_range_info:
            print(
                f"Data spans: {date_range_info['earliest']} to {date_range_info['latest']}"
            )

        # Build date filter for FILTERED data (stats)
        date_conditions = {
            "today": "DATE(c.created_at) = DATE('now')",
            "7days": "c.created_at >= DATE('now', '-7 days')",
            "month": "c.created_at >= DATE('now', '-30 days')",
            "all": "1=1",  # No date filter - this is what finds all 14!
        }

        # Handle custom date range
        if date_range == "custom" and start_date and end_date:
            date_clause = "DATE(c.created_at) BETWEEN ? AND ?"
        else:
            date_clause = date_conditions.get(date_range, date_conditions["month"])

        # Build status filter
        status_conditions = {
            "free": "c.charged = 0 AND c.excluded = 0 AND c.status = 'approved'",
            "charged": "c.charged = 1 AND c.excluded = 0",
            "excluded": "c.excluded = 1",
            "deadline_passed": "c.deadline_passed = 1",
            "override": "c.is_override = 1",
            "note": "c.cancellation_note IS NOT NULL AND c.cancellation_note != ''",
        }
        status_clause = status_conditions.get(status_filter, "1=1")

        # Build membership filter
        where_clause = f"WHERE {date_clause} AND {status_clause}"
        params = []

        # Add custom date parameters if needed
        if date_range == "custom" and start_date and end_date:
            params.extend([start_date, end_date])

        if membership_filter:
            where_clause += " AND s.membership_level = ?"
            params.append(membership_filter)

        print(f"Filter WHERE clause: {where_clause}")
        print(f"Filter params: {params}")

        # 1. Summary statistics with applied filters
        stats_query = f"""
            SELECT
                CAST(COUNT(*) AS INTEGER) as total_cancellations,
                CAST(SUM(CASE WHEN c.charged = 0 AND c.excluded = 0 AND c.status = 'approved' THEN 1 ELSE 0 END) AS INTEGER) as free_cancellations,
                CAST(SUM(CASE WHEN c.charged = 1 AND c.excluded = 0 THEN 1 ELSE 0 END) AS INTEGER) as charged_cancellations,
                CAST(SUM(CASE WHEN c.excluded = 1 THEN 1 ELSE 0 END) AS INTEGER) as excluded_cancellations,
                CAST(SUM(CASE WHEN c.is_override = 1 THEN 1 ELSE 0 END) AS INTEGER) as override_cancellations,
                CAST(SUM(CASE WHEN c.deadline_passed = 1 THEN 1 ELSE 0 END) AS INTEGER) as deadline_passed,
                CAST(SUM(CASE WHEN 
                    (c.sequential_lessons IS NOT NULL AND c.sequential_lessons != '' AND c.sequential_lessons != '[]') 
                    OR 
                    (c.error_report IS NOT NULL AND c.error_report != '')
                    OR
                    (c.reschedule_requested = 1)
                THEN 1 ELSE 0 END) AS INTEGER) as with_notes
            FROM cancellations c
            JOIN students s ON c.student_id = s.id
            {where_clause}
        """

        stats_raw = conn.execute(stats_query, params).fetchone()

        # Process stats
        stats = {}
        if stats_raw:
            for key in stats_raw.keys():
                value = stats_raw[key]
                stats[key] = int(value) if value is not None else 0
        else:
            stats = {
                "total_cancellations": 0,
                "free_cancellations": 0,
                "charged_cancellations": 0,
                "excluded_cancellations": 0,
                "override_cancellations": 0,
                "deadline_passed": 0,
                "with_notes": 0,
            }

        print(f"STATS with current filters: {stats}")

        # 2. Monthly trends - USE SAME FILTERS for consistency!
        monthly_trends_query = f"""
            SELECT 
                strftime('%Y-%m', c.created_at) as month,
                CAST(COUNT(*) AS INTEGER) as total_cancellations,
                CAST(SUM(CASE WHEN c.charged = 0 AND c.excluded = 0 AND c.status = 'approved' THEN 1 ELSE 0 END) AS INTEGER) as free_cancellations,
                CAST(SUM(CASE WHEN c.charged = 1 AND c.excluded = 0 THEN 1 ELSE 0 END) AS INTEGER) as charged_cancellations,
                CAST(SUM(CASE WHEN c.is_override = 1 THEN 1 ELSE 0 END) AS INTEGER) as override_cancellations
            FROM cancellations c
            JOIN students s ON c.student_id = s.id
            {where_clause}
            GROUP BY strftime('%Y-%m', c.created_at)
            ORDER BY month ASC
        """

        print("Monthly trends query with same filters:")
        print(monthly_trends_query)
        print(f"With params: {params}")

        monthly_trends_raw = conn.execute(monthly_trends_query, params).fetchall()
        monthly_trends = []

        print(f"Monthly trends raw results: {len(monthly_trends_raw)} months")

        if monthly_trends_raw:
            total_from_trends = 0
            for trend in monthly_trends_raw:
                try:
                    month_date = datetime.strptime(trend["month"], "%Y-%m")
                    trend_data = {
                        "month": month_date.strftime("%b %Y"),
                        "month_short": month_date.strftime("%m/%y"),
                        "total_cancellations": int(trend["total_cancellations"] or 0),
                        "free_cancellations": int(trend["free_cancellations"] or 0),
                        "charged_cancellations": int(
                            trend["charged_cancellations"] or 0
                        ),
                        "override_cancellations": int(
                            trend["override_cancellations"] or 0
                        ),
                    }
                    monthly_trends.append(trend_data)
                    total_from_trends += trend_data["total_cancellations"]
                    print(
                        f"  {trend_data['month']}: {trend_data['total_cancellations']} total"
                    )
                except (ValueError, TypeError) as e:
                    print(f"Error parsing month {trend['month']}: {e}")
                    continue

            print(f"Total from trends: {total_from_trends}")
            print(f"Total from stats: {stats['total_cancellations']}")

            if total_from_trends != stats["total_cancellations"]:
                print("⚠️  WARNING: Trends total doesn't match stats total!")

        # If no monthly data found, create structure based on actual data
        if not monthly_trends:
            print("No monthly trends found, checking if any cancellations exist...")

            if stats["total_cancellations"] > 0:
                print("Stats show cancellations exist, but trends query found none!")

                # Let's see what months actually have data (no filters)
                all_months_query = """
                    SELECT DISTINCT 
                        strftime('%Y-%m', created_at) as month,
                        COUNT(*) as count
                    FROM cancellations 
                    GROUP BY strftime('%Y-%m', created_at)
                    ORDER BY month DESC
                """
                all_months = conn.execute(all_months_query).fetchall()
                print(
                    f"All months in DB: {[(m['month'], m['count']) for m in all_months]}"
                )

                # Create trend for current month with aggregated data
                current_month = toronto_now()
                monthly_trends = [
                    {
                        "month": current_month.strftime("%b %Y"),
                        "month_short": current_month.strftime("%m/%y"),
                        "total_cancellations": stats["total_cancellations"],
                        "free_cancellations": stats["free_cancellations"],
                        "charged_cancellations": stats["charged_cancellations"],
                        "override_cancellations": stats["override_cancellations"],
                    }
                ]
                print(f"Created synthetic monthly trend: {monthly_trends[0]}")
            else:
                # Truly no data
                current_month = toronto_now()
                monthly_trends = [
                    {
                        "month": current_month.strftime("%b %Y"),
                        "month_short": current_month.strftime("%m/%y"),
                        "total_cancellations": 0,
                        "free_cancellations": 0,
                        "charged_cancellations": 0,
                        "override_cancellations": 0,
                    }
                ]

        # 3. Top students (with same filters)
        top_students_query = f"""
            SELECT 
                s.first_name, s.last_name, s.membership_level, 
                CAST(COUNT(c.id) AS INTEGER) as cancellation_count
            FROM students s
            JOIN cancellations c ON s.id = c.student_id
            {where_clause}
            GROUP BY s.id, s.first_name, s.last_name, s.membership_level
            HAVING cancellation_count > 0
            ORDER BY cancellation_count DESC
            LIMIT 10
        """

        top_students_raw = conn.execute(top_students_query, params).fetchall()
        top_students = [dict(s) for s in top_students_raw]

        # 4. Membership tier analysis (with same filters)
        tier_analysis_query = f"""
            SELECT 
                s.membership_level,
                CAST(COUNT(DISTINCT s.id) AS INTEGER) as student_count,
                CAST(COUNT(c.id) AS INTEGER) as total_cancellations,
                CAST(SUM(CASE WHEN c.charged = 0 AND c.excluded = 0 AND c.status = 'approved' THEN 1 ELSE 0 END) AS INTEGER) as free_cancellations,
                CAST(SUM(CASE WHEN c.charged = 1 AND c.excluded = 0 THEN 1 ELSE 0 END) AS INTEGER) as charged_cancellations,
                CAST(ROUND(COUNT(c.id) * 1.0 / NULLIF(COUNT(DISTINCT s.id), 0), 2) AS REAL) as avg_monthly_cancellations
            FROM students s
            JOIN cancellations c ON s.id = c.student_id
            {where_clause}
            GROUP BY s.membership_level
            HAVING total_cancellations > 0
            ORDER BY student_count DESC
        """

        tier_distribution_raw = conn.execute(tier_analysis_query, params).fetchall()
        tier_distribution = [dict(t) for t in tier_distribution_raw]

        # 5. Filtered cancellations for table (with same filters)
        filtered_cancellations_query = f"""
            SELECT c.*, s.first_name, s.last_name, s.membership_level,
                   (s.first_name || ' ' || s.last_name) as student_name
            FROM cancellations c
            JOIN students s ON c.student_id = s.id
            {where_clause}
            ORDER BY c.created_at DESC
            LIMIT 50
        """

        filtered_cancellations_raw = conn.execute(
            filtered_cancellations_query, params
        ).fetchall()

        # Process cancellations
        filtered_cancellations = []
        for cancellation in filtered_cancellations_raw:
            processed_cancellation = process_cancellation_dates(cancellation)
            filtered_cancellations.append(processed_cancellation)

        # 6. Revenue impact (no filters - always show all for broader context)
        revenue_impact_query = """
            SELECT
                strftime('%Y-%m', c.created_at) as month,
                CAST(SUM(CASE WHEN c.charged = 1 THEN 1 ELSE 0 END) AS INTEGER) as charged_count,
                CAST(SUM(CASE WHEN c.charged = 1 THEN 1 ELSE 0 END) * 25 AS INTEGER) as estimated_revenue
            FROM cancellations c
            WHERE c.created_at >= DATE('now', '-12 months')
            GROUP BY strftime('%Y-%m', c.created_at)
            ORDER BY month ASC
        """

        revenue_impact_raw = conn.execute(revenue_impact_query).fetchall()
        revenue_impact = []

        for revenue in revenue_impact_raw:
            try:
                month_date = datetime.strptime(revenue["month"], "%Y-%m")
                revenue_impact.append(
                    {
                        "month": month_date.strftime("%b %Y"),
                        "charged_count": int(revenue["charged_count"] or 0),
                        "estimated_revenue": int(revenue["estimated_revenue"] or 0),
                    }
                )
            except (ValueError, TypeError):
                continue

        conn.close()

        print(f"FINAL RESULTS:")
        print(f"  Stats: {stats}")
        print(f"  Monthly trends: {len(monthly_trends)} months")
        print(f"  First month: {monthly_trends[0] if monthly_trends else 'None'}")
        print(f"=== ANALYTICS DEBUG END ===\n")

        # Handle JSON request for AJAX
        if format_type == "json":
            json_data = {
                "success": True,
                "stats": stats,
                "monthly_trends": monthly_trends,
                "top_students": top_students,
                "tier_distribution": tier_distribution,
                "filtered_cancellations": prepare_cancellations_for_json(
                    filtered_cancellations
                ),
                "revenue_impact": revenue_impact,
            }
            return jsonify(json_data)

        # Regular HTML response
        log_action(
            "analytics_viewed",
            f"Filters: {date_range}, {status_filter}, {membership_filter}",
        )

        return render_template(
            "manager_analytics.html",
            stats=stats,
            monthly_trends=monthly_trends,
            tier_distribution=tier_distribution,
            top_students=top_students,
            filtered_cancellations=filtered_cancellations,
            revenue_impact=revenue_impact,
            current_filters={
                "date_range": date_range,
                "status": status_filter,
                "membership": membership_filter,
            },
            debug_info={
                "trends_count": len(monthly_trends),
                "has_data": len(monthly_trends) > 0
                and any(t["total_cancellations"] > 0 for t in monthly_trends),
                "first_month": monthly_trends[0]["month"] if monthly_trends else "None",
                "total_in_first": (
                    monthly_trends[0]["total_cancellations"] if monthly_trends else 0
                ),
            },
        )

    except Exception as e:
        conn.close()
        print(f"ERROR in manager_analytics: {str(e)}")
        import traceback

        traceback.print_exc()

        log_action("analytics_error", f"Error loading analytics: {str(e)}")

        # Return safe fallback
        fallback_stats = {
            "total_cancellations": 0,
            "free_cancellations": 0,
            "charged_cancellations": 0,
            "excluded_cancellations": 0,
            "override_cancellations": 0,
            "deadline_passed": 0,
            "with_notes": 0,
        }

        fallback_trends = [
            {
                "month": toronto_now().strftime("%b %Y"),
                "month_short": toronto_now().strftime("%m/%y"),
                "total_cancellations": 0,
                "free_cancellations": 0,
                "charged_cancellations": 0,
                "override_cancellations": 0,
            }
        ]

        if format_type == "json":
            return jsonify(
                {
                    "success": False,
                    "error": str(e),
                    "stats": fallback_stats,
                    "monthly_trends": fallback_trends,
                    "tier_distribution": [],
                    "top_students": [],
                    "filtered_cancellations": [],
                    "revenue_impact": [],
                }
            )

        return render_template(
            "manager_analytics.html",
            stats=fallback_stats,
            monthly_trends=fallback_trends,
            tier_distribution=[],
            top_students=[],
            filtered_cancellations=[],
            revenue_impact=[],
            error_message=f"Analytics error: {str(e)}",
            current_filters={"date_range": "month", "status": "", "membership": ""},
            debug_info={
                "trends_count": 1,
                "has_data": False,
                "first_month": "Error",
                "total_in_first": 0,
            },
        )


# Manager API endpoints
@app.route("/manager/api/student/<int:student_id>", methods=["PUT", "POST"])
@login_required
@admin_required
def update_student(student_id):
    """Update student information"""
    conn = None
    try:
        data = request.json
        conn = get_db()
        conn.execute("PRAGMA busy_timeout = 5000")  # 5 second timeout
        
        # Start transaction
        conn.execute("BEGIN IMMEDIATE")
        try:
            conn.execute(
                """
                UPDATE students 
                SET first_name = ?, last_name = ?, parent_first = ?, parent_last = ?,
                    email = ?, phone = ?, membership_level = ?, updated_at = ?
                WHERE id = ?
            """,
                (
                    data["first_name"],
                    data["last_name"],
                    data.get("parent_first", ""),
                    data.get("parent_last", ""),
                    data["email"],
                    data.get("phone", ""),
                    data["membership_level"],
                    toronto_now().strftime("%Y-%m-%d %H:%M:%S"),
                    student_id,
                ),
            )
            conn.commit()
        except Exception as e:
            conn.rollback()
            raise e

        log_action("student_updated", f"Student ID: {student_id}")
        return jsonify({"success": True, "message": "Student updated successfully"})
    
    except Exception as e:
        if conn:
            try:
                conn.rollback()
            except:
                pass
        print(f"Error updating student {student_id}: {str(e)}")
        return jsonify({"success": False, "message": str(e)})
    finally:
        if conn:
            conn.close()


@app.route("/manager/api/student", methods=["POST"])
@login_required
@admin_required
def add_student():
    """Add new student"""
    conn = None
    try:
        data = request.json
        conn = get_db()
        conn.execute("PRAGMA busy_timeout = 5000")  # 5 second timeout
        
        # Start transaction
        conn.execute("BEGIN IMMEDIATE")
        try:
            current_time = toronto_now().strftime("%Y-%m-%d %H:%M:%S")
            membership_level = data["membership_level"]
            welcome_start_date = current_time if membership_level == "Welcome Package" else None
            
            cursor = conn.execute(
                """
                INSERT INTO students (first_name, last_name, parent_first, parent_last, email, phone, membership_level, welcome_package_date_started)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """,
                (
                    data["first_name"],
                    data["last_name"],
                    data["parent_first"],
                    data["parent_last"],
                    data["email"],
                    data["phone"],
                    membership_level,
                    welcome_start_date,
                ),
            )
            student_id = cursor.lastrowid
            
            # Track Welcome Package start
            if membership_level == "Welcome Package":
                track_package_upgrade(student_id, "None", "Welcome Package", False, conn)
            
            conn.commit()
        except Exception as e:
            conn.rollback()
            raise e

        log_action("student_added", f"New student: {data['email']}")
        return jsonify({"success": True, "student_id": student_id})
        
    except sqlite3.IntegrityError:
        if conn:
            try:
                conn.rollback()
            except:
                pass
        return jsonify({"success": False, "error": "Email already exists"})
    except Exception as e:
        if conn:
            try:
                conn.rollback()
            except:
                pass
        print(f"Error adding student: {str(e)}")
        return jsonify({"success": False, "error": str(e)})
    finally:
        if conn:
            conn.close()


@app.route("/manager/api/student/<int:student_id>", methods=["DELETE"])
@login_required
@senior_admin_required
def delete_student(student_id):
    """Delete student - Senior Manager only"""
    conn = None
    try:
        conn = get_db()
        
        # Set timeout for database operations
        conn.execute("PRAGMA busy_timeout = 5000")  # 5 second timeout
        
        # Get student info for logging before deletion
        student = conn.execute(
            "SELECT email, first_name, last_name FROM students WHERE id = ?", (student_id,)
        ).fetchone()

        if not student:
            if conn:
                conn.close()
            return jsonify({"success": False, "message": "Student not found"})

        # Delete in transaction
        conn.execute("BEGIN IMMEDIATE")
        try:
            # Delete cancellations first (foreign key constraint)
            conn.execute("DELETE FROM cancellations WHERE student_id = ?", (student_id,))
            # Delete student
            conn.execute("DELETE FROM students WHERE id = ?", (student_id,))
            conn.commit()
        except Exception as e:
            conn.rollback()
            raise e

        log_action(
            "student_deleted",
            f"Student: {student['first_name']} {student['last_name']} ({student['email']})",
        )
        return jsonify({"success": True, "message": "Student deleted successfully"})

    except Exception as e:
        if conn:
            try:
                conn.rollback()
            except:
                pass
        print(f"Error deleting student {student_id}: {str(e)}")
        return jsonify({"success": False, "message": f"Error deleting student: {str(e)}"})
    finally:
        if conn:
            conn.close()


@app.route("/manager/api/cancellation/<int:cancellation_id>/exclude", methods=["POST"])
@login_required
@admin_required
def exclude_cancellation(cancellation_id):
    """Exclude cancellation from policy (illness, etc.) - sends notifications to student and manager"""
    data = request.json
    reason = data.get("reason", "")
    suppress_notifications = data.get("suppress_notifications", False)

    if not reason.strip():
        return jsonify({"success": False, "message": "Exclusion reason is required"})

    conn = get_db()

    try:
        conn.execute(
            """UPDATE cancellations 
               SET excluded = 1, exclusion_reason = ?, approved_by = ?, 
                   is_override = 1, updated_at = ?, excluded_notification_suppressed = ?
               WHERE id = ?""",
            (
                reason,
                session["user_email"],
                toronto_now().strftime("%Y-%m-%d %H:%M:%S"),
                1 if suppress_notifications else 0,
                cancellation_id,
            ),
        )
        conn.commit()
        
        # Send notifications if not suppressed
        if not suppress_notifications:
            # Send to student
            send_exclusion_notification(cancellation_id, "student")
            # Send to managers from system settings
            send_exclusion_notification(cancellation_id, "manager")
        
        conn.close()

        log_action(
            "cancellation_excluded",
            f"Cancellation ID: {cancellation_id}, Reason: {reason}, By: {session['user_email']}, Notifications: {'suppressed' if suppress_notifications else 'sent'}",
        )
        return jsonify({"success": True})

    except Exception as e:
        conn.close()
        return jsonify({"success": False, "message": str(e)})


@app.route("/manager/api/cancellation/<int:cancellation_id>/override-excluded", methods=["POST"])
@login_required
@admin_required
def override_excluded_cancellation(cancellation_id):
    """Override an excluded cancellation back to free or charged status"""
    data = request.json
    new_status = data.get("status", "").lower()  # 'free' or 'charged'
    reason = data.get("reason", "")
    suppress_notifications = data.get("suppress_notifications", False)

    if new_status not in ["free", "charged"]:
        return jsonify({"success": False, "message": "Invalid status. Must be 'free' or 'charged'"})

    if not reason.strip():
        return jsonify({"success": False, "message": "Override reason is required"})

    conn = get_db()

    try:
        # Update the cancellation
        charged = 1 if new_status == "charged" else 0
        status_value = "charged" if new_status == "charged" else "approved"
        
        print(f"\n{'='*60}")
        print(f"DEBUG: Override Excluded Cancellation")
        print(f"{'='*60}")
        print(f"Cancellation ID: {cancellation_id}")
        print(f"New Status: {new_status}")
        print(f"Charged: {charged}")
        print(f"Status Value: {status_value}")
        print(f"Reason: {reason}")
        
        # Before update - check current status
        before = conn.execute(
            "SELECT id, excluded, charged, status FROM cancellations WHERE id = ?",
            (cancellation_id,)
        ).fetchone()
        print(f"\nBEFORE UPDATE:")
        print(f"  excluded: {before['excluded']}")
        print(f"  charged: {before['charged']}")
        print(f"  status: {before['status']}")
        
        # Perform update
        result = conn.execute(
            """UPDATE cancellations 
               SET excluded = 0, charged = ?, status = ?, manager_notes = ?, 
                   approved_by = ?, is_override = 1, updated_at = ?,
                   override_notification_suppressed = ?
               WHERE id = ?""",
            (
                charged,
                status_value,
                f"Override from exclusion to {new_status}: {reason}",
                session["user_email"],
                toronto_now().strftime("%Y-%m-%d %H:%M:%S"),
                1 if suppress_notifications else 0,
                cancellation_id,
            ),
        )
        print(f"\nUPDATE Result: {result.rowcount} rows affected")
        
        conn.commit()
        print(f"COMMIT: Success")
        
        # After update - verify changes
        after = conn.execute(
            "SELECT id, excluded, charged, status FROM cancellations WHERE id = ?",
            (cancellation_id,)
        ).fetchone()
        print(f"\nAFTER UPDATE:")
        print(f"  excluded: {after['excluded']}")
        print(f"  charged: {after['charged']}")
        print(f"  status: {after['status']}")
        print(f"{'='*60}\n")
        
        # Fetch updated cancellation and student data for email sending
        if not suppress_notifications:
            try:
                updated_cancellation = conn.execute(
                    "SELECT * FROM cancellations WHERE id = ?", (cancellation_id,)
                ).fetchone()
                
                if updated_cancellation:
                    student_id = updated_cancellation["student_id"]
                    student_record = conn.execute(
                        "SELECT * FROM students WHERE id = ?", (student_id,)
                    ).fetchone()
                    
                    if student_record:
                        # Use send_override_notification_emails for comprehensive email handling
                        override_action = "charge" if new_status == "charged" else "approve"
                        email_results = send_override_notification_emails(
                            student_record,
                            updated_cancellation,
                            override_action,
                            reason,
                            session.get("user_email", "Unknown Manager")
                        )
            except Exception as email_error:
                print(f"⚠️ Error sending override notification emails: {str(email_error)}")
        
        conn.close()

        log_action(
            "cancellation_override",
            f"Cancellation ID: {cancellation_id}, Overridden to {new_status}, Reason: {reason}, By: {session['user_email']}, Notifications: {'suppressed' if suppress_notifications else 'sent'}",
        )
        return jsonify({"success": True})

    except Exception as e:
        conn.close()
        return jsonify({"success": False, "message": str(e)})


def notification_settings():
    """Manage manager notification suppression preferences"""
    if request.method == "POST":
        data = request.json
        manager_id = session.get("user_id")
        
        # Store preferences in system_settings or create a new manager_preferences table
        # For now, we'll use system_settings with a key pattern
        try:
            conn = get_db()
            
            suppress_exclusions = data.get("suppress_exclusion_notifications", False)
            suppress_overrides = data.get("suppress_override_notifications", False)
            suppress_cancellations = data.get("suppress_cancellation_notifications", False)
            
            # Update or insert settings for this manager
            for key, value in [
                (f"manager_{manager_id}_suppress_exclusions", suppress_exclusions),
                (f"manager_{manager_id}_suppress_overrides", suppress_overrides),
                (f"manager_{manager_id}_suppress_cancellations", suppress_cancellations),
            ]:
                conn.execute(
                    """INSERT OR REPLACE INTO system_settings (key, value, description, category)
                       VALUES (?, ?, ?, ?)""",
                    (key, str(int(value)), "Manager notification preference", "notifications")
                )
            
            conn.commit()
            conn.close()
            
            return jsonify({"success": True, "message": "Notification settings updated"})
        except Exception as e:
            return jsonify({"success": False, "message": str(e)})
    
    else:  # GET
        try:
            conn = get_db()
            manager_id = session.get("user_id")
            
            # Get current settings
            settings = {}
            for key_pattern in [
                f"manager_{manager_id}_suppress_exclusions",
                f"manager_{manager_id}_suppress_overrides",
                f"manager_{manager_id}_suppress_cancellations",
            ]:
                result = conn.execute(
                    "SELECT value FROM system_settings WHERE key = ?", (key_pattern,)
                ).fetchone()
                
                setting_name = key_pattern.replace(f"manager_{manager_id}_", "")
                settings[setting_name] = bool(int(result["value"])) if result else False
            
            conn.close()
            
            return jsonify({
                "success": True,
                "settings": settings
            })
        except Exception as e:
            return jsonify({"success": False, "message": str(e)})



@login_required
@admin_required
def bulk_import_students():
    """Bulk import students from CSV data"""
    data = request.json
    csv_data = data.get("csv_data", "")

    if not csv_data:
        return jsonify({"success": False, "error": "No data provided"})

    # Parse CSV data
    csv_reader = csv.reader(io.StringIO(csv_data))
    imported = 0
    errors = []

    conn = get_db()

    for row_num, row in enumerate(csv_reader, 1):
        if len(row) < 6:  # Minimum required columns
            errors.append(f"Row {row_num}: Insufficient data")
            continue

        try:
            current_time = toronto_now().strftime("%Y-%m-%d %H:%M:%S")
            membership_level = row[6] if len(row) > 6 else "Bronze"
            welcome_start_date = current_time if membership_level == "Welcome Package" else None
            
            cursor = conn.execute(
                """
                INSERT INTO students (first_name, last_name, parent_first, parent_last, email, phone, membership_level, welcome_package_date_started)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """,
                (
                    row[0],
                    row[1],
                    row[2],
                    row[3],
                    row[4],
                    row[5],
                    membership_level,
                    welcome_start_date,
                ),
            )
            
            # Track Welcome Package start
            if membership_level == "Welcome Package":
                student_id = cursor.lastrowid
                track_package_upgrade(student_id, "None", "Welcome Package", False, conn)
            
            imported += 1
        except sqlite3.IntegrityError:
            errors.append(f"Row {row_num}: Email {row[4]} already exists")
        except Exception as e:
            errors.append(f"Row {row_num}: {str(e)}")

    conn.commit()
    conn.close()

    log_action("bulk_import", f"Imported {imported} students, {len(errors)} errors")
    return jsonify({"success": True, "imported": imported, "errors": errors})


@app.route("/manager/export/students")
@login_required
@admin_required
def export_students():
    """Export students to CSV"""
    conn = get_db()
    students = conn.execute(
        "SELECT * FROM students ORDER BY last_name, first_name"
    ).fetchall()
    conn.close()

    # Create CSV output
    output = io.StringIO()
    writer = csv.writer(output)

    # Header
    writer.writerow(
        [
            "First Name",
            "Last Name",
            "Parent First",
            "Parent Last",
            "Email",
            "Phone",
            "Membership Level",
            "Created",
        ]
    )

    # Data rows
    for student in students:
        writer.writerow(
            [
                student["first_name"],
                student["last_name"],
                student["parent_first"],
                student["parent_last"],
                student["email"],
                student["phone"],
                student["membership_level"],
                student["created_at"],
            ]
        )

    # Create response
    response = app.response_class(
        output.getvalue(),
        mimetype="text/csv",
        headers={"Content-Disposition": "attachment; filename=students_export.csv"},
    )

    log_action("export_students", f"Exported {len(students)} students")
    return response


@app.route("/manager/api/student/<int:student_id>", methods=["POST"])
@login_required
@admin_required
def update_student_post(student_id):
    """Update student information via POST (for the edit form)"""
    conn = None
    try:
        data = request.json
        conn = get_db()
        conn.execute("PRAGMA busy_timeout = 5000")  # 5 second timeout

        # Start transaction
        conn.execute("BEGIN IMMEDIATE")
        try:
            conn.execute(
                """
                UPDATE students 
                SET first_name = ?, last_name = ?, parent_first = ?, parent_last = ?,
                    email = ?, phone = ?, membership_level = ?, updated_at = ?
                WHERE id = ?
            """,
                (
                    data["first_name"],
                    data["last_name"],
                    data.get("parent_first", ""),
                    data.get("parent_last", ""),
                    data["email"],
                    data.get("phone", ""),
                    data["membership_level"],
                    toronto_now().strftime("%Y-%m-%d %H:%M:%S"),
                    student_id,
                ),
            )
            conn.commit()
        except Exception as e:
            conn.rollback()
            raise e

        log_action("student_updated", f"Student ID: {student_id}")
        return jsonify({"success": True, "message": "Student updated successfully"})

    except Exception as e:
        if conn:
            try:
                conn.rollback()
            except:
                pass
        print(f"Error updating student {student_id}: {str(e)}")
        return jsonify({"success": False, "message": str(e)})
    finally:
        if conn:
            conn.close()


@app.route("/manager/api/student/<int:student_id>/delete", methods=["DELETE"])
@login_required
@senior_admin_required
def delete_student_api(student_id):
    """Delete student via API - Senior Manager only"""
    conn = None
    try:
        conn = get_db()
        conn.execute("PRAGMA busy_timeout = 5000")  # 5 second timeout

        # Get student info for logging
        student = conn.execute(
            "SELECT email, first_name, last_name FROM students WHERE id = ?",
            (student_id,),
        ).fetchone()

        if not student:
            if conn:
                conn.close()
            return jsonify({"success": False, "message": "Student not found"})

        # Delete in transaction
        conn.execute("BEGIN IMMEDIATE")
        try:
            # Delete cancellations first (foreign key constraint)
            conn.execute("DELETE FROM cancellations WHERE student_id = ?", (student_id,))
            conn.execute("DELETE FROM students WHERE id = ?", (student_id,))
            conn.commit()
        except Exception as e:
            conn.rollback()
            raise e

        log_action(
            "student_deleted",
            f"Student: {student['first_name']} {student['last_name']} ({student['email']})",
        )
        return jsonify({"success": True, "message": "Student deleted successfully"})

    except Exception as e:
        if conn:
            try:
                conn.rollback()
            except:
                pass
        print(f"Error deleting student {student_id}: {str(e)}")
        return jsonify({"success": False, "message": str(e)})
    finally:
        if conn:
            conn.close()


@app.route("/manager/api/students/bulk-delete", methods=["POST"])
@login_required
@senior_admin_required
def bulk_delete_students():
    """Delete multiple students via API - Senior Manager only"""
    conn = None
    try:
        data = request.json
        student_ids = data.get("student_ids", [])
        
        if not student_ids:
            return jsonify({"success": False, "message": "No students selected"})
        
        conn = get_db()
        conn.execute("PRAGMA busy_timeout = 5000")  # 5 second timeout
        deleted_count = 0
        
        # Start transaction
        conn.execute("BEGIN IMMEDIATE")
        try:
            for student_id in student_ids:
                # Get student info for logging
                student = conn.execute(
                    "SELECT email, first_name, last_name FROM students WHERE id = ?",
                    (student_id,),
                ).fetchone()
                
                if student:
                    # Delete cancellations first (foreign key constraint)
                    conn.execute("DELETE FROM cancellations WHERE student_id = ?", (student_id,))
                    conn.execute("DELETE FROM students WHERE id = ?", (student_id,))
                    deleted_count += 1
                    
                    log_action(
                        "student_deleted",
                        f"Student: {student['first_name']} {student['last_name']} ({student['email']})",
                    )
            
            conn.commit()
        except Exception as e:
            conn.rollback()
            raise e
        
        return jsonify({
            "success": True,
            "message": f"Successfully deleted {deleted_count} student(s)",
            "deleted_count": deleted_count
        })
    
    except Exception as e:
        if conn:
            try:
                conn.rollback()
            except:
                pass
        print(f"Error in bulk delete: {str(e)}")
        return jsonify({"success": False, "message": str(e)})
    finally:
        if conn:
            conn.close()


@app.route("/manager/api/student/<int:student_id>/membership", methods=["POST"])
@login_required
@admin_required
def change_student_membership(student_id):
    """Change student membership level"""
    conn = None
    try:
        data = request.json
        new_membership = data.get("membership_level")

        if not new_membership:
            return jsonify(
                {"success": False, "message": "Membership level is required"}
            )

        conn = get_db()
        conn.execute("PRAGMA busy_timeout = 5000")  # 5 second timeout

        # Get student info for logging
        student = conn.execute(
            "SELECT first_name, last_name, membership_level FROM students WHERE id = ?",
            (student_id,),
        ).fetchone()

        if not student:
            if conn:
                conn.close()
            return jsonify({"success": False, "message": "Student not found"})

        # Start transaction
        conn.execute("BEGIN IMMEDIATE")
        try:
            # Update membership
            old_package = student['membership_level']
            conn.execute(
                "UPDATE students SET membership_level = ?, updated_at = ? WHERE id = ?",
                (new_membership, toronto_now().strftime("%Y-%m-%d %H:%M:%S"), student_id),
            )
            
            # Track package upgrade/change in history
            if old_package != new_membership:
                # Check if Welcome free was used when upgrading FROM Welcome Package
                welcome_free_used_at_upgrade = False
                if old_package == "Welcome Package":
                    student_check = conn.execute(
                        "SELECT welcome_free_used FROM students WHERE id = ?",
                        (student_id,)
                    ).fetchone()
                    welcome_free_used_at_upgrade = bool(student_check["welcome_free_used"]) if student_check else False
                    
                    # Set upgrade date if leaving Welcome Package
                    conn.execute(
                        "UPDATE students SET welcome_package_upgrade_date = ? WHERE id = ?",
                        (toronto_now().strftime("%Y-%m-%d %H:%M:%S"), student_id)
                    )
                
                # Record in package history
                track_package_upgrade(student_id, old_package, new_membership, welcome_free_used_at_upgrade, conn)
                
                # If joining Welcome Package, record the start date
                if new_membership == "Welcome Package":
                    record_welcome_package_start(student_id, conn)
            
            conn.commit()
        except Exception as e:
            conn.rollback()
            raise e

        log_action(
            "membership_changed",
            f"Student: {student['first_name']} {student['last_name']}, "
            f"From: {student['membership_level']}, To: {new_membership}",
        )

        return jsonify({"success": True, "message": "Membership updated successfully"})

    except Exception as e:
        if conn:
            try:
                conn.rollback()
            except:
                pass
        print(f"Error changing student membership: {str(e)}")
        return jsonify({"success": False, "message": str(e)})
    finally:
        if conn:
            conn.close()


# ===================================
# SENIOR MANAGER ROUTES
# ===================================


@app.route("/senior/dashboard")
@login_required
@senior_admin_required
def senior_dashboard():
    """Senior manager dashboard with improved data handling"""
    conn = get_db()

    try:
        # Get comprehensive dashboard stats - these are correct values
        stats = get_dashboard_stats()
        print(f"DEBUG: Stats from get_dashboard_stats: {stats}")

        # System usage analytics (last 30 days) with better error handling
        usage_analytics_raw = conn.execute(
            """
            SELECT 
                date(created_at) as date,
                COUNT(*) as cancellations,
                COUNT(DISTINCT student_id) as active_students
            FROM cancellations
            WHERE created_at >= date('now', '-30 days')
            GROUP BY date(created_at)
            ORDER BY date DESC
            LIMIT 30
        """
        ).fetchall()

        # Convert to list of dicts for template
        usage_analytics = []
        for row in usage_analytics_raw:
            usage_analytics.append(
                {
                    "date": row["date"],
                    "cancellations": row["cancellations"] or 0,
                    "active_students": row["active_students"] or 0,
                }
            )

        # Recent system activity with better formatting
        recent_activity_raw = conn.execute(
            """
            SELECT 
                l.action,
                l.details,
                l.created_at,
                l.user_type,
                CASE 
                    WHEN l.user_type = 'client' AND s.first_name IS NOT NULL 
                        THEN s.first_name || ' ' || s.last_name
                    WHEN l.user_type = 'admin' AND a.first_name IS NOT NULL 
                        THEN a.first_name || ' ' || a.last_name
                    ELSE 'System User'
                END as user_name
            FROM system_logs l
            LEFT JOIN students s ON l.user_id = s.id AND l.user_type = 'client'
            LEFT JOIN admin_users a ON l.user_id = a.id AND l.user_type = 'admin'
            WHERE l.created_at >= date('now', '-7 days')
            ORDER BY l.created_at DESC
            LIMIT 15
        """
        ).fetchall()

        # Convert to list of dicts
        recent_activity = []
        for row in recent_activity_raw:
            recent_activity.append(
                {
                    "action": row["action"],
                    "details": row["details"] or "",
                    "created_at": row["created_at"],
                    "user_type": row["user_type"] or "unknown",
                    "user_name": row["user_name"] or "Unknown User",
                }
            )

        # Recent trends for insights
        monthly_trends = conn.execute(
            """
            SELECT 
                strftime('%Y-%m', created_at) as month,
                COUNT(*) as total_cancellations,
                SUM(CASE WHEN charged = 0 THEN 1 ELSE 0 END) as free_cancellations,
                SUM(CASE WHEN charged = 1 THEN 1 ELSE 0 END) as charged_cancellations
            FROM cancellations
            WHERE created_at >= date('now', '-6 months')
            GROUP BY strftime('%Y-%m', created_at)
            ORDER BY month DESC
        """
        ).fetchall()

        conn.close()

        # Make sure all needed properties exist in stats for template compatibility
        final_stats = {
            # Core stats that work correctly
            "active_students": stats.get("active_students", 0),
            "month_cancellations": stats.get("month_cancellations", 0),
            "pending_reviews": stats.get("pending_reviews", 0),
            "active_tiers": stats.get("active_tiers", 7),
            # Additional properties for completeness
            "today_cancellations": stats.get("today_cancellations", 0),
            "today_vs_yesterday": stats.get("today_vs_yesterday", "No data"),
            "free_cancellations": stats.get("free_cancellations", 0),
            "charged_cancellations": stats.get("charged_cancellations", 0),
            "excluded_cancellations": stats.get("excluded_cancellations", 0),
            "new_students_month": stats.get("new_students_month", 0),
            "total_students": stats.get(
                "total_students", stats.get("active_students", 0)
            ),
            "monthly_cancellations": stats.get(
                "monthly_cancellations", stats.get("month_cancellations", 0)
            ),
            "system_alerts": stats.get("system_alerts", 0),
        }

        print(f"DEBUG: Final stats being sent to template: {final_stats}")

        log_action("senior_dashboard_viewed", "Senior manager accessed dashboard")

        return render_template(
            "senior_dashboard.html",
            stats=final_stats,
            usage_analytics=usage_analytics,
            recent_activity=recent_activity,
            monthly_trends=[dict(row) for row in monthly_trends],
        )

    except Exception as e:
        conn.close()
        print(f"ERROR in senior_dashboard: {str(e)}")
        import traceback

        traceback.print_exc()

        # Log error
        log_action(
            "senior_dashboard_error", f"Error loading senior dashboard: {str(e)}"
        )

        # Fallback with minimal data
        fallback_stats = {
            "active_students": 0,
            "month_cancellations": 0,
            "pending_reviews": 0,
            "total_students": 0,
            "active_tiers": 7,
            "monthly_cancellations": 0,
            "today_cancellations": 0,
            "today_vs_yesterday": "Error",
            "free_cancellations": 0,
            "charged_cancellations": 0,
            "excluded_cancellations": 0,
            "new_students_month": 0,
            "system_alerts": 0,
        }

        return render_template(
            "senior_dashboard.html",
            stats=fallback_stats,
            usage_analytics=[],
            recent_activity=[],
            monthly_trends=[],
            error_message=f"Dashboard data unavailable: {str(e)}",
        )


@app.route("/senior/generate-report", methods=["GET"])
@login_required
@senior_admin_required
def generate_system_report():
    """Generate comprehensive system report"""
    try:
        conn = get_db()

        # Collect comprehensive system data
        report_data = {
            "timestamp": datetime.now().isoformat(),
            "students": conn.execute(
                "SELECT COUNT(*) as count FROM students"
            ).fetchone()["count"],
            "cancellations": conn.execute(
                "SELECT COUNT(*) as count FROM cancellations"
            ).fetchone()["count"],
            "this_month": conn.execute(
                "SELECT COUNT(*) as count FROM cancellations WHERE created_at >= date('now', 'start of month')"
            ).fetchone()["count"],
            "pending": conn.execute(
                "SELECT COUNT(*) as count FROM cancellations WHERE status = 'pending'"
            ).fetchone()["count"],
        }

        conn.close()

        # Create simple HTML report
        html_report = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <title>System Report - {report_data['timestamp'][:10]}</title>
            <style>
                body {{ font-family: Arial, sans-serif; margin: 40px; }}
                .header {{ border-bottom: 2px solid #333; padding-bottom: 20px; }}
                .metric {{ margin: 20px 0; padding: 15px; background: #f5f5f5; border-radius: 5px; }}
                .metric h3 {{ margin-top: 0; }}
            </style>
        </head>
        <body>
            <div class="header">
                <h1>Riverside Equestrian System Report</h1>
                <p>Generated: {report_data['timestamp']}</p>
            </div>
            
            <div class="metric">
                <h3>Total Students</h3>
                <p>{report_data['students']}</p>
            </div>
            
            <div class="metric">
                <h3>Total Cancellations</h3>
                <p>{report_data['cancellations']}</p>
            </div>
            
            <div class="metric">
                <h3>This Month's Cancellations</h3>
                <p>{report_data['this_month']}</p>
            </div>
            
            <div class="metric">
                <h3>Pending Reviews</h3>
                <p>{report_data['pending']}</p>
            </div>
        </body>
        </html>
        """

        log_action(
            "system_report_generated", f"Generated at {report_data['timestamp']}"
        )

        return html_report

    except Exception as e:
        log_action("system_report_error", f"Failed to generate report: {str(e)}")
        return (
            f"<html><body><h1>Error generating report</h1><p>{str(e)}</p></body></html>"
        )


@app.route("/senior/backup", methods=["POST"])
@login_required
@senior_admin_required
def senior_backup():
    """Create system backup - complete implementation"""
    import shutil
    import os
    import sqlite3
    from datetime import datetime
    import tempfile
    import zipfile

    try:
        # Create timestamp for backup files
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_dir = "backups"

        # Create backups directory if it doesn't exist
        os.makedirs(backup_dir, exist_ok=True)

        # Backup filename
        backup_filename = f"riverside_backup_{timestamp}"
        backup_path = os.path.join(backup_dir, backup_filename)

        # Create a zip file for the complete backup
        zip_path = f"{backup_path}.zip"

        with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as backup_zip:

            # 1. Backup the database file
            db_backup_path = f"{backup_path}_database.db"

            # Create a proper SQLite backup
            source_conn = sqlite3.connect(app.config["DATABASE"])
            backup_conn = sqlite3.connect(db_backup_path)

            # Use SQLite's backup API for a clean backup
            source_conn.backup(backup_conn)

            # Close connections
            source_conn.close()
            backup_conn.close()

            # Add database to zip
            backup_zip.write(db_backup_path, f"database_{timestamp}.db")

            # 2. Create a SQL dump as well
            sql_dump_path = f"{backup_path}_dump.sql"
            with open(sql_dump_path, "w") as sql_file:
                conn = sqlite3.connect(app.config["DATABASE"])
                for line in conn.iterdump():
                    sql_file.write("%s\n" % line)
                conn.close()

            # Add SQL dump to zip
            backup_zip.write(sql_dump_path, f"database_dump_{timestamp}.sql")

            # 3. Backup configuration and logs if they exist
            config_files = []

            # Add any config files that exist
            if os.path.exists("config.py"):
                config_files.append("config.py")
            if os.path.exists("requirements.txt"):
                config_files.append("requirements.txt")

            for config_file in config_files:
                try:
                    backup_zip.write(config_file, f"config/{config_file}")
                except:
                    pass  # Skip if file doesn't exist or can't be read

            # 4. Create a backup manifest
            manifest_path = f"{backup_path}_manifest.txt"
            with open(manifest_path, "w") as manifest:
                manifest.write(f"Riverside Equestrian System Backup\n")
                manifest.write(f"Created: {datetime.now().isoformat()}\n")
                manifest.write(f"Created by: {session.get('user_email', 'Unknown')}\n")
                manifest.write(f"Database: {app.config['DATABASE']}\n")
                manifest.write(f"Backup includes:\n")
                manifest.write(f"- Complete database backup\n")
                manifest.write(f"- SQL dump\n")
                manifest.write(f"- Configuration files\n")

                # Add some statistics
                conn = get_db()
                try:
                    stats = conn.execute(
                        """
                        SELECT 
                            (SELECT COUNT(*) FROM students) as students,
                            (SELECT COUNT(*) FROM cancellations) as cancellations,
                            (SELECT COUNT(*) FROM admin_users) as admins,
                            (SELECT COUNT(*) FROM membership_tiers) as tiers
                    """
                    ).fetchone()

                    manifest.write(f"\nDatabase Statistics:\n")
                    manifest.write(f"- Students: {stats['students']}\n")
                    manifest.write(f"- Cancellations: {stats['cancellations']}\n")
                    manifest.write(f"- Admin Users: {stats['admins']}\n")
                    manifest.write(f"- Membership Tiers: {stats['tiers']}\n")
                except:
                    manifest.write(f"\nCould not retrieve database statistics\n")
                finally:
                    conn.close()

            # Add manifest to zip
            backup_zip.write(manifest_path, f"backup_manifest_{timestamp}.txt")

        # Clean up temporary files
        temp_files = [db_backup_path, sql_dump_path, manifest_path]
        for temp_file in temp_files:
            try:
                os.remove(temp_file)
            except:
                pass  # Ignore cleanup errors

        # Calculate backup size
        backup_size = os.path.getsize(zip_path)
        backup_size_mb = round(backup_size / (1024 * 1024), 2)

        # Log the successful backup
        log_action(
            "backup_created",
            f"System backup created: {backup_filename}.zip ({backup_size_mb} MB)",
        )

        # Return success response
        return jsonify(
            {
                "success": True,
                "filename": f"{backup_filename}.zip",
                "size_mb": backup_size_mb,
                "path": zip_path,
                "timestamp": timestamp,
                "message": f"Backup created successfully! File size: {backup_size_mb} MB",
            }
        )

    except sqlite3.Error as db_error:
        # Database-specific error
        error_msg = f"Database backup failed: {str(db_error)}"
        log_action("backup_failed", error_msg)
        return jsonify({"success": False, "error": error_msg, "error_type": "database"})

    except OSError as file_error:
        # File system error
        error_msg = f"File system error during backup: {str(file_error)}"
        log_action("backup_failed", error_msg)
        return jsonify(
            {"success": False, "error": error_msg, "error_type": "filesystem"}
        )

    except Exception as e:
        # General error
        error_msg = f"Backup failed: {str(e)}"
        log_action("backup_failed", error_msg)
        return jsonify({"success": False, "error": error_msg, "error_type": "general"})


@app.route("/senior/list-backups", methods=["GET"])
@login_required
@senior_admin_required
def list_backups():
    """List all available backups"""
    try:
        backup_dir = "backups"
        backups = []

        if os.path.exists(backup_dir):
            for filename in os.listdir(backup_dir):
                if filename.endswith(".zip") and filename.startswith(
                    "riverside_backup_"
                ):
                    file_path = os.path.join(backup_dir, filename)
                    file_stats = os.stat(file_path)

                    # Extract timestamp from filename
                    timestamp_str = filename.replace("riverside_backup_", "").replace(
                        ".zip", ""
                    )
                    try:
                        created_date = datetime.strptime(timestamp_str, "%Y%m%d_%H%M%S")
                    except:
                        created_date = datetime.fromtimestamp(file_stats.st_mtime)

                    backups.append(
                        {
                            "filename": filename,
                            "size_mb": round(file_stats.st_size / (1024 * 1024), 2),
                            "created": created_date.isoformat(),
                            "created_readable": created_date.strftime(
                                "%Y-%m-%d %H:%M:%S"
                            ),
                        }
                    )

        # Sort by creation date (newest first)
        backups.sort(key=lambda x: x["created"], reverse=True)

        return jsonify({"success": True, "backups": backups, "count": len(backups)})

    except Exception as e:
        return jsonify({"success": False, "error": str(e)})


@app.route("/senior/download-backup/<filename>", methods=["GET"])
@login_required
@senior_admin_required
def download_backup(filename):
    """Download a specific backup file"""
    try:
        # Sanitize filename to prevent directory traversal
        safe_filename = os.path.basename(filename)
        if not safe_filename.endswith(".zip") or not safe_filename.startswith(
            "riverside_backup_"
        ):
            return jsonify({"success": False, "error": "Invalid backup file"}), 400

        backup_path = os.path.join("backups", safe_filename)

        if not os.path.exists(backup_path):
            return jsonify({"success": False, "error": "Backup file not found"}), 404

        log_action("backup_downloaded", f"Downloaded backup: {safe_filename}")

        return send_file(
            backup_path,
            as_attachment=True,
            download_name=safe_filename,
            mimetype="application/zip",
        )

    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


@app.route("/senior/delete-backup/<filename>", methods=["DELETE"])
@login_required
@senior_admin_required
def delete_backup(filename):
    """Delete a specific backup file"""
    try:
        # Sanitize filename
        safe_filename = os.path.basename(filename)
        if not safe_filename.endswith(".zip") or not safe_filename.startswith(
            "riverside_backup_"
        ):
            return jsonify({"success": False, "error": "Invalid backup file"}), 400

        backup_path = os.path.join("backups", safe_filename)

        if not os.path.exists(backup_path):
            return jsonify({"success": False, "error": "Backup file not found"}), 404

        os.remove(backup_path)
        log_action("backup_deleted", f"Deleted backup: {safe_filename}")

        return jsonify(
            {"success": True, "message": f"Backup {safe_filename} deleted successfully"}
        )

    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


@app.route("/senior/settings")
@login_required
@senior_admin_required
def senior_settings():
    """Senior manager settings"""
    conn = get_db()

    # Get all system settings
    settings = {}
    settings_rows = conn.execute("SELECT key, value FROM system_settings").fetchall()
    for setting in settings_rows:
        settings[setting["key"]] = setting["value"]

    conn.close()

    return render_template("senior_settings.html", settings=settings)


@app.route("/senior/tiers")
@login_required
@senior_admin_required
def senior_tiers():
    """Senior manager tier management"""
    conn = get_db()

    # Get all tiers with student counts
    tiers = conn.execute(
        """
        SELECT mt.*, COUNT(s.id) as student_count
        FROM membership_tiers mt
        LEFT JOIN students s ON mt.level = s.membership_level
        GROUP BY mt.level
        ORDER BY mt.sort_order
    """
    ).fetchall()

    conn.close()

    return render_template("senior_tiers.html", tiers=tiers)


# ===================================
# TEMPLATE MANAGEMENT API ROUTES
# ===================================


@app.route("/senior/api/template-variables", methods=["GET"])
@login_required
@senior_admin_required
def get_template_variables_api():
    """Get available template variables with categories"""
    try:
        # Sample variables for demonstration
        sample_variables = {
            "client_name": "John Smith",
            "client_first_name": "John",
            "client_last_name": "Smith",
            "client_email": "john@example.com",
            "client_phone": "604-123-4567",
            "membership_tier": "Silver",
            "parent_name": "Jane Smith",
            "lesson_date": "March 15, 2024",
            "lesson_time": "3:00 PM",
            "cancellation_status": "Free cancellation",
            "company_name": "Riverside Equestrian",
            "contact_email": "managers@riversideequestrian.ca",
            "current_date": "March 10, 2024",
            "policy_url": "https://www.riversideequestrian.ca/cancellations",
        }

        # Organize variables by category
        variable_categories = {
            "Student Information": [
                "client_name",
                "client_first_name",
                "client_last_name",
                "client_email",
                "client_phone",
                "membership_tier",
                "parent_name",
            ],
            "Lesson Details": ["lesson_date", "lesson_time", "cancellation_status"],
            "Company Information": [
                "company_name",
                "contact_email",
                "policy_url",
                "current_date",
            ],
        }

        return jsonify(
            {
                "success": True,
                "variables": sample_variables,
                "categories": variable_categories,
            }
        )

    except Exception as e:
        return jsonify({"success": False, "message": str(e)})


@app.route("/senior/api/template/<template_id>", methods=["GET"])
@login_required
@senior_admin_required
def get_template_for_edit(template_id):
    """Get template data for editing"""
    try:
        conn = get_db()
        template = conn.execute(
            "SELECT * FROM email_templates WHERE id = ?", (template_id,)
        ).fetchone()
        conn.close()

        if not template:
            return jsonify({"success": False, "message": "Template not found"})

        template_dict = dict(template)
        return jsonify({"success": True, "template": template_dict})

    except Exception as e:
        return jsonify({"success": False, "message": str(e)})


@app.route("/senior/api/template/save", methods=["POST"])
@login_required
@senior_admin_required
def save_template_enhanced():
    """Save or update email template"""
    try:
        data = request.json

        # Validate required fields
        required_fields = ["id", "name", "subject", "body", "type"]
        for field in required_fields:
            if not data.get(field):
                return jsonify(
                    {"success": False, "message": f"Missing required field: {field}"}
                )

        # Clean and prepare data
        template_data = {
            "id": data["id"].strip().replace(" ", "_").lower(),
            "name": data["name"].strip(),
            "subject": data["subject"].strip(),
            "body": data["body"],
            "type": data.get("type", "client"),
            "active": bool(data.get("active", True)),
            "auto_send": bool(data.get("auto_send", True)),
            "priority": data.get("priority", "normal"),
            "delay_minutes": int(data.get("delay_minutes", 0)),
            "include_attachment": bool(data.get("include_attachment", False)),
        }

        # Extract variables used in template
        import re

        variables_used = []
        variable_pattern = r"\{\{(\w+)\}\}"

        # Find variables in subject
        subject_vars = re.findall(variable_pattern, template_data["subject"])
        variables_used.extend(subject_vars)

        # Find variables in body
        body_vars = re.findall(variable_pattern, template_data["body"])
        variables_used.extend(body_vars)

        template_data["variables_used"] = ",".join(set(variables_used))

        conn = get_db()

        # Check if template exists
        existing = conn.execute(
            "SELECT id FROM email_templates WHERE id = ?", (template_data["id"],)
        ).fetchone()

        if existing:
            # Update existing template
            conn.execute(
                """
                UPDATE email_templates 
                SET name = ?, subject = ?, body = ?, type = ?, active = ?, 
                    auto_send = ?, priority = ?, delay_minutes = ?, 
                    include_attachment = ?, variables_used = ?, updated_at = ?
                WHERE id = ?
                """,
                (
                    template_data["name"],
                    template_data["subject"],
                    template_data["body"],
                    template_data["type"],
                    template_data["active"],
                    template_data["auto_send"],
                    template_data["priority"],
                    template_data["delay_minutes"],
                    template_data["include_attachment"],
                    template_data["variables_used"],
                    toronto_now().strftime("%Y-%m-%d %H:%M:%S"),
                    template_data["id"],
                ),
            )
            action = "updated"
        else:
            # Insert new template
            conn.execute(
                """
                INSERT INTO email_templates 
                (id, name, subject, body, type, active, auto_send, priority, 
                 delay_minutes, include_attachment, variables_used, created_at, updated_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    template_data["id"],
                    template_data["name"],
                    template_data["subject"],
                    template_data["body"],
                    template_data["type"],
                    template_data["active"],
                    template_data["auto_send"],
                    template_data["priority"],
                    template_data["delay_minutes"],
                    template_data["include_attachment"],
                    template_data["variables_used"],
                    toronto_now().strftime("%Y-%m-%d %H:%M:%S"),
                    toronto_now().strftime("%Y-%m-%d %H:%M:%S"),
                ),
            )
            action = "created"

        conn.commit()
        conn.close()

        return jsonify(
            {
                "success": True,
                "message": f"Template {action} successfully",
                "template_id": template_data["id"],
            }
        )

    except Exception as e:
        return jsonify({"success": False, "message": str(e)})


@app.route("/senior/api/template/<template_id>/delete", methods=["DELETE"])
@login_required
@senior_admin_required
def delete_template_enhanced(template_id):
    """Delete email template"""
    try:
        conn = get_db()

        # Check if template exists
        template = conn.execute(
            "SELECT name FROM email_templates WHERE id = ?", (template_id,)
        ).fetchone()

        if not template:
            return jsonify({"success": False, "message": "Template not found"})

        # Delete template
        conn.execute("DELETE FROM email_templates WHERE id = ?", (template_id,))
        conn.commit()
        conn.close()

        return jsonify({"success": True, "message": "Template deleted successfully"})

    except Exception as e:
        return jsonify({"success": False, "message": str(e)})


@app.route("/senior/api/template/<template_id>/toggle", methods=["POST"])
@login_required
@senior_admin_required
def toggle_template_enhanced(template_id):
    """Toggle template active status"""
    try:
        conn = get_db()

        # Get current status
        template = conn.execute(
            "SELECT active, name FROM email_templates WHERE id = ?", (template_id,)
        ).fetchone()

        if not template:
            return jsonify({"success": False, "message": "Template not found"})

        new_status = not bool(template["active"])

        # Update status
        conn.execute(
            "UPDATE email_templates SET active = ?, updated_at = ? WHERE id = ?",
            (new_status, toronto_now().strftime("%Y-%m-%d %H:%M:%S"), template_id),
        )
        conn.commit()
        conn.close()

        status_text = "activated" if new_status else "deactivated"
        return jsonify(
            {
                "success": True,
                "message": f"Template '{template['name']}' {status_text}",
                "active": new_status,
            }
        )

    except Exception as e:
        return jsonify({"success": False, "message": str(e)})


@app.route("/senior/api/template/<template_id>/duplicate", methods=["POST"])
@login_required
@senior_admin_required
def duplicate_template_enhanced(template_id):
    """Duplicate an email template"""
    try:
        conn = get_db()
        template = conn.execute(
            "SELECT * FROM email_templates WHERE id = ?", (template_id,)
        ).fetchone()

        if not template:
            return jsonify({"success": False, "message": "Template not found"})

        template_dict = dict(template)

        # Create new template data
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        new_id = f"{template_id}_copy_{timestamp}"

        # Insert duplicated template
        conn.execute(
            """
            INSERT INTO email_templates 
            (id, name, subject, body, type, active, auto_send, priority, 
             delay_minutes, include_attachment, variables_used, created_at, updated_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                new_id,
                f"Copy of {template_dict['name']}",
                template_dict["subject"],
                template_dict["body"],
                template_dict["type"],
                False,  # Start inactive
                template_dict.get("auto_send", True),
                template_dict.get("priority", "normal"),
                template_dict.get("delay_minutes", 0),
                template_dict.get("include_attachment", False),
                template_dict.get("variables_used", ""),
                toronto_now().strftime("%Y-%m-%d %H:%M:%S"),
                toronto_now().strftime("%Y-%m-%d %H:%M:%S"),
            ),
        )
        conn.commit()
        conn.close()

        log_action("template_duplicated", f"Duplicated {template_id} to {new_id}")

        return jsonify(
            {
                "success": True,
                "message": "Template duplicated successfully",
                "new_id": new_id,
            }
        )

    except Exception as e:
        return jsonify({"success": False, "message": str(e)})


@app.route("/senior/api/template/<template_id>/test", methods=["POST"])
@login_required
@senior_admin_required
def send_template_test_enhanced(template_id):
    """Send test email for a template"""
    try:
        data = request.json
        test_email = data.get("test_email")

        if not test_email:
            return jsonify({"success": False, "message": "Test email address required"})

        # Validate email format
        import re

        email_pattern = r"^[^@]+@[^@]+\.[^@]+$"
        if not re.match(email_pattern, test_email):
            return jsonify({"success": False, "message": "Invalid email format"})

        # Get template
        conn = get_db()
        template = conn.execute(
            "SELECT * FROM email_templates WHERE id = ?", (template_id,)
        ).fetchone()
        conn.close()

        if not template:
            return jsonify({"success": False, "message": "Template not found"})

        template_dict = dict(template)

        # Create sample data for testing
        sample_student = {
            "first_name": "Sarah",
            "last_name": "Johnson",
            "parent_first": "Michael",
            "parent_last": "Johnson",
            "email": test_email,
            "phone": "604-123-4567",
            "membership_level": "Silver",
            "id": 999,
        }

        sample_cancellation = {
            "lesson_date": "2024-03-15",
            "lesson_time": "15:00:00",
            "charged": False,
            "sequential_lessons": None,
            "reschedule_requested": False,
            "reschedule_preferences": "",
            "error_report": "",
            "manager_notes": "",
            "id": 999,
        }

        # Generate template variables
        variables = get_template_variables(
            sample_student,
            sample_cancellation,
            {
                "status_message": "This is a test email with sample data.",
                "test_mode": "true",
            },
        )

        # Process template
        body, subject = process_template_variables(
            template_dict["body"], template_dict["subject"], variables
        )

        # Add test disclaimer
        test_subject = f"[TEST] {subject}"
        test_body = f"""
        <div style="background-color: #fff3cd; border: 1px solid #ffeaa7; padding: 15px; margin-bottom: 20px; border-radius: 5px;">
            <h4 style="color: #856404; margin: 0 0 10px 0;">⚠️ TEST EMAIL</h4>
            <p style="margin: 0; color: #856404;">
                This is a test email with sample data from template: <strong>{template_dict['name']}</strong><br>
                Template ID: {template_id}<br>
                Sent from: {email_config.from_email}<br>
                Test sent at: {toronto_now().strftime('%Y-%m-%d %H:%M:%S')}
            </p>
        </div>
        {body}
        <hr style="margin: 30px 0;">
        <p style="font-size: 12px; color: #666;">
            <strong>Template Variables Used:</strong><br>
            {', '.join([f'{{{{{k}}}}}' for k in variables.keys() if f'{{{{{k}}}}}' in template_dict['body'] or f'{{{{{k}}}}}' in template_dict['subject']])}
        </p>
        """

        # Send test email
        result = send_email(
            test_email,
            test_subject,
            test_body,
            "test",
            template_id=f"test_{template_id}",
        )

        if result["success"]:
            log_action("test_email_sent", f"Template: {template_id}, To: {test_email}")

        return jsonify(result)

    except Exception as e:
        return jsonify({"success": False, "message": str(e)})


@app.route("/senior/api/template/<template_id>/preview", methods=["GET"])
@login_required
@senior_admin_required
def preview_template_enhanced(template_id):
    """Get template preview with sample data"""
    try:
        conn = get_db()
        template = conn.execute(
            "SELECT * FROM email_templates WHERE id = ?", (template_id,)
        ).fetchone()
        conn.close()

        if not template:
            return jsonify({"success": False, "message": "Template not found"})

        template_dict = dict(template)

        # Create sample data
        sample_variables = {
            "client_name": "Sarah Johnson",
            "client_first_name": "Sarah",
            "lesson_date": "March 15, 2024",
            "lesson_time": "3:00 PM",
            "membership_tier": "Silver",
            "cancellation_status": "Free cancellation",
            "company_name": "Riverside Equestrian",
            "current_date": toronto_now().strftime("%B %d, %Y"),
        }

        # Simple variable replacement
        preview_body = template_dict["body"]
        preview_subject = template_dict["subject"]

        for key, value in sample_variables.items():
            preview_body = preview_body.replace(f"{{{{{key}}}}}", str(value))
            preview_subject = preview_subject.replace(f"{{{{{key}}}}}", str(value))

        return jsonify(
            {
                "success": True,
                "preview": {
                    "subject": preview_subject,
                    "body": preview_body,
                    "template_name": template_dict["name"],
                    "template_type": template_dict["type"],
                },
            }
        )

    except Exception as e:
        return jsonify({"success": False, "message": str(e)})


@app.route("/senior/api/templates/export", methods=["GET"])
@login_required
@senior_admin_required
def export_templates_enhanced():
    """Export all email templates to JSON"""
    try:
        conn = get_db()
        templates = conn.execute(
            "SELECT * FROM email_templates ORDER BY name"
        ).fetchall()
        conn.close()

        # Convert to list of dictionaries
        template_list = [dict(template) for template in templates]

        # Create export data
        export_data = {
            "export_info": {
                "exported_at": datetime.now().isoformat(),
                "exported_by": session.get("user_email", "Unknown"),
                "system": "Riverside Equestrian Cancellation System",
                "version": "1.0",
            },
            "templates": template_list,
        }

        # Create JSON response
        import json

        response_data = json.dumps(export_data, indent=2, default=str)

        response = app.response_class(
            response_data,
            mimetype="application/json",
            headers={
                "Content-Disposition": f"attachment; filename=riverside_email_templates_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
            },
        )

        return response

    except Exception as e:
        return jsonify({"success": False, "message": str(e)})


@app.route("/senior/api/templates/import", methods=["POST"])
@login_required
@senior_admin_required
def import_templates_enhanced():
    """Import email templates from JSON"""
    try:
        data = request.json

        # Handle both old and new format
        if "templates" in data:
            templates = data["templates"]
        else:
            templates = data

        if not templates:
            return jsonify(
                {"success": False, "message": "No templates found in import data"}
            )

        imported_count = 0
        updated_count = 0
        errors = []

        conn = get_db()

        for template_data in templates:
            try:
                # Validate required fields
                if not all(
                    key in template_data for key in ["id", "name", "subject", "body"]
                ):
                    errors.append(
                        f"Template missing required fields: {template_data.get('id', 'unknown')}"
                    )
                    continue

                # Check if template exists
                existing = conn.execute(
                    "SELECT id FROM email_templates WHERE id = ?",
                    (template_data["id"],),
                ).fetchone()

                if existing:
                    # Update existing
                    conn.execute(
                        """
                        UPDATE email_templates 
                        SET name = ?, subject = ?, body = ?, type = ?, active = ?, 
                            auto_send = ?, priority = ?, delay_minutes = ?, 
                            include_attachment = ?, variables_used = ?, updated_at = ?
                        WHERE id = ?
                        """,
                        (
                            template_data["name"],
                            template_data["subject"],
                            template_data["body"],
                            template_data.get("type", "client"),
                            template_data.get("active", True),
                            template_data.get("auto_send", True),
                            template_data.get("priority", "normal"),
                            template_data.get("delay_minutes", 0),
                            template_data.get("include_attachment", False),
                            template_data.get("variables_used", ""),
                            toronto_now().strftime("%Y-%m-%d %H:%M:%S"),
                            template_data["id"],
                        ),
                    )
                    updated_count += 1
                else:
                    # Insert new
                    conn.execute(
                        """
                        INSERT INTO email_templates 
                        (id, name, subject, body, type, active, auto_send, priority, 
                         delay_minutes, include_attachment, variables_used, created_at, updated_at)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                        """,
                        (
                            template_data["id"],
                            template_data["name"],
                            template_data["subject"],
                            template_data["body"],
                            template_data.get("type", "client"),
                            template_data.get("active", True),
                            template_data.get("auto_send", True),
                            template_data.get("priority", "normal"),
                            template_data.get("delay_minutes", 0),
                            template_data.get("include_attachment", False),
                            template_data.get("variables_used", ""),
                            toronto_now().strftime("%Y-%m-%d %H:%M:%S"),
                            toronto_now().strftime("%Y-%m-%d %H:%M:%S"),
                        ),
                    )
                    imported_count += 1

            except Exception as e:
                errors.append(
                    f"Template {template_data.get('id', 'unknown')}: {str(e)}"
                )

        conn.commit()
        conn.close()

        return jsonify(
            {
                "success": True,
                "message": f"Import completed: {imported_count} new, {updated_count} updated",
                "imported": imported_count,
                "updated": updated_count,
                "errors": errors[:10],  # Limit error list
            }
        )

    except Exception as e:
        return jsonify({"success": False, "message": str(e)})


# ===================================
# EMAIL TRIGGER MANAGEMENT ROUTES - NEW
# ===================================

@app.route("/senior/api/triggers", methods=["GET"])
@login_required
@senior_admin_required
def get_all_triggers():
    """Get all email triggers with their current mappings"""
    try:
        conn = get_db()
        conn.row_factory = sqlite3.Row
        
        # Get all triggers
        triggers = conn.execute(
            """
            SELECT id, name, description, category, event_condition, active
            FROM email_triggers
            ORDER BY category, name
            """
        ).fetchall()
        
        result = []
        for trigger in triggers:
            trigger_dict = dict(trigger)
            
            # Get all mappings for this trigger
            mappings = conn.execute(
                """
                SELECT tm.id, tm.template_id, tm.recipient_type, tm.enabled, tm.priority,
                       et.name as template_name
                FROM email_trigger_mappings tm
                JOIN email_templates et ON tm.template_id = et.id
                WHERE tm.trigger_id = ?
                ORDER BY tm.priority, et.name
                """,
                (trigger['id'],)
            ).fetchall()
            
            trigger_dict['mappings'] = [dict(m) for m in mappings]
            result.append(trigger_dict)
        
        conn.close()
        return jsonify({"success": True, "triggers": result})
    
    except Exception as e:
        return jsonify({"success": False, "message": str(e)}), 500


@app.route("/senior/api/trigger/<trigger_id>/mappings", methods=["GET"])
@login_required
@senior_admin_required
def get_trigger_mappings(trigger_id):
    """Get all template mappings for a specific trigger"""
    try:
        conn = get_db()
        conn.row_factory = sqlite3.Row
        
        # Get trigger info
        trigger = conn.execute(
            "SELECT * FROM email_triggers WHERE id = ?",
            (trigger_id,)
        ).fetchone()
        
        if not trigger:
            conn.close()
            return jsonify({"success": False, "message": "Trigger not found"}), 404
        
        # Get all mappings
        mappings = conn.execute(
            """
            SELECT tm.id, tm.template_id, tm.recipient_type, tm.enabled, tm.priority,
                   et.name as template_name, et.type as template_type
            FROM email_trigger_mappings tm
            LEFT JOIN email_templates et ON tm.template_id = et.id
            WHERE tm.trigger_id = ?
            ORDER BY tm.priority, et.name
            """,
            (trigger_id,)
        ).fetchall()
        
        # Get all available templates
        templates = conn.execute(
            "SELECT id, name, type FROM email_templates ORDER BY type, name"
        ).fetchall()
        
        conn.close()
        
        return jsonify({
            "success": True,
            "trigger": dict(trigger),
            "mappings": [dict(m) for m in mappings],
            "available_templates": [dict(t) for t in templates]
        })
    
    except Exception as e:
        return jsonify({"success": False, "message": str(e)}), 500


@app.route("/senior/api/trigger/mapping/add", methods=["POST"])
@login_required
@senior_admin_required
def add_trigger_mapping():
    """Add a new template mapping to a trigger"""
    try:
        data = request.json
        trigger_id = data.get('trigger_id')
        template_id = data.get('template_id')
        recipient_type = data.get('recipient_type')
        
        if not all([trigger_id, template_id, recipient_type]):
            return jsonify({"success": False, "message": "Missing required fields"}), 400
        
        conn = get_db()
        
        # Check if mapping already exists
        existing = conn.execute(
            """
            SELECT id FROM email_trigger_mappings 
            WHERE trigger_id = ? AND template_id = ? AND recipient_type = ?
            """,
            (trigger_id, template_id, recipient_type)
        ).fetchone()
        
        if existing:
            conn.close()
            return jsonify({"success": False, "message": "This mapping already exists"}), 400
        
        # Add mapping
        cursor = conn.execute(
            """
            INSERT INTO email_trigger_mappings 
            (trigger_id, template_id, recipient_type, enabled, priority)
            VALUES (?, ?, ?, 1, 0)
            """,
            (trigger_id, template_id, recipient_type)
        )
        
        conn.commit()
        mapping_id = cursor.lastrowid
        conn.close()
        
        log_action("trigger_mapping_added", f"Added mapping for trigger {trigger_id}")
        return jsonify({"success": True, "mapping_id": mapping_id})
    
    except Exception as e:
        return jsonify({"success": False, "message": str(e)}), 500


@app.route("/senior/api/trigger/mapping/<int:mapping_id>/remove", methods=["DELETE"])
@login_required
@senior_admin_required
def remove_trigger_mapping(mapping_id):
    """Remove a template mapping from a trigger"""
    try:
        conn = get_db()
        
        # Get mapping info for logging
        mapping = conn.execute(
            "SELECT trigger_id FROM email_trigger_mappings WHERE id = ?",
            (mapping_id,)
        ).fetchone()
        
        if not mapping:
            conn.close()
            return jsonify({"success": False, "message": "Mapping not found"}), 404
        
        conn.execute("DELETE FROM email_trigger_mappings WHERE id = ?", (mapping_id,))
        conn.commit()
        conn.close()
        
        log_action("trigger_mapping_removed", f"Removed mapping {mapping_id}")
        return jsonify({"success": True})
    
    except Exception as e:
        return jsonify({"success": False, "message": str(e)}), 500


@app.route("/senior/api/trigger/mapping/<int:mapping_id>/toggle", methods=["POST"])
@login_required
@senior_admin_required
def toggle_trigger_mapping(mapping_id):
    """Toggle a trigger mapping enabled/disabled"""
    try:
        conn = get_db()
        
        # Get current status
        mapping = conn.execute(
            "SELECT enabled FROM email_trigger_mappings WHERE id = ?",
            (mapping_id,)
        ).fetchone()
        
        if not mapping:
            conn.close()
            return jsonify({"success": False, "message": "Mapping not found"}), 404
        
        new_status = 1 - mapping['enabled']
        conn.execute(
            "UPDATE email_trigger_mappings SET enabled = ? WHERE id = ?",
            (new_status, mapping_id)
        )
        
        conn.commit()
        conn.close()
        
        log_action("trigger_mapping_toggled", f"Toggled mapping {mapping_id} to {new_status}")
        return jsonify({"success": True, "enabled": new_status})
    
    except Exception as e:
        return jsonify({"success": False, "message": str(e)}), 500


@app.route("/senior/api/trigger/mapping/<int:mapping_id>/priority", methods=["POST"])
@login_required
@senior_admin_required
def update_mapping_priority(mapping_id):
    """Update priority of a trigger mapping"""
    try:
        data = request.json
        priority = data.get('priority')
        
        if priority is None:
            return jsonify({"success": False, "message": "Priority value required"}), 400
        
        conn = get_db()
        conn.execute(
            "UPDATE email_trigger_mappings SET priority = ? WHERE id = ?",
            (priority, mapping_id)
        )
        conn.commit()
        conn.close()
        
        return jsonify({"success": True})
    
    except Exception as e:
        return jsonify({"success": False, "message": str(e)}), 500


@app.route("/senior/triggers")
@login_required
@senior_admin_required
def manage_triggers():
    """Email triggers management page"""
    try:
        conn = get_db()
        
        # Get all triggers
        triggers = conn.execute(
            """
            SELECT id, name, description, category, active
            FROM email_triggers
            ORDER BY category, name
            """
        ).fetchall()
        
        conn.close()
        
        return render_template("email_triggers.html", triggers=[dict(t) for t in triggers])
    
    except Exception as e:
        return render_template("email_triggers.html", triggers=[], error=str(e))


# ===================================
# UPDATE EXISTING SENIOR TEMPLATES ROUTE
# ===================================


@app.route("/senior/templates")
@login_required
@senior_admin_required
def senior_templates():
    """Senior manager email templates page"""
    try:
        conn = get_db()

        # Get all email templates
        templates = conn.execute(
            """
            SELECT id, name, subject, body, type, active, auto_send, priority, 
                   delay_minutes, include_attachment, variables_used, created_at, updated_at
            FROM email_templates 
            ORDER BY active DESC, type, name
        """
        ).fetchall()

        conn.close()

        # Convert to list of dicts for template
        template_list = []
        for template in templates:
            template_dict = dict(template)

            # Parse dates properly
            for date_field in ["created_at", "updated_at"]:
                if template_dict.get(date_field):
                    try:
                        template_dict[date_field] = datetime.strptime(
                            template_dict[date_field], "%Y-%m-%d %H:%M:%S"
                        )
                    except (ValueError, TypeError):
                        template_dict[date_field] = toronto_now()

            template_list.append(template_dict)

        return render_template("senior_templates.html", templates=template_list)

    except Exception as e:
        print(f"Error in senior_templates: {e}")
        return render_template("senior_templates.html", templates=[], error=str(e))


# Senior Manager API endpoints
@app.route("/senior/save-settings", methods=["POST"])
@login_required
@senior_admin_required
def save_settings():
    """Save system settings"""
    data = request.json

    conn = get_db()

    try:
        # Update each setting
        for form_name, form_data in data.items():
            for key, value in form_data.items():
                conn.execute(
                    "INSERT OR REPLACE INTO system_settings (key, value, updated_at) VALUES (?, ?, ?)",
                    (key, str(value), toronto_now().isoformat()),
                )

        conn.commit()
        conn.close()

        log_action("settings_updated", f"Updated settings: {list(data.keys())}")
        return jsonify({"success": True})
    except Exception as e:
        conn.close()
        return jsonify({"success": False, "error": str(e)})


@app.route("/senior/save-tiers", methods=["POST"])
@login_required
@senior_admin_required
def save_tiers():
    """Save membership tiers"""
    data = request.json
    tiers = data.get("tiers", [])

    conn = get_db()

    try:
        for tier in tiers:
            # Calculate deadline_hours based on deadline type
            deadline_mapping = {
                "6pm_previous_day": 18,
                "2_hours_before": 2,
                "24_hours_before": 24,
                "same_day": 0,
            }
            deadline_hours = deadline_mapping.get(tier["deadline"], 18)

            conn.execute(
                """
                UPDATE membership_tiers 
                SET free_notices = ?, deadline_hours = ?, deadline_display = ?, active = ?
                WHERE level = ?
            """,
                (
                    tier["limit"],
                    deadline_hours,
                    tier["deadline"].replace("_", " "),
                    1 if tier["status"] == "active" else 0,
                    tier["name"],
                ),
            )

        conn.commit()
        conn.close()

        log_action("tiers_updated", f"Updated {len(tiers)} tiers")
        return jsonify({"success": True})
    except Exception as e:
        conn.close()
        return jsonify({"success": False, "error": str(e)})


@app.route("/senior/save-template", methods=["POST"])
@login_required
@senior_admin_required
def save_template():
    """Save email template"""
    data = request.json

    conn = get_db()

    try:
        conn.execute(
            """
            INSERT OR REPLACE INTO email_templates 
            (id, name, subject, body, type, active, auto_send, priority, delay_minutes, include_attachment, updated_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
            (
                data["id"],
                data["id"].replace("_", " ").title(),
                data["subject"],
                data["body"],
                data["type"],
                data["active"],
                data["autoSend"],
                data["priority"],
                data["delay"],
                data["includeAttachment"],
                toronto_now().isoformat(),
            ),
        )

        conn.commit()
        conn.close()

        log_action("template_updated", f"Template: {data['id']}")
        return jsonify({"success": True})
    except Exception as e:
        conn.close()
        return jsonify({"success": False, "error": str(e)})


@app.route("/senior/migrate-students", methods=["POST"])
@login_required
@senior_admin_required
def migrate_students():
    """Migrate students between tiers"""
    data = request.json
    from_tier = data.get("from_tier")
    to_tier = data.get("to_tier")

    conn = get_db()

    try:
        # Count students to migrate
        count = conn.execute(
            "SELECT COUNT(*) as count FROM students WHERE membership_level = ?",
            (from_tier,),
        ).fetchone()["count"]

        # Perform migration
        conn.execute(
            "UPDATE students SET membership_level = ? WHERE membership_level = ?",
            (to_tier, from_tier),
        )

        conn.commit()
        conn.close()

        log_action(
            "students_migrated",
            f"Migrated {count} students from {from_tier} to {to_tier}",
        )
        return jsonify({"success": True, "count": count})
    except Exception as e:
        conn.close()
        return jsonify({"success": False, "error": str(e)})


@app.route("/senior/create-backup", methods=["POST"])
@login_required
@senior_admin_required
def create_backup():
    """Create system backup"""
    try:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_filename = f"backup_{timestamp}.sql"

        # In a real implementation, you would create an actual database backup
        # For this demo, we'll just return a success message

        log_action("backup_created", f"Backup file: {backup_filename}")
        return jsonify({"success": True, "filename": backup_filename})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})


@app.route("/test-email")
def test_email_route():
    result = test_email_configuration()
    if result["success"]:
        # Send a test email to yourself
        test_result = send_email(
            "stav@riversideequestrian.ca",
            "Test Email from Riverside System",
            "<h1>Success!</h1><p>Your email system is working correctly!</p>",
            template_id="test",
        )
        return (
            f"Email config: {result['message']}<br>Test email: {test_result['message']}"
        )
    else:
        return f"Email config failed: {result['message']}"


@app.route("/test-timezone")
def test_timezone_route():
    """Debug route to test timezone functionality"""
    import time

    pacific_time = toronto_now()
    system_time = datetime.now()
    utc_time = datetime.utcnow()

    html_output = f"""
    <h2>Timezone Debug Information</h2>
    <p><strong>Current Pacific Time:</strong> {pacific_time}</p>
    <p><strong>Formatted Pacific Time:</strong> {format_datetime_for_display(pacific_time)}</p>
    <p><strong>System Time:</strong> {system_time}</p>
    <p><strong>UTC Time:</strong> {utc_time}</p>
    <p><strong>Pacific Timezone:</strong> {pacific_time.tzname()}</p>
    <p><strong>UTC Offset:</strong> {pacific_time.strftime('%z')}</p>
    <p><strong>Is DST Active:</strong> {'Yes (PDT)' if pacific_time.dst().total_seconds() > 0 else 'No (PST)'}</p>
    <p><strong>System TZ:</strong> {time.tzname}</p>
    <hr>
    <p><em>The Pacific time should be 8 hours behind UTC in winter (PST) or 7 hours behind in summer (PDT).</em></p>
    """

    return html_output


# ===================================
# API ENDPOINTS FOR AJAX CALLS
# ===================================


@app.route("/api/preview-cancellation", methods=["POST"])
@login_required
def preview_cancellation():
    """Preview cancellation status before submission"""
    if session["user_role"] != "client":
        return jsonify({"error": "Unauthorized"}), 403
 
    data = request.json
    lesson_date = data.get("lesson_date")
    lesson_time = data.get("lesson_time")
 
    if not lesson_date or not lesson_time:
        return jsonify({"error": "Missing date or time"}), 400
 
    try:
        # Parse and create timezone-aware datetime
        lesson_datetime = parse_lesson_datetime(lesson_date, lesson_time)
    except ValueError:
        return jsonify({"error": "Invalid date or time format"}), 400
 
    # Get student info
    conn = get_db()
    student = conn.execute(
        "SELECT * FROM students WHERE id = ?", (session["user_id"],)
    ).fetchone()
    conn.close()
 
    if not student:
        return jsonify({"error": "Student not found"}), 404
 
    # Get current time for comparison
    current_time = toronto_now()
    
    # Check if lesson is in the past
    same_day_past_lesson = False
    cannot_cancel = False
    
    if lesson_datetime <= current_time:
        # Lesson has passed - check if it's the same day
        lesson_date_obj = lesson_datetime.date()
        current_date_obj = current_time.date()
        
        if lesson_date_obj == current_date_obj:
            # Same-day past lesson - will be charged
            same_day_past_lesson = True
            will_charge = True
            reason = "Same-day cancellation after lesson time (charged)"
        else:
            # Past lesson from earlier day - cannot cancel
            cannot_cancel = True
            will_charge = True
            reason = "Cannot cancel past lessons"
    else:
        # Normal future lesson - use standard charge logic
        will_charge, reason = will_be_charged(student, lesson_datetime)
 
    # Calculate status
    status = calculate_cancellation_status(student)
 
    return jsonify({
        "will_be_charged": will_charge,
        "reason": reason,
        "current_status": status,
        "same_day_past_lesson": same_day_past_lesson,
        "cannot_cancel": cannot_cancel
    })

# ===================================
# TEMPLATE CONTEXT PROCESSORS
# ===================================


@app.context_processor
def inject_now():
    """Inject current datetime into all templates"""
    return {"now": toronto_now()}


# Replace your current inject_moment context processor with this:


@app.context_processor
def inject_moment():
    """Inject moment-like function for templates"""

    def moment_func(date_str):
        """Format relative time like moment.js"""
        if not date_str:
            return "Unknown"

        try:
            # Parse the date string using localize_datetime to handle both formats
            if isinstance(date_str, str):
                date_obj = localize_datetime(date_str)
                if date_obj is None:
                    return str(date_str)
            else:
                date_obj = date_str
                # Ensure it's timezone-aware
                if isinstance(date_obj, datetime) and date_obj.tzinfo is None:
                    date_obj = localize_datetime(date_obj)

            now = toronto_now()

            # Make sure both are timezone-aware before comparison
            if date_obj.tzinfo is None:
                date_obj = localize_datetime(date_obj)

            time_diff = now - date_obj

            total_seconds = int(time_diff.total_seconds())

            if total_seconds < 60:
                return "Just now"
            elif total_seconds < 3600:
                minutes = total_seconds // 60
                return f"{minutes} minute{'s' if minutes != 1 else ''} ago"
            elif total_seconds < 86400:
                hours = total_seconds // 3600
                return f"{hours} hour{'s' if hours != 1 else ''} ago"
            elif total_seconds < 2592000:  # 30 days
                days = total_seconds // 86400
                return f"{days} day{'s' if days != 1 else ''} ago"
            else:
                return date_obj.strftime("%Y-%m-%d")

        except (ValueError, TypeError) as e:
            return str(date_str)

    return {"moment": moment_func}


# ===================================
# STATIC FILE FIXES
# ===================================


@app.route("/static/js/main.js")
def serve_main_js():
    """Serve main.js file"""
    js_content = """
// Main JavaScript file for Riverside Equestrian
console.log('Riverside Equestrian Cancellation System loaded');

// Global utilities
window.RiversideUtils = {
    formatDate: function(date) {
        return new Date(date).toLocaleDateString();
    },
    
    formatTime: function(time) {
        return new Date('1970-01-01T' + time).toLocaleTimeString([], {hour: '2-digit', minute:'2-digit'});
    },
    
    showAlert: function(message, type = 'info') {
        const alertDiv = document.createElement('div');
        alertDiv.className = `alert alert-${type} alert-dismissible fade show position-fixed`;
        alertDiv.style.cssText = 'top: 20px; right: 20px; z-index: 9999; min-width: 300px;';
        alertDiv.innerHTML = `
            ${message}
            <button type="button" class="btn-close" data-bs-dismiss="alert"></button>
        `;
        document.body.appendChild(alertDiv);
        
        setTimeout(() => {
            if (alertDiv.parentNode) {
                alertDiv.remove();
            }
        }, 5000);
    }
};

// Initialize tooltips if Bootstrap is available
document.addEventListener('DOMContentLoaded', function() {
    if (typeof bootstrap !== 'undefined' && bootstrap.Tooltip) {
        var tooltipTriggerList = [].slice.call(document.querySelectorAll('[data-bs-toggle="tooltip"]'));
        var tooltipList = tooltipTriggerList.map(function (tooltipTriggerEl) {
            return new bootstrap.Tooltip(tooltipTriggerEl);
        });
    }
});
"""
    return app.response_class(js_content, mimetype="application/javascript")


@app.route("/favicon.ico")
def serve_favicon():
    """Serve favicon or return 204 No Content"""
    return "", 204


@app.route("/manager/export/cancellations")
@login_required
@admin_required
def export_cancellations_excel():
    """Export cancellations with current filters to Excel"""
    try:
        conn = get_db()

        # Get filter parameters from the request - same as manager_cancellations route
        filter_status = request.args.get("status", "")
        search = request.args.get("search", "")
        date_range = request.args.get("date_range", "month")
        membership = request.args.get("membership", "")
        sort_by = request.args.get("sort", "submit_date")
        student_id = request.args.get("student")
        date_from = request.args.get("date_from")
        date_to = request.args.get("date_to")

        # Build the same query as manager_cancellations route
        where_clauses = []
        params = []

        # Add student filter
        if student_id:
            where_clauses.append("c.student_id = ?")
            params.append(student_id)

        # Status filters
        if filter_status == "free":
            where_clauses.append("c.charged = 0 AND c.status = 'approved'")
        elif filter_status == "charged":
            where_clauses.append("c.charged = 1 AND c.excluded = 0")
        elif filter_status == "excluded":
            where_clauses.append("c.excluded = 1")
        elif filter_status == "note":
            where_clauses.append(
                "c.cancellation_note IS NOT NULL AND c.cancellation_note != ''"
            )
        elif filter_status == "deadline_passed":
            where_clauses.append(
                """
        (
            (s.membership_level = 'Gold' AND 
             datetime(c.lesson_date || ' ' || c.lesson_time) <= datetime(c.created_at, '+2 hours'))
            OR
            (s.membership_level != 'Gold' AND 
             datetime(c.lesson_date || ' 18:00:00', '-1 day') <= datetime(c.created_at))
        )
    """
            )
        elif filter_status == "override":
            where_clauses.append("c.is_override = 1")

        # Search filter
        if search:
            where_clauses.append(
                "(s.first_name LIKE ? OR s.last_name LIKE ? OR s.email LIKE ?)"
            )
            search_param = f"%{search}%"
            params.extend([search_param, search_param, search_param])

        # Date range filter
        if date_range == "today":
            where_clauses.append("DATE(c.created_at) = DATE('now')")
        elif date_range == "yesterday":
            where_clauses.append("DATE(c.created_at) = DATE('now', '-1 day')")
        elif date_range == "7days":
            where_clauses.append("c.created_at >= DATE('now', '-7 days')")
        elif date_range == "month":
            where_clauses.append("c.created_at >= DATE('now', '-30 days')")
        elif date_range == "all":
            pass  # No date filter
        elif date_range == "custom" and date_from and date_to:
            where_clauses.append("DATE(c.created_at) BETWEEN ? AND ?")
            params.extend([date_from, date_to])

        # Membership filter
        if membership:
            where_clauses.append("s.membership_level = ?")
            params.append(membership)

        where_sql = " AND ".join(where_clauses) if where_clauses else "1=1"

        # Sort order
        if sort_by == "submit_date":
            order_by = "c.created_at DESC"
        elif sort_by == "lesson_date":
            order_by = "c.lesson_date DESC"
        elif sort_by == "student":
            order_by = "s.last_name, s.first_name"
        elif sort_by == "status":
            order_by = "c.status, c.charged, c.excluded"
        else:
            order_by = "c.created_at DESC"

        # Get cancellations data with the same query as the page
        cancellations = conn.execute(
            f"""
            SELECT 
                c.id,
                s.first_name,
                s.last_name,
                s.parent_first,
                s.parent_last,
                s.email,
                s.membership_level,
                c.lesson_date,
                c.lesson_time,
                c.created_at,
                c.charged,
                c.excluded,
                c.status,
                c.deadline_passed,
                c.is_override,
                c.manager_notes,
                c.cancellation_note
            FROM cancellations c
            JOIN students s ON c.student_id = s.id
            WHERE {where_sql}
            ORDER BY {order_by}
        """,
            params,
        ).fetchall()

        conn.close()

        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Cancellations Export"

        # Headers
        headers = [
            "ID",
            "Student First Name",
            "Student Last Name",
            "Parent First Name",
            "Parent Last Name",
            "Email",
            "Membership Level",
            "Lesson Date",
            "Lesson Time",
            "Submitted Date",
            "Status",
            "Cancellation Note",
            "Manager Notes",
        ]

        # Add headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(
                start_color="E6F3FF", end_color="E6F3FF", fill_type="solid"
            )

        # Add data
        for row, cancellation in enumerate(cancellations, 2):
            ws.cell(row=row, column=1, value=cancellation["id"])
            ws.cell(row=row, column=2, value=cancellation["first_name"])
            ws.cell(row=row, column=3, value=cancellation["last_name"])
            ws.cell(row=row, column=4, value=cancellation["parent_first"] or "")
            ws.cell(row=row, column=5, value=cancellation["parent_last"] or "")
            ws.cell(row=row, column=6, value=cancellation["email"])
            ws.cell(row=row, column=7, value=cancellation["membership_level"])
            ws.cell(row=row, column=8, value=cancellation["lesson_date"])
            ws.cell(row=row, column=9, value=cancellation["lesson_time"])

            # Format submitted date
            submitted_date = datetime.fromisoformat(
                cancellation["created_at"]
            ).strftime("%Y-%m-%d %H:%M")
            ws.cell(row=row, column=10, value=submitted_date)

            # Determine status
            if cancellation["excluded"]:
                status = "Excluded"
            elif cancellation["charged"]:
                status = "Charged"
            elif cancellation["status"] == "approved":
                status = "Free"
            elif cancellation["deadline_passed"]:
                status = "Deadline Passed"
            elif cancellation["is_override"]:
                status = "Override"
            else:
                status = "Pending"

            ws.cell(row=row, column=11, value=status)
            ws.cell(row=row, column=12, value=cancellation["cancellation_note"] or "")
            ws.cell(row=row, column=13, value=cancellation["manager_notes"] or "")

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = max(len(str(cell.value or "")) for cell in column)
            ws.column_dimensions[column[0].column_letter].width = min(
                max_length + 2, 50
            )

        # Save to temporary file
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
        wb.save(temp_file.name)
        temp_file.close()

        # Generate filename with current filters
        filter_parts = []
        if filter_status:
            filter_parts.append(filter_status)
        if date_range and date_range != "all":
            filter_parts.append(date_range)
        if search:
            filter_parts.append("search")

        filename_suffix = "_".join(filter_parts) if filter_parts else "all"
        filename = (
            f'cancellations_{filename_suffix}_{datetime.now().strftime("%Y%m%d")}.xlsx'
        )

        log_action(
            "export_cancellations", f"Exported filtered cancellations: {filename}"
        )

        return send_file(
            temp_file.name,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=filename,
        )

    except Exception as e:
        log_action("export_error", f"Error exporting filtered cancellations: {str(e)}")
        return jsonify({"success": False, "message": f"Export failed: {str(e)}"}), 500


@app.route("/manager/export/cancellations/monthly")
@login_required
@admin_required
def export_cancellations_monthly_excel():
    """Export monthly cancellation report to Excel"""
    try:
        conn = get_db()

        # Get current month and previous months data
        current_date = toronto_now()
        start_date = current_date.replace(day=1) - relativedelta(months=5)

        # Get cancellations data
        cancellations = conn.execute(
            """
            SELECT 
                c.id,
                s.first_name,
                s.last_name,
                s.parent_first,
                s.parent_last,
                s.email,
                s.membership_level,
                c.lesson_date,
                c.lesson_time,
                c.created_at,
                c.charged,
                c.excluded,
                c.manager_notes,
                c.cancellation_note
            FROM cancellations c
            JOIN students s ON c.student_id = s.id
            WHERE c.created_at >= ?
            ORDER BY c.created_at DESC
        """,
            (start_date.strftime("%Y-%m-%d"),),
        ).fetchall()

        conn.close()

        # Create Excel workbook
        wb = Workbook()

        # Create summary sheet
        ws_summary = wb.active
        ws_summary.title = "Monthly Summary"

        # Summary headers
        headers = [
            "Month",
            "Total Cancellations",
            "Free",
            "Charged",
            "Excluded",
            "Bronze Members",
            "Silver Members",
            "Gold Members",
        ]

        for col, header in enumerate(headers, 1):
            cell = ws_summary.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(
                start_color="366092", end_color="366092", fill_type="solid"
            )
            cell.alignment = Alignment(horizontal="center")

        # Calculate monthly summaries
        monthly_data = {}
        for cancellation in cancellations:
            # Use localize_datetime to handle both old and new timestamp formats
            created_dt = localize_datetime(cancellation["created_at"])
            if created_dt is None:
                continue  # Skip if parsing fails
            month_key = created_dt.strftime("%Y-%m")
            if month_key not in monthly_data:
                monthly_data[month_key] = {
                    "total": 0,
                    "free": 0,
                    "charged": 0,
                    "excluded": 0,
                    "bronze": 0,
                    "silver": 0,
                    "gold": 0,
                }

            monthly_data[month_key]["total"] += 1
            if cancellation["excluded"]:
                monthly_data[month_key]["excluded"] += 1
            elif cancellation["charged"]:
                monthly_data[month_key]["charged"] += 1
            else:
                monthly_data[month_key]["free"] += 1

            # Count by membership
            membership = cancellation["membership_level"].lower()
            if membership == "bronze":
                monthly_data[month_key]["bronze"] += 1
            elif membership == "silver":
                monthly_data[month_key]["silver"] += 1
            elif membership == "gold":
                monthly_data[month_key]["gold"] += 1

        # Add summary data
        row = 2
        for month in sorted(monthly_data.keys(), reverse=True):
            data = monthly_data[month]
            month_name = datetime.strptime(month, "%Y-%m").strftime("%B %Y")

            ws_summary.cell(row=row, column=1, value=month_name)
            ws_summary.cell(row=row, column=2, value=data["total"])
            ws_summary.cell(row=row, column=3, value=data["free"])
            ws_summary.cell(row=row, column=4, value=data["charged"])
            ws_summary.cell(row=row, column=5, value=data["excluded"])
            ws_summary.cell(row=row, column=6, value=data["bronze"])
            ws_summary.cell(row=row, column=7, value=data["silver"])
            ws_summary.cell(row=row, column=8, value=data["gold"])
            row += 1

        # Auto-adjust column widths
        for column in ws_summary.columns:
            max_length = max(len(str(cell.value or "")) for cell in column)
            ws_summary.column_dimensions[column[0].column_letter].width = max_length + 2

        # Create detailed sheet
        ws_detail = wb.create_sheet("Detailed Cancellations")

        # Detailed headers
        detail_headers = [
            "ID",
            "Student Name",
            "Parent Name",
            "Email",
            "Membership",
            "Lesson Date",
            "Lesson Time",
            "Submitted Date",
            "Status",
            "Cancellation Note",
            "Manager Notes",
        ]

        for col, header in enumerate(detail_headers, 1):
            cell = ws_detail.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(
                start_color="366092", end_color="366092", fill_type="solid"
            )
            cell.alignment = Alignment(horizontal="center")

        # Add detailed data
        for row, cancellation in enumerate(cancellations, 2):
            ws_detail.cell(row=row, column=1, value=cancellation["id"])

            student_name = f"{cancellation['first_name']} {cancellation['last_name']}"
            ws_detail.cell(row=row, column=2, value=student_name)

            parent_name = ""
            if cancellation["parent_first"] or cancellation["parent_last"]:
                parent_name = f"{cancellation['parent_first'] or ''} {cancellation['parent_last'] or ''}".strip()
            ws_detail.cell(row=row, column=3, value=parent_name)

            ws_detail.cell(row=row, column=4, value=cancellation["email"])
            ws_detail.cell(row=row, column=5, value=cancellation["membership_level"])
            ws_detail.cell(row=row, column=6, value=cancellation["lesson_date"])
            ws_detail.cell(row=row, column=7, value=cancellation["lesson_time"])

            submitted_date = datetime.fromisoformat(
                cancellation["created_at"]
            ).strftime("%Y-%m-%d %H:%M")
            ws_detail.cell(row=row, column=8, value=submitted_date)

            if cancellation["excluded"]:
                status = "Excluded"
            elif cancellation["charged"]:
                status = "Charged"
            else:
                status = "Free"
            ws_detail.cell(row=row, column=9, value=status)

            ws_detail.cell(
                row=row, column=10, value=cancellation["cancellation_note"] or ""
            )
            ws_detail.cell(
                row=row, column=11, value=cancellation["manager_notes"] or ""
            )

        # Auto-adjust column widths for detailed sheet
        for column in ws_detail.columns:
            max_length = max(len(str(cell.value or "")) for cell in column)
            ws_detail.column_dimensions[column[0].column_letter].width = min(
                max_length + 2, 50
            )

        # Save to temporary file
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
        wb.save(temp_file.name)
        temp_file.close()

        log_action("export_cancellations", f"Exported cancellations report")

        return send_file(
            temp_file.name,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=f'cancellations_report_{datetime.now().strftime("%Y%m%d")}.xlsx',
        )

    except Exception as e:
        log_action("export_error", f"Error exporting cancellations: {str(e)}")
        return jsonify({"success": False, "message": f"Export failed: {str(e)}"}), 500


@app.route("/manager/export/students")
@login_required
@admin_required
def export_students_excel():
    """Export students to Excel with enhanced formatting"""
    try:
        conn = get_db()

        # Get students data with cancellation statistics
        students = conn.execute(
            """
            SELECT 
                s.*,
                COUNT(c.id) as total_cancellations,
                SUM(CASE WHEN c.charged = 0 AND c.excluded = 0 THEN 1 ELSE 0 END) as free_cancellations,
                SUM(CASE WHEN c.charged = 1 THEN 1 ELSE 0 END) as charged_cancellations,
                SUM(CASE WHEN c.excluded = 1 THEN 1 ELSE 0 END) as excluded_cancellations,
                MAX(c.created_at) as last_cancellation_date
            FROM students s
            LEFT JOIN cancellations c ON s.id = c.student_id
            GROUP BY s.id
            ORDER BY s.last_name, s.first_name
        """
        ).fetchall()

        conn.close()

        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Students Directory"

        # Headers
        headers = [
            "ID",
            "First Name",
            "Last Name",
            "Parent First",
            "Parent Last",
            "Email",
            "Phone",
            "Membership Level",
            "Created Date",
            "Total Cancellations",
            "Free",
            "Charged",
            "Excluded",
            "Last Cancellation",
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(
                start_color="366092", end_color="366092", fill_type="solid"
            )
            cell.alignment = Alignment(horizontal="center")

        # Add student data
        for row, student in enumerate(students, 2):
            ws.cell(row=row, column=1, value=student["id"])
            ws.cell(row=row, column=2, value=student["first_name"])
            ws.cell(row=row, column=3, value=student["last_name"])
            ws.cell(row=row, column=4, value=student["parent_first"] or "")
            ws.cell(row=row, column=5, value=student["parent_last"] or "")
            ws.cell(row=row, column=6, value=student["email"])
            ws.cell(row=row, column=7, value=student["phone"] or "")
            ws.cell(row=row, column=8, value=student["membership_level"])

            created_date = datetime.fromisoformat(student["created_at"]).strftime(
                "%Y-%m-%d"
            )
            ws.cell(row=row, column=9, value=created_date)

            ws.cell(row=row, column=10, value=student["total_cancellations"])
            ws.cell(row=row, column=11, value=student["free_cancellations"])
            ws.cell(row=row, column=12, value=student["charged_cancellations"])
            ws.cell(row=row, column=13, value=student["excluded_cancellations"])

            last_cancellation = ""
            if student["last_cancellation_date"]:
                last_cancellation = datetime.fromisoformat(
                    student["last_cancellation_date"]
                ).strftime("%Y-%m-%d")
            ws.cell(row=row, column=14, value=last_cancellation)

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = max(len(str(cell.value or "")) for cell in column)
            ws.column_dimensions[column[0].column_letter].width = min(
                max_length + 2, 50
            )

        # Save to temporary file
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
        wb.save(temp_file.name)
        temp_file.close()

        log_action("export_students", f"Exported {len(students)} students")

        return send_file(
            temp_file.name,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=f'students_directory_{datetime.now().strftime("%Y%m%d")}.xlsx',
        )

    except Exception as e:
        log_action("export_error", f"Error exporting students: {str(e)}")
        return jsonify({"success": False, "message": f"Export failed: {str(e)}"}), 500


@app.route("/manager/export/charged-cancellations")
@login_required
@admin_required
def export_charged_cancellations():
    """Export charged cancellations report to Excel"""
    try:
        conn = get_db()

        # Get charged cancellations data
        charged_cancellations = conn.execute(
            """
            SELECT 
                c.id,
                s.first_name,
                s.last_name,
                s.parent_first,
                s.parent_last,
                s.email,
                s.phone,
                s.membership_level,
                c.lesson_date,
                c.lesson_time,
                c.created_at,
                c.cancellation_note,
                c.manager_notes
            FROM cancellations c
            JOIN students s ON c.student_id = s.id
            WHERE c.charged = 1
            ORDER BY c.created_at DESC
        """
        ).fetchall()

        conn.close()

        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Charged Cancellations"

        # Headers
        headers = [
            "Cancellation ID",
            "Student Name",
            "Parent Name",
            "Email",
            "Phone",
            "Membership Level",
            "Lesson Date",
            "Lesson Time",
            "Submitted Date",
            "Cancellation Note",
            "Manager Notes",
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(
                start_color="F59E0B", end_color="F59E0B", fill_type="solid"
            )
            cell.alignment = Alignment(horizontal="center")

        # Add data
        for row, cancellation in enumerate(charged_cancellations, 2):
            ws.cell(row=row, column=1, value=cancellation["id"])

            student_name = f"{cancellation['first_name']} {cancellation['last_name']}"
            ws.cell(row=row, column=2, value=student_name)

            parent_name = ""
            if cancellation["parent_first"] or cancellation["parent_last"]:
                parent_name = f"{cancellation['parent_first'] or ''} {cancellation['parent_last'] or ''}".strip()
            ws.cell(row=row, column=3, value=parent_name)

            ws.cell(row=row, column=4, value=cancellation["email"])
            ws.cell(row=row, column=5, value=cancellation["phone"] or "")
            ws.cell(row=row, column=6, value=cancellation["membership_level"])
            ws.cell(row=row, column=7, value=cancellation["lesson_date"])
            ws.cell(row=row, column=8, value=cancellation["lesson_time"])

            submitted_date = datetime.fromisoformat(
                cancellation["created_at"]
            ).strftime("%Y-%m-%d %H:%M")
            ws.cell(row=row, column=9, value=submitted_date)

            ws.cell(row=row, column=10, value=cancellation["cancellation_note"] or "")
            ws.cell(row=row, column=11, value=cancellation["manager_notes"] or "")

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = max(len(str(cell.value or "")) for cell in column)
            ws.column_dimensions[column[0].column_letter].width = min(
                max_length + 2, 50
            )

        # Add summary at top
        ws.insert_rows(1, 2)
        ws.cell(
            row=1,
            column=1,
            value=f"Charged Cancellations Report - Generated {toronto_now().strftime('%Y-%m-%d %H:%M')}",
        )
        ws.cell(row=1, column=1).font = Font(bold=True, size=14)
        ws.cell(
            row=2,
            column=1,
            value=f"Total Charged Cancellations: {len(charged_cancellations)}",
        )
        ws.cell(row=2, column=1).font = Font(bold=True)

        # Save to temporary file
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
        wb.save(temp_file.name)
        temp_file.close()

        log_action(
            "export_charged",
            f"Exported {len(charged_cancellations)} charged cancellations",
        )

        return send_file(
            temp_file.name,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=f'charged_cancellations_{datetime.now().strftime("%Y%m%d")}.xlsx',
        )

    except Exception as e:
        log_action("export_error", f"Error exporting charged cancellations: {str(e)}")
        return jsonify({"success": False, "message": f"Export failed: {str(e)}"}), 500


@app.route("/manager/export/analytics")
@login_required
@admin_required
def export_full_analytics():
    """Export comprehensive analytics report to Excel with multiple sheets"""
    try:
        conn = get_db()

        # Create Excel workbook with multiple sheets
        wb = Workbook()

        # 1. Executive Summary Sheet
        ws_summary = wb.active
        ws_summary.title = "Executive Summary"

        # Get overall statistics
        current_month = toronto_now().replace(day=1)
        last_month = current_month - relativedelta(months=1)

        stats = {}

        # Current month stats
        stats["current_month"] = conn.execute(
            """
            SELECT 
                COUNT(*) as total,
                SUM(CASE WHEN charged = 0 AND excluded = 0 THEN 1 ELSE 0 END) as free,
                SUM(CASE WHEN charged = 1 THEN 1 ELSE 0 END) as charged,
                SUM(CASE WHEN excluded = 1 THEN 1 ELSE 0 END) as excluded
            FROM cancellations 
            WHERE created_at >= ? AND created_at < ?
        """,
            (
                current_month.strftime("%Y-%m-%d"),
                (current_month + relativedelta(months=1)).strftime("%Y-%m-%d"),
            ),
        ).fetchone()

        # Previous month stats
        stats["last_month"] = conn.execute(
            """
            SELECT 
                COUNT(*) as total,
                SUM(CASE WHEN charged = 0 AND excluded = 0 THEN 1 ELSE 0 END) as free,
                SUM(CASE WHEN charged = 1 THEN 1 ELSE 0 END) as charged,
                SUM(CASE WHEN excluded = 1 THEN 1 ELSE 0 END) as excluded
            FROM cancellations 
            WHERE created_at >= ? AND created_at < ?
        """,
            (last_month.strftime("%Y-%m-%d"), current_month.strftime("%Y-%m-%d")),
        ).fetchone()

        # Student counts by membership
        membership_stats = conn.execute(
            """
            SELECT membership_level, COUNT(*) as count
            FROM students
            GROUP BY membership_level
            ORDER BY count DESC
        """
        ).fetchall()

        # Executive Summary content
        summary_data = [
            ["Riverside Equestrian - Analytics Report", "", ""],
            [f"Generated: {toronto_now().strftime('%Y-%m-%d %H:%M')}", "", ""],
            ["", "", ""],
            ["MONTHLY COMPARISON", "", ""],
            ["Metric", "Current Month", "Previous Month"],
            [
                "Total Cancellations",
                stats["current_month"]["total"],
                stats["last_month"]["total"],
            ],
            [
                "Free Cancellations",
                stats["current_month"]["free"],
                stats["last_month"]["free"],
            ],
            [
                "Charged Cancellations",
                stats["current_month"]["charged"],
                stats["last_month"]["charged"],
            ],
            [
                "Excluded Cancellations",
                stats["current_month"]["excluded"],
                stats["last_month"]["excluded"],
            ],
            ["", "", ""],
            ["MEMBERSHIP DISTRIBUTION", "", ""],
            ["Membership Level", "Student Count", "Percentage"],
        ]

        total_students = sum(m["count"] for m in membership_stats)
        for membership in membership_stats:
            percentage = f"{(membership['count'] / total_students * 100):.1f}%"
            summary_data.append(
                [membership["membership_level"], membership["count"], percentage]
            )

        # Write summary data
        for row_idx, row_data in enumerate(summary_data, 1):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws_summary.cell(row=row_idx, column=col_idx, value=value)
                if row_idx in [1, 4, 11]:  # Header rows
                    cell.font = Font(bold=True, size=12)
                elif row_idx == 5 or row_idx == 12:  # Sub-header rows
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(
                        start_color="E5E7EB", end_color="E5E7EB", fill_type="solid"
                    )

        # 2. Monthly Trends Sheet
        ws_trends = wb.create_sheet("Monthly Trends")

        # Get 12 months of data
        start_date = datetime.now().replace(day=1) - relativedelta(months=11)
        monthly_trends = conn.execute(
            """
            SELECT 
                strftime('%Y-%m', created_at) as month,
                COUNT(*) as total,
                SUM(CASE WHEN charged = 0 AND excluded = 0 THEN 1 ELSE 0 END) as free,
                SUM(CASE WHEN charged = 1 THEN 1 ELSE 0 END) as charged,
                SUM(CASE WHEN excluded = 1 THEN 1 ELSE 0 END) as excluded
            FROM cancellations 
            WHERE created_at >= ?
            GROUP BY strftime('%Y-%m', created_at)
            ORDER BY month
        """,
            (start_date.strftime("%Y-%m-%d"),),
        ).fetchall()

        # Write trends data
        trends_headers = ["Month", "Total", "Free", "Charged", "Excluded"]
        for col, header in enumerate(trends_headers, 1):
            cell = ws_trends.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(
                start_color="366092", end_color="366092", fill_type="solid"
            )
            cell.font = Font(bold=True, color="FFFFFF")

        for row, trend in enumerate(monthly_trends, 2):
            month_name = datetime.strptime(trend["month"], "%Y-%m").strftime("%B %Y")
            ws_trends.cell(row=row, column=1, value=month_name)
            ws_trends.cell(row=row, column=2, value=trend["total"])
            ws_trends.cell(row=row, column=3, value=trend["free"])
            ws_trends.cell(row=row, column=4, value=trend["charged"])
            ws_trends.cell(row=row, column=5, value=trend["excluded"])

        # 3. Student Analysis Sheet
        ws_students = wb.create_sheet("Student Analysis")

        # Get top students by cancellation frequency
        top_students = conn.execute(
            """
            SELECT 
                s.first_name,
                s.last_name,
                s.membership_level,
                COUNT(c.id) as total_cancellations,
                SUM(CASE WHEN c.charged = 1 THEN 1 ELSE 0 END) as charged_count,
                MAX(c.created_at) as last_cancellation
            FROM students s
            JOIN cancellations c ON s.id = c.student_id
            GROUP BY s.id
            HAVING COUNT(c.id) > 0
            ORDER BY total_cancellations DESC
            LIMIT 50
        """
        ).fetchall()

        # Write student analysis
        student_headers = [
            "Student Name",
            "Membership",
            "Total Cancellations",
            "Charged",
            "Last Cancellation",
        ]
        for col, header in enumerate(student_headers, 1):
            cell = ws_students.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(
                start_color="10B981", end_color="10B981", fill_type="solid"
            )
            cell.font = Font(bold=True, color="FFFFFF")

        for row, student in enumerate(top_students, 2):
            student_name = f"{student['first_name']} {student['last_name']}"
            ws_students.cell(row=row, column=1, value=student_name)
            ws_students.cell(row=row, column=2, value=student["membership_level"])
            ws_students.cell(row=row, column=3, value=student["total_cancellations"])
            ws_students.cell(row=row, column=4, value=student["charged_count"])

            last_date = ""
            if student["last_cancellation"]:
                last_date = datetime.fromisoformat(
                    student["last_cancellation"]
                ).strftime("%Y-%m-%d")
            ws_students.cell(row=row, column=5, value=last_date)

        # Auto-adjust column widths for all sheets
        for ws in wb.worksheets:
            for column in ws.columns:
                max_length = max(len(str(cell.value or "")) for cell in column)
                ws.column_dimensions[column[0].column_letter].width = min(
                    max_length + 2, 50
                )

        conn.close()

        # Save to temporary file
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
        wb.save(temp_file.name)
        temp_file.close()

        log_action("export_analytics", "Exported full analytics report")

        return send_file(
            temp_file.name,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=f'full_analytics_report_{datetime.now().strftime("%Y%m%d")}.xlsx',
        )

    except Exception as e:
        log_action("export_error", f"Error exporting analytics: {str(e)}")
        return jsonify({"success": False, "message": f"Export failed: {str(e)}"}), 500


# Add route to get cancellation details for management card
@app.route("/manager/api/cancellation/<int:cancellation_id>/card", methods=["GET"])
@login_required
@admin_required
def get_cancellation_card(cancellation_id):
    """Get cancellation data formatted for management card view"""
    try:
        conn = get_db()

        # Get cancellation with student and policy info
        cancellation_data = conn.execute(
            """
            SELECT 
                c.*,
                s.first_name,
                s.last_name,
                s.parent_first,
                s.parent_last,
                s.email,
                s.phone,
                s.membership_level,
                s.created_at as student_created_at
            FROM cancellations c
            JOIN students s ON c.student_id = s.id
            WHERE c.id = ?
        """,
            (cancellation_id,),
        ).fetchone()

        if not cancellation_data:
            return jsonify({"success": False, "message": "Cancellation not found"}), 404

        # Get student's current month cancellation count
        current_month = toronto_now().replace(day=1)
        month_cancellations = conn.execute(
            """
            SELECT COUNT(*) as count
            FROM cancellations
            WHERE student_id = ? AND created_at >= ?
        """,
            (cancellation_data["student_id"], current_month.strftime("%Y-%m-%d")),
        ).fetchone()

        conn.close()

        # Format response
        response_data = {
            "success": True,
            "cancellation": dict(cancellation_data),
            "month_count": month_cancellations["count"],
        }

        return jsonify(response_data)

    except Exception as e:
        return jsonify({"success": False, "message": str(e)}), 500


# ===================================
# IMPROVED ERROR HANDLERS
# ===================================


@app.errorhandler(404)
def not_found_error(error):
    """Handle 404 errors with safe template"""
    return (
        render_template_string(
            """
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Page Not Found - Riverside Equestrian</title>
    <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/css/bootstrap.min.css" rel="stylesheet">
</head>
<body class="bg-light">
    <div class="container mt-5">
        <div class="row justify-content-center">
            <div class="col-md-8 text-center">
                <div class="card shadow">
                    <div class="card-body p-5">
                        <h1 class="display-1 text-primary">404</h1>
                        <h2 class="mb-3">Page Not Found</h2>
                        <p class="lead mb-4">The page you're looking for doesn't exist.</p>
                        <p class="text-muted mb-4">
                            Error occurred at: {{ now.strftime('%Y-%m-%d %H:%M:%S') }}
                        </p>
                        <div class="d-grid gap-2 d-md-block">
                            <a href="{{ url_for('login') }}" class="btn btn-primary">Go to Login</a>
                            <a href="javascript:history.back()" class="btn btn-outline-secondary">Go Back</a>
                        </div>
                    </div>
                </div>
            </div>
        </div>
    </div>
</body>
</html>
    """,
            now=datetime.now(),
        ),
        404,
    )


@app.errorhandler(403)
def forbidden_error(error):
    """Handle 403 errors with safe template"""
    return (
        render_template_string(
            """
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Access Forbidden - Riverside Equestrian</title>
    <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/css/bootstrap.min.css" rel="stylesheet">
</head>
<body class="bg-light">
    <div class="container mt-5">
        <div class="row justify-content-center">
            <div class="col-md-8 text-center">
                <div class="card shadow">
                    <div class="card-body p-5">
                        <h1 class="display-1 text-warning">403</h1>
                        <h2 class="mb-3">Access Forbidden</h2>
                        <p class="lead mb-4">You don't have permission to access this resource.</p>
                        <p class="text-muted mb-4">
                            Error occurred at: {{ now.strftime('%Y-%m-%d %H:%M:%S') }}
                        </p>
                        <div class="d-grid gap-2 d-md-block">
                            <a href="{{ url_for('dashboard') }}" class="btn btn-primary">Go to Dashboard</a>
                            <a href="{{ url_for('logout') }}" class="btn btn-outline-secondary">Logout</a>
                        </div>
                    </div>
                </div>
            </div>
        </div>
    </div>
</body>
</html>
    """,
            now=datetime.now(),
        ),
        403,
    )


@app.errorhandler(500)
def internal_error(error):
    """Handle 500 errors with safe template"""
    return (
        render_template_string(
            """
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Server Error - Riverside Equestrian</title>
    <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/css/bootstrap.min.css" rel="stylesheet">
</head>
<body class="bg-light">
    <div class="container mt-5">
        <div class="row justify-content-center">
            <div class="col-md-8 text-center">
                <div class="card shadow">
                    <div class="card-body p-5">
                        <h1 class="display-1 text-danger">500</h1>
                        <h2 class="mb-3">Server Error</h2>
                        <p class="lead mb-4">An internal server error occurred.</p>
                        <p class="text-muted mb-4">
                            Error occurred at: {{ now.strftime('%Y-%m-%d %H:%M:%S') }}
                        </p>
                        <div class="d-grid gap-2 d-md-block">
                            <a href="{{ url_for('login') }}" class="btn btn-primary">Go to Login</a>
                            <a href="javascript:location.reload()" class="btn btn-outline-secondary">Reload Page</a>
                        </div>
                    </div>
                </div>
            </div>
        </div>
    </div>
</body>
</html>
    """,
            now=datetime.now(),
        ),
        500,
    )


# ===================================
# CREATE STATIC DIRECTORIES
# ===================================


def create_static_structure():
    """Create static file structure if it doesn't exist"""
    static_dirs = ["static", "static/css", "static/js", "static/images"]

    for directory in static_dirs:
        os.makedirs(directory, exist_ok=True)

    # Create basic main.css if it doesn't exist
    main_css_path = "static/css/main.css"
    if not os.path.exists(main_css_path):
        with open(main_css_path, "w") as f:
            f.write(
                """
/* Main CSS for Riverside Equestrian Cancellation System */
:root {
    --primary-color: #4e73df;
    --secondary-color: #858796;
    --success-color: #1cc88a;
    --warning-color: #f6c23e;
    --danger-color: #e74a3b;
    --info-color: #36b9cc;
}

body {
    font-family: 'Nunito', -apple-system, BlinkMacSystemFont, 'Segoe UI', 'Roboto', 'Helvetica Neue', Arial, sans-serif;
    background-color: #f8f9fc;
}

.navbar-brand {
    font-weight: 800;
    color: var(--primary-color) !important;
}

.card {
    border: none;
    border-radius: 0.35rem;
    box-shadow: 0 0.15rem 1.75rem 0 rgba(58, 59, 69, 0.15);
}

.card-header {
    background-color: #f8f9fc;
    border-bottom: 1px solid #e3e6f0;
}

.btn-primary {
    background-color: var(--primary-color);
    border-color: var(--primary-color);
}

.btn-primary:hover {
    background-color: #2e59d9;
    border-color: #2e59d9;
}

.alert {
    border: none;
    border-radius: 0.35rem;
}

.form-control:focus {
    border-color: var(--primary-color);
    box-shadow: 0 0 0 0.2rem rgba(78, 115, 223, 0.25);
}

.sidebar {
    background: linear-gradient(180deg, var(--primary-color) 10%, #224abe 100%);
    background-size: cover;
}

.sidebar .nav-link {
    color: rgba(255, 255, 255, 0.8);
}

.sidebar .nav-link:hover {
    color: #fff;
}

.sidebar .nav-link.active {
    color: #fff;
    background-color: rgba(255, 255, 255, 0.1);
    border-radius: 0.35rem;
}

@media (max-width: 768px) {
    .sidebar {
        position: relative !important;
        width: 100% !important;
        height: auto !important;
    }
}

.table th {
    border-top: none;
    font-weight: 800;
    font-size: 0.8rem;
    color: #5a5c69;
    text-transform: uppercase;
    letter-spacing: 0.5px;
}

.badge {
    font-size: 0.75rem;
    font-weight: 700;
    padding: 0.375rem 0.75rem;
}

.text-truncate {
    max-width: 200px;
}

/* Animation for loading states */
.loading {
    opacity: 0.6;
    pointer-events: none;
}

@keyframes fadeIn {
    from { opacity: 0; transform: translateY(10px); }
    to { opacity: 1; transform: translateY(0); }
}

.fade-in {
    animation: fadeIn 0.3s ease-out;
}

/* Print styles */
@media print {
    .sidebar, .navbar, .btn, .no-print {
        display: none !important;
    }
    
    .container-fluid {
        margin: 0 !important;
        padding: 0 !important;
    }
}
            """
            )

    # Create basic main.js if it doesn't exist
    main_js_path = "static/js/main.js"
    if not os.path.exists(main_js_path):
        with open(main_js_path, "w") as f:
            f.write(
                """
// Main JavaScript for Riverside Equestrian Cancellation System
console.log('Riverside Equestrian system loaded');

// Global utilities
window.RiversideUtils = {
    formatDate: function(date) {
        return new Date(date).toLocaleDateString();
    },
    
    formatTime: function(time) {
        return new Date('1970-01-01T' + time).toLocaleTimeString([], {hour: '2-digit', minute:'2-digit'});
    }
};
            """
            )


class EmailConfig:
    """Enhanced Email configuration class with Office 365 support"""

    def __init__(self):
        # Email server configuration
        self.smtp_server = os.getenv("SMTP_SERVER", "smtp.office365.com")
        self.smtp_port = int(os.getenv("SMTP_PORT", "587"))
        self.use_tls = os.getenv("USE_TLS", "True").lower() == "true"
        self.use_ssl = os.getenv("USE_SSL", "False").lower() == "true"
        self.timeout = int(os.getenv("EMAIL_TIMEOUT", "30"))

        # Authentication
        self.smtp_user = os.getenv("SMTP_USER", "")
        self.smtp_password = os.getenv("SMTP_PASSWORD", "")

        # Email addresses
        self.from_email = os.getenv("FROM_EMAIL", "noreply@riversideequestrian.ca")
        self.from_name = os.getenv("FROM_NAME", "Riverside Equestrian")
        self.admin_email = os.getenv("ADMIN_EMAIL", "stav@riversideequestrian.ca")
        self.manager_emails = [
            email.strip()
            for email in os.getenv(
                "MANAGER_EMAIL", "managers@riversideequestrian.ca"
            ).split(",")
        ]

        # Email behavior
        self.send_emails = os.getenv("SEND_EMAILS", "True").lower() == "true"
        self.debug_mode = os.getenv("EMAIL_DEBUG", "False").lower() == "true"
        self.log_emails = True

        # Validate configuration
        self.is_configured = bool(self.smtp_user and self.smtp_password)

    def get_connection(self):
        """Get SMTP connection"""
        try:
            if self.use_ssl:
                # Use SSL connection
                context = ssl.create_default_context()
                server = smtplib.SMTP_SSL(
                    self.smtp_server,
                    self.smtp_port,
                    context=context,
                    timeout=self.timeout,
                )
            else:
                # Use regular SMTP with STARTTLS (Office 365 default)
                server = smtplib.SMTP(
                    self.smtp_server, self.smtp_port, timeout=self.timeout
                )
                if self.use_tls:
                    server.starttls()

            if self.smtp_user and self.smtp_password:
                server.login(self.smtp_user, self.smtp_password)

            return server
        except Exception as e:
            raise Exception(f"Failed to connect to email server: {str(e)}")


email_config = EmailConfig()

# ===================================
# EMAIL NOTIFICATION SETTINGS HELPERS
# ===================================


def get_notification_settings():
    """
    Get notification settings from database
    Returns dict with settings or defaults if database unavailable
    """
    try:
        conn = get_db()
        settings = {}
        settings_rows = conn.execute(
            "SELECT key, value FROM system_settings WHERE key IN (?, ?, ?)",
            (
                "client_email_notifications",
                "manager_email_notifications",
                "manager_emails",
            ),
        ).fetchall()

        for setting in settings_rows:
            settings[setting["key"]] = setting["value"]

        conn.close()

        # Return settings with defaults
        return {
            "client_enabled": settings.get("client_email_notifications", "true")
            == "true",
            "manager_enabled": settings.get("manager_email_notifications", "true")
            == "true",
            "manager_emails": [
                email.strip()
                for email in settings.get(
                    "manager_emails", "managers@riversideequestrian.ca"
                ).split(",")
                if email.strip()
            ],
        }
    except Exception as e:
        # Fallback to defaults if database error
        print(f"Warning: Could not load notification settings from database: {e}")
        return {
            "client_enabled": True,
            "manager_enabled": True,
            "manager_emails": ["managers@riversideequestrian.ca"],
        }


def should_send_email(recipient_type):
    """
    Check if emails should be sent for the given recipient type

    Args:
        recipient_type (str): 'client' or 'manager'

    Returns:
        bool: True if emails should be sent, False otherwise
    """
    settings = get_notification_settings()

    if recipient_type == "client":
        return settings["client_enabled"]
    elif recipient_type == "manager":
        return settings["manager_enabled"]
    else:
        return True  # Default to sending for unknown types


def get_manager_emails_from_settings():
    """
    Get manager emails from database settings

    Returns:
        list: List of manager email addresses
    """
    settings = get_notification_settings()
    return settings["manager_emails"]


# ===================================
# EMAIL SENDING FUNCTIONS
# ===================================


def send_email_async(func):
    """Decorator to send emails asynchronously"""

    @wraps(func)
    def wrapper(*args, **kwargs):
        thread = threading.Thread(target=func, args=args, kwargs=kwargs)
        thread.daemon = True
        thread.start()
        return thread

    return wrapper


def log_email_attempt(to_email, subject, success, error=None, template_id=None):
    """Enhanced email logging"""
    try:
        conn = get_db()
        details = f"To: {to_email}, Subject: {subject}"
        if template_id:
            details += f", Template: {template_id}"
        if error:
            details += f", Error: {error}"

        conn.execute(
            """
            INSERT INTO system_logs (user_id, user_type, action, details, created_at)
            VALUES (?, ?, ?, ?, ?)
        """,
            (
                session.get("user_id"),
                "email_system",
                "email_sent" if success else "email_failed",
                details,
                toronto_now().strftime("%Y-%m-%d %H:%M:%S"),
            ),
        )
        conn.commit()
        conn.close()
    except Exception as e:
        if email_config.debug_mode:
            print(f"Failed to log email attempt: {e}")


def wrap_email_html(body_content):
    """
    Wrap email body in a minimal HTML email shell with inline styles only.

    WHY INLINE STYLES (not <style> blocks):
        Gmail (especially mobile and forwarded mail) strips <style> blocks
        in <head>. Inline style="..." attributes always survive.

    WHAT THIS DOES:
        - Provides a minimal HTML document shell + centered white card.
        - Does NOT inject styles into <p>, <div>, <a>, etc. — your template's
          own inline styles take full effect.
        - DOES force font-weight: bold inline onto bare <strong>/<b> tags,
          because Gmail intermittently fails to honor the default browser
          bold weight when the parent element has custom font-family.
        - Same for <em>/<i> → italic.
        - Skips tags that already have a style attribute, so any
          template-customized weight/style is preserved.

    Args:
        body_content (str): Raw HTML body content from the template,
                            with template variables already substituted.

    Returns:
        str: Complete HTML email document.
    """
    import re as _re

    # Force bold on bare <strong> and <b> (no existing attributes).
    def _style_bold_tag(m):
        tag = m.group(1)
        attrs = m.group(2)
        if attrs and 'style=' in attrs.lower():
            return m.group(0)
        return f'<{tag} style="font-weight: bold; font-family: Arial, Helvetica, sans-serif;">'

    styled_body = _re.sub(
        r"<(strong|b)(\s[^>]*)?>",
        _style_bold_tag,
        body_content,
        flags=_re.IGNORECASE,
    )
    # Force italic on bare <em> and <i>
    styled_body = _re.sub(
        r"<(em|i)>",
        lambda m: f'<{m.group(1)} style="font-style: italic;">',
        styled_body,
        flags=_re.IGNORECASE,
    )
    # Inline-style bare <p> tags (no existing style= attribute) for cross-client consistency.
    # Gmail/Hotmail reset paragraph margins without explicit inline styles.
    styled_body = _re.sub(
        r"<p(?!\s[^>]*style=)(?=[>\s])",
        '<p style="margin: 0 0 14px 0; font-size: 14px; line-height: 1.6; color: #333333; font-family: Arial, Helvetica, sans-serif;">',
        styled_body,
        flags=_re.IGNORECASE,
    )
    # Replace bare <hr> with an explicitly styled one (Outlook renders unstyled <hr> as invisible)
    styled_body = _re.sub(
        r"<hr\s*/?>",
        '<hr style="border: 0; border-top: 2px solid #dddddd; margin: 20px 0;" />',
        styled_body,
        flags=_re.IGNORECASE,
    )
    # Inline-style bare <a> tags so link colour survives Gmail's reset.
    def _style_a_tag(m):
        attrs = m.group(1)
        if 'style=' in attrs.lower():
            return m.group(0)
        return f'<a {attrs} style="color: #2B7BC2; text-decoration: underline;">'

    styled_body = _re.sub(
        r"<a\s([^>]*)>",
        _style_a_tag,
        styled_body,
        flags=_re.IGNORECASE,
    )
    # Headings
    for tag, size, weight in [("h1", "22px", "bold"), ("h2", "18px", "bold"), ("h3", "16px", "bold")]:
        styled_body = _re.sub(
            rf"<{tag}>",
            f'<{tag} style="font-size: {size}; font-weight: {weight}; color: #222222; margin: 0 0 12px 0; font-family: Arial, Helvetica, sans-serif;">',
            styled_body,
            flags=_re.IGNORECASE,
        )
    # Lists
    styled_body = _re.sub(
        r"<ul>",
        '<ul style="margin: 0 0 14px 0; padding-left: 20px; font-size: 14px; line-height: 1.6; color: #333333; font-family: Arial, Helvetica, sans-serif;">',
        styled_body,
        flags=_re.IGNORECASE,
    )
    styled_body = _re.sub(
        r"<ol>",
        '<ol style="margin: 0 0 14px 0; padding-left: 20px; font-size: 14px; line-height: 1.6; color: #333333; font-family: Arial, Helvetica, sans-serif;">',
        styled_body,
        flags=_re.IGNORECASE,
    )
    styled_body = _re.sub(
        r"<li>",
        '<li style="margin-bottom: 6px; font-size: 14px; color: #333333;">',
        styled_body,
        flags=_re.IGNORECASE,
    )

    html = f"""<!DOCTYPE html PUBLIC "-//W3C//DTD XHTML 1.0 Transitional//EN" "http://www.w3.org/TR/xhtml1/DTD/xhtml1-transitional.dtd">
<html xmlns="http://www.w3.org/1999/xhtml">
<head>
    <meta http-equiv="Content-Type" content="text/html; charset=UTF-8" />
    <meta name="viewport" content="width=device-width, initial-scale=1.0" />
    <meta http-equiv="X-UA-Compatible" content="IE=edge" />
    <title>Riverside Equestrian</title>
</head>
<body style="margin: 0; padding: 20px; background-color: #f4f5f7; font-family: Arial, Helvetica, sans-serif; font-size: 14px; color: #333333;">
    <table role="presentation" cellpadding="0" cellspacing="0" border="0" width="100%" style="background-color: #f4f5f7;">
        <tr>
            <td align="center" style="padding: 20px 10px;">
                <table role="presentation" cellpadding="0" cellspacing="0" border="0" width="600" style="max-width: 600px; background-color: #ffffff; border: 1px solid #e0e0e0;">
                    <tr>
                        <td style="padding: 32px; font-family: Arial, Helvetica, sans-serif; font-size: 14px; color: #333333; line-height: 1.6;">
                            {styled_body}
                        </td>
                    </tr>
                </table>
            </td>
        </tr>
    </table>
</body>
</html>"""
    return html


def send_email(
    to_email, subject, body, template_type="client", attachments=None, template_id=None
):
    """
    Enhanced email sending with Office 365 support and notification settings check

    Args:
        to_email (str): Recipient email address
        subject (str): Email subject
        body (str): HTML email body
        template_type (str): Type of email template ('client' or 'manager')
        attachments (list): List of file paths to attach
        template_id (str): Template ID for logging

    Returns:
        dict: Result with success status and message
    """

    # NEW: Check if notifications are enabled for this recipient type
    if not should_send_email(template_type):
        message = f"Email notifications disabled for {template_type}s"
        if email_config.debug_mode:
            print(f"🔕 {message} - would send to {to_email}: {subject}")
        log_email_attempt(to_email, subject, True, message, template_id)
        return {"success": True, "message": message, "skipped": True}

    if not email_config.send_emails:
        if email_config.debug_mode:
            print(f"Email sending disabled - would send to {to_email}: {subject}")
        return {"success": True, "message": "Email sending disabled (debug mode)"}

    if not email_config.is_configured:
        error_msg = "Email not configured - missing SMTP credentials"
        log_email_attempt(to_email, subject, False, error_msg, template_id)
        return {"success": False, "message": error_msg}

    try:
        # Create message
        msg = MIMEMultipart("alternative")
        msg["From"] = f"{email_config.from_name} <{email_config.from_email}>"
        msg["To"] = to_email
        msg["Subject"] = subject
        msg["Reply-To"] = email_config.from_email

        # Wrap body in proper HTML structure for Gmail compatibility
        wrapped_body = wrap_email_html(body)

        # Build a plain-text fallback by stripping tags from the raw body
        from html import unescape
        import re

        plain_text = re.sub("<[^<]+?>", "", body)
        plain_text = unescape(plain_text)

        # IMPORTANT (RFC 2046): For multipart/alternative, parts must be
        # ordered from least-preferred (plain text) to MOST-preferred (HTML).
        # Email clients display the LAST part they can render. Attaching
        # plain text last would make Gmail show the unformatted version.
        text_part = MIMEText(plain_text, "plain", "utf-8")
        msg.attach(text_part)

        html_part = MIMEText(wrapped_body, "html", "utf-8")
        msg.attach(html_part)

        # Add attachments if provided
        if attachments:
            for file_path in attachments:
                try:
                    with open(file_path, "rb") as attachment:
                        part = MIMEBase("application", "octet-stream")
                        part.set_payload(attachment.read())
                        encoders.encode_base64(part)
                        part.add_header(
                            "Content-Disposition",
                            f"attachment; filename= {os.path.basename(file_path)}",
                        )
                        msg.attach(part)
                except Exception as e:
                    if email_config.debug_mode:
                        print(f"Failed to attach file {file_path}: {e}")

        # Send email using connection
        server = email_config.get_connection()
        server.send_message(msg)
        server.quit()

        result = {"success": True, "message": "Email sent successfully"}

        if email_config.debug_mode:
            print(f"✓ Email sent to {to_email}: {subject}")

        log_email_attempt(to_email, subject, True, template_id=template_id)
        return result

    except Exception as e:
        error_msg = str(e)
        result = {"success": False, "message": f"Failed to send email: {error_msg}"}

        if email_config.debug_mode:
            print(f"✗ Email failed to {to_email}: {error_msg}")

        log_email_attempt(to_email, subject, False, error_msg, template_id)
        return result


def send_email_by_trigger(trigger_id, student, cancellation, recipient_type=None, extra_vars=None):
    """
    New function: Send emails based on trigger mappings
    This replaces hardcoded email sending with trigger-based system
    
    Args:
        trigger_id (str): The trigger ID (e.g., 'free_cancellation_triggered')
        student (dict): Student data
        cancellation (dict): Cancellation data
        recipient_type (str): Optional filter for recipient type ('student', 'manager', 'client')
        extra_vars (dict): Extra variables to include in template
    
    Returns:
        dict: Results with success status and details
    """
    
    # Convert sqlite3.Row to dict if needed
    if hasattr(student, "keys"):
        student_dict = dict(student)
    else:
        student_dict = student
    
    if hasattr(cancellation, "keys"):
        cancellation_dict = dict(cancellation)
    else:
        cancellation_dict = cancellation
    
    # Get templates for this trigger
    templates = get_templates_by_trigger(trigger_id, recipient_type)
    
    if not templates:
        if email_config.debug_mode:
            print(f"No templates found for trigger: {trigger_id}")
        return {"success": False, "message": f"No templates configured for trigger: {trigger_id}"}
    
    results = []
    
    # Generate variables once
    variables = get_template_variables(student_dict, cancellation_dict, extra_vars)
    
    # Send email for each mapped template
    for template in templates:
        try:
            recipient_type_email = template.get('recipient_type', 'client')
            
            # Determine recipient email
            if recipient_type_email == 'manager':
                manager_emails = get_manager_emails_from_settings()
                recipient_emails = manager_emails
            elif recipient_type_email == 'student':
                recipient_emails = [student_dict.get('email')]
            else:  # client
                recipient_emails = [student_dict.get('email')]
            
            # Send to each recipient
            for to_email in recipient_emails:
                if not to_email:
                    continue
                
                # Process template
                body, subject = process_template_variables(
                    template["body"], template["subject"], variables
                )
                
                # Send email
                result = send_email(
                    to_email,
                    subject,
                    body,
                    recipient_type_email,
                    template_id=template["id"]
                )
                
                results.append({
                    "template_id": template["id"],
                    "recipient": to_email,
                    "recipient_type": recipient_type_email,
                    "result": result
                })
        
        except Exception as e:
            if email_config.debug_mode:
                print(f"Error sending template {template.get('id')}: {str(e)}")
            results.append({
                "template_id": template.get("id"),
                "error": str(e)
            })
    
    # Return aggregate results
    success_count = sum(1 for r in results if r.get('result', {}).get('success'))
    return {
        "success": success_count > 0,
        "message": f"Sent {success_count}/{len(results)} emails for trigger {trigger_id}",
        "details": results
    }


def send_cancellation_emails_by_trigger(student, cancellation):
    """
    Enhanced version: Send all appropriate cancellation emails using trigger system
    Replaces the old hardcoded logic
    
    Args:
        student (dict): Student data
        cancellation (dict): Cancellation data
    
    Returns:
        dict: Results with success status
    """
    
    # Convert to dicts
    if hasattr(student, "keys"):
        student_dict = dict(student)
    else:
        student_dict = student
    
    if hasattr(cancellation, "keys"):
        cancellation_dict = dict(cancellation)
    else:
        cancellation_dict = cancellation
    
    # Determine which trigger to use based on charge status
    if cancellation_dict.get('charged'):
        trigger_id = 'charged_cancellation_triggered'
    else:
        trigger_id = 'free_cancellation_triggered'
    
    if email_config.debug_mode:
        print(f"📧 Using trigger: {trigger_id} for student {student_dict.get('email')}")
    
    # Send using trigger-based system
    return send_email_by_trigger(trigger_id, student_dict, cancellation_dict)


def send_manager_notification_by_trigger(student, cancellation):
    """
    Enhanced version: Send manager notification using trigger system
    
    Args:
        student (dict): Student data
        cancellation (dict): Cancellation data
    
    Returns:
        dict: Results with success status
    """
    
    # Convert to dicts
    if hasattr(student, "keys"):
        student_dict = dict(student)
    else:
        student_dict = student
    
    if hasattr(cancellation, "keys"):
        cancellation_dict = dict(cancellation)
    else:
        cancellation_dict = cancellation
    
    extra_vars = {
        "action_required": "Review cancellation and approve/charge as needed",
        "dashboard_url": f"{request.url_root if 'request' in globals() else 'http://localhost:5000/'}manager/cancellations?student={student_dict['id']}",
    }
    
    if email_config.debug_mode:
        print(f"📧 Sending manager notification for student {student_dict.get('email')}")
    
    # Send using trigger-based system
    return send_email_by_trigger('manager_new_cancellation', student_dict, cancellation_dict, 'manager', extra_vars)





# ===================================
# TEMPLATE PROCESSING FUNCTIONS
# ===================================


def get_email_template(template_id):
    """Get email template from database with error handling"""
    try:
        conn = get_db()
        template = conn.execute(
            "SELECT * FROM email_templates WHERE id = ? AND active = 1", (template_id,)
        ).fetchone()
        conn.close()
        return dict(template) if template else None
    except Exception as e:
        if email_config.debug_mode:
            print(f"Error getting template {template_id}: {e}")
        return None


def get_templates_by_trigger(trigger_id, recipient_type=None):
    """Get all templates mapped to a specific trigger, optionally filtered by recipient type"""
    try:
        conn = get_db()
        conn.row_factory = sqlite3.Row
        
        query = """
            SELECT et.id, et.name, et.subject, et.body, et.type, et.active,
                   tm.recipient_type, tm.enabled, tm.priority
            FROM email_trigger_mappings tm
            JOIN email_templates et ON tm.template_id = et.id
            WHERE tm.trigger_id = ? AND tm.enabled = 1 AND et.active = 1
        """
        params = [trigger_id]
        
        if recipient_type:
            query += " AND tm.recipient_type = ?"
            params.append(recipient_type)
        
        query += " ORDER BY tm.priority, et.name"
        
        templates = conn.execute(query, params).fetchall()
        conn.close()
        
        return [dict(t) for t in templates]
    except Exception as e:
        if email_config.debug_mode:
            print(f"Error getting templates for trigger {trigger_id}: {e}")
        return []


def process_template_variables(template_body, template_subject, variables):
    """
    Enhanced template variable processing with better error handling

    Args:
        template_body (str): Template HTML body
        template_subject (str): Template subject
        variables (dict): Variables to replace

    Returns:
        tuple: (processed_body, processed_subject)
    """
    processed_body = template_body
    processed_subject = template_subject

    for key, value in variables.items():
        placeholder = f"{{{{{key}}}}}"
        processed_body = processed_body.replace(placeholder, str(value))
        processed_subject = processed_subject.replace(placeholder, str(value))

    return processed_body, processed_subject


# ===================================
# TEMPLATE VARIABLE GENERATORS
# ===================================


def safe_value(val, default="NONE"):
    """Convert None or empty string to NONE, otherwise return value"""
    if val is None or val == "" or val == "None":
        return default
    if isinstance(val, str):
        stripped = val.strip()
        return default if stripped == "" else stripped
    return str(val)


def get_template_variables(student=None, cancellation=None, extra_vars=None):
    """Generate template variables for your cancellation system - FIXED for sqlite3.Row"""
    # Use Pacific timezone for all timestamps
    pacific_now = toronto_now()

    variables = {
        "current_date": pacific_now.strftime("%B %d, %Y"),
        "current_time": pacific_now.strftime("%I:%M %p"),
        "company_name": "Riverside Equestrian",
        "contact_email": email_config.admin_email,
        "policy_url": "https://www.riversideequestrian.ca/cancellations",
        "website_url": "https://www.riversideequestrian.ca",
        "system_email": email_config.from_email,
    }

    if student:
        # Convert sqlite3.Row to dict if needed
        if hasattr(student, "keys"):
            student_dict = dict(student)
        else:
            student_dict = student

        variables.update(
            {
                "client_name": safe_value(f"{student_dict.get('first_name', '')} {student_dict.get('last_name', '')}".strip()),
                "client_first_name": safe_value(student_dict.get("first_name")),
                "client_last_name": safe_value(student_dict.get("last_name")),
                "client_email": safe_value(student_dict.get("email")),
                "client_phone": safe_value(student_dict.get("phone")),
                "membership_tier": safe_value(student_dict.get("membership_level")),
                "parent_name": safe_value(f"{student_dict.get('parent_first', '')} {student_dict.get('parent_last', '')}".strip()),
                "parent_first_name": safe_value(student_dict.get("parent_first")),
                "parent_last_name": safe_value(student_dict.get("parent_last")),
            }
        )

        # Get membership tier info
        tier = get_membership_tier(student_dict.get("membership_level"))
        if tier:
            tier_dict = dict(tier) if hasattr(tier, "keys") else tier
            variables.update(
                {
                    "allowed_cancellations": safe_value(tier_dict.get("free_notices")),
                    "cancellation_deadline": safe_value(tier_dict.get("deadline_display")),
                }
            )

        # Get current usage
        status = calculate_cancellation_status(student_dict)
        variables.update(
            {
                "used_cancellations": safe_value(status.get("used")),
                "remaining_cancellations": safe_value(status.get("remaining")),
            }
        )

    if cancellation:
        # Convert sqlite3.Row to dict if needed
        if hasattr(cancellation, "keys"):
            cancellation_dict = dict(cancellation)
        else:
            cancellation_dict = cancellation

        # Format dates consistently
        lesson_date_str = cancellation_dict.get("lesson_date")
        if isinstance(lesson_date_str, str):
            try:
                lesson_date = datetime.strptime(lesson_date_str, "%Y-%m-%d")
            except ValueError:
                lesson_date = toronto_now()
        else:
            lesson_date = lesson_date_str or toronto_now()

        lesson_time_str = cancellation_dict.get("lesson_time")
        if isinstance(lesson_time_str, str):
            try:
                if len(lesson_time_str.split(":")) == 3:
                    lesson_time = datetime.strptime(lesson_time_str, "%H:%M:%S").time()
                else:
                    lesson_time = datetime.strptime(lesson_time_str, "%H:%M").time()
            except ValueError:
                lesson_time = toronto_now().time()
        else:
            lesson_time = lesson_time_str or toronto_now().time()

        # Get submission time - use manual_submission_date for manager submissions
        # This ensures emails show the correct "policy" submission date
        is_manager_submission = cancellation_dict.get("is_manager_submission", False)
        manual_submission_date = cancellation_dict.get("manual_submission_date")

        if is_manager_submission and manual_submission_date:
            # Manager submission: use the policy submission date they specified
            try:
                manual_date_str = manual_submission_date.strip()
                # Handle both "YYYY-MM-DD HH:MM" and "YYYY-MM-DD HH:MM:SS" formats
                if manual_date_str.count(':') == 1:  # HH:MM format
                    submission_dt = datetime.strptime(manual_date_str + ":00", "%Y-%m-%d %H:%M:%S")
                else:  # HH:MM:SS format
                    submission_dt = datetime.strptime(manual_date_str, "%Y-%m-%d %H:%M:%S")
                pacific_tz = pytz.timezone("America/Los_Angeles")
                if submission_dt.tzinfo is None:
                    submission_dt = pacific_tz.localize(submission_dt)
            except Exception as e:
                # Fallback to created_at if parsing fails
                created_at = cancellation_dict.get("created_at")
                submission_dt = localize_datetime(created_at)
                if submission_dt is None:
                    submission_dt = toronto_now()
        else:
            # Student submission: use created_at (actual submission timestamp)
            created_at = cancellation_dict.get("created_at")
            submission_dt = localize_datetime(created_at)
            if submission_dt is None:
                submission_dt = toronto_now()

        submission_time_str = submission_dt.strftime("%B %d, %Y at %I:%M %p")
        submission_month = submission_dt.strftime("%B %Y")

        variables.update(
            {
                "lesson_date": lesson_date.strftime("%B %d, %Y"),
                "lesson_time": lesson_time.strftime("%I:%M %p"),
                "cancellation_status": (
                    "Free cancellation"
                    if not cancellation_dict.get("charged")
                    else "Charged cancellation"
                ),
                "will_be_charged": "Yes" if cancellation_dict.get("charged") else "No",
                "charge_reason": safe_value(get_charge_reason(cancellation_dict)),
                "submission_time": submission_time_str,
                "submission_month": submission_month,
                "cancellation_id": safe_value(cancellation_dict.get("id")),
            }
        )

        # Sequential lessons
        sequential_lessons = cancellation_dict.get("sequential_lessons")
        lesson_date_for_seq = cancellation_dict.get("lesson_date")
        if sequential_lessons:
            variables["sequential_lessons"] = format_sequential_lessons(
                sequential_lessons, lesson_date_for_seq
            )
        else:
            variables["sequential_lessons"] = "None"

        # Reschedule info
        variables.update(
            {
                "reschedule_requested": (
                    "Yes" if cancellation_dict.get("reschedule_requested") else "No"
                ),
                "reschedule_preferences": safe_value(cancellation_dict.get("reschedule_preferences"), "None provided"),
                "error_report": safe_value(cancellation_dict.get("error_report"), "None reported"),
                "cancellation_note": safe_value(cancellation_dict.get("cancellation_note")),
                "manager_notes": safe_value(cancellation_dict.get("manager_notes")),
            }
        )

        # Calculate lesson month usage (for the month when the lesson is scheduled)
        if cancellation_dict.get("student_id") and lesson_date:
            lesson_month = lesson_date.month
            lesson_year = lesson_date.year
            lesson_month_usage = get_monthly_cancellation_count(cancellation_dict.get("student_id"), lesson_month, lesson_year)
            
            # Get the tier to know how many are allowed
            if student:
                tier = get_membership_tier(student_dict.get("membership_level"))
                if tier:
                    tier_dict = dict(tier) if hasattr(tier, "keys") else tier
                    allowed = tier_dict.get("free_notices", 1)
                    lesson_month_str = lesson_date.strftime("%B %Y")
                    
                    variables.update(
                        {
                            "lesson_month": lesson_month_str,
                            "lesson_month_used_cancellations": safe_value(str(lesson_month_usage)),
                            "lesson_month_allowed_cancellations": safe_value(str(allowed)),
                            "lesson_month_remaining_cancellations": safe_value(str(max(0, allowed - lesson_month_usage))),
                        }
                    )

    # Add extra variables and sanitize them
    if extra_vars:
        sanitized_extra = {}
        for key, value in extra_vars.items():
            sanitized_extra[key] = safe_value(value) if not isinstance(value, (int, float, bool)) else value
        variables.update(sanitized_extra)

    return variables


def get_charge_reason(cancellation):
    """Get human-readable reason for cancellation charge - FIXED for sqlite3.Row"""
    # Convert sqlite3.Row to dict if needed
    if hasattr(cancellation, "keys"):
        cancellation_dict = dict(cancellation)
    else:
        cancellation_dict = cancellation

    if not cancellation_dict.get("charged"):
        return "Within policy - no charge applied"

    if cancellation_dict.get("manager_notes"):
        return cancellation_dict["manager_notes"]

    # Default reasons based on common scenarios
    return "Late cancellation or monthly limit exceeded"


def format_sequential_lessons(sequential_data, cancelled_lesson_date=None):
    """Format sequential lessons for email and view history"""
    if not sequential_data:
        return "None"

    try:
        if isinstance(sequential_data, str):
            sequential_lessons = eval(sequential_data)
        else:
            sequential_lessons = sequential_data

        if not sequential_lessons:
            return "None"

        # Get the last lesson date and add one day to get the return date
        last_lesson = sequential_lessons[-1]
        if isinstance(last_lesson, dict):
            date_str = last_lesson.get("date", "")
        else:
            return "None"

        if date_str:
            try:
                # Parse the last cancelled lesson date
                last_cancelled_date = datetime.strptime(str(date_str), "%Y-%m-%d")
                # Add one day to get the return date
                return_date_obj = last_cancelled_date + timedelta(days=1)
                return_date = return_date_obj.strftime("%B %d, %Y")

                # If cancelled_lesson_date is provided, show the full message
                if cancelled_lesson_date:
                    if isinstance(cancelled_lesson_date, str):
                        cancelled_date = datetime.strptime(
                            cancelled_lesson_date, "%Y-%m-%d"
                        ).strftime("%B %d, %Y")
                    else:
                        cancelled_date = cancelled_lesson_date.strftime("%B %d, %Y")

                    return f"All lessons are cancelled from {cancelled_date} until the return date of {return_date}"
                else:
                    # Fallback if cancelled date is not provided
                    return f"All lessons are cancelled until the return date of {return_date}"
            except:
                return "None"

        return "None"

    except Exception as e:
        if email_config.debug_mode:
            print(f"Error formatting sequential lessons: {e}")
        return "No additional lessons"


def safe_dict_convert(row_or_dict):
    """Safely convert sqlite3.Row to dict, or return dict as-is"""
    if row_or_dict is None:
        return {}
    if hasattr(row_or_dict, "keys"):
        return dict(row_or_dict)
    return row_or_dict


# ===================================
# EMAIL SENDING WRAPPER FUNCTIONS (FOR YOUR CANCELLATION SYSTEM)
# ===================================


def debug_email_templates():
    """Debug function to check email template status"""
    print("=" * 60)
    print("DEBUGGING EMAIL TEMPLATES")
    print("=" * 60)

    # Check what templates exist in database
    conn = get_db()
    templates = conn.execute(
        "SELECT id, name, active FROM email_templates ORDER BY id"
    ).fetchall()
    conn.close()

    print(f"Templates in database ({len(templates)} found):")
    for template in templates:
        print(
            f"  - {template['id']}: {template['name']} (Active: {template['active']})"
        )

    print()

    # Test template retrieval
    required_templates = [
        "client_confirmation",
        "cancellation_charged",
        "free_cancellation",
        "manager_notification",
    ]

    for template_id in required_templates:
        template = get_email_template(template_id)
        if template:
            print(f"✅ {template_id}: Found and active")
        else:
            print(f"❌ {template_id}: NOT FOUND or INACTIVE")

    print("\n" + "=" * 60)

    # Test cancellation logic
    print("TESTING CANCELLATION LOGIC")
    print("=" * 60)

    # Mock Bronze student data (like you)
    mock_student = {
        "id": 1,
        "first_name": "Test",
        "last_name": "Student",
        "email": "test@example.com",
        "membership_level": "Bronze",
    }

    # Mock 4th cancellation (should be charged)
    mock_charged_cancellation = {
        "id": 999,
        "lesson_date": "2024-12-15",
        "lesson_time": "14:00:00",
        "charged": True,  # This should trigger charged email
        "manager_notes": "Monthly limit exceeded",
        "cancellation_note": "Family emergency",
    }

    # Mock 1st cancellation (should be free)
    mock_free_cancellation = {
        "id": 998,
        "lesson_date": "2024-12-20",
        "lesson_time": "15:00:00",
        "charged": False,  # This should trigger free email
        "manager_notes": "Within policy",
        "cancellation_note": None,
    }

    print("Test 1: Charged cancellation (Bronze member, 4th cancellation)")
    print(
        f"  - Student: {mock_student['first_name']} {mock_student['last_name']} ({mock_student['membership_level']})"
    )
    print(f"  - Charged status: {mock_charged_cancellation['charged']}")
    print(f"  - Expected template: cancellation_charged")

    # Test the logic
    if mock_charged_cancellation.get("charged"):
        expected_template = "cancellation_charged"
        template = get_email_template("cancellation_charged")
        if template:
            print(f"  ✅ Would use: {expected_template}")
        else:
            print(
                f"  ⚠️  Template '{expected_template}' not found, would fall back to client_confirmation"
            )
    else:
        print(f"  ❌ Logic error: charged=True but treated as free")

    print()
    print("Test 2: Free cancellation (Bronze member, 1st cancellation)")
    print(
        f"  - Student: {mock_student['first_name']} {mock_student['last_name']} ({mock_student['membership_level']})"
    )
    print(f"  - Charged status: {mock_free_cancellation['charged']}")
    print(f"  - Expected template: free_cancellation")

    if not mock_free_cancellation.get("charged"):
        expected_template = "free_cancellation"
        template = get_email_template("free_cancellation")
        if template:
            print(f"  ✅ Would use: {expected_template}")
        else:
            print(
                f"  ⚠️  Template '{expected_template}' not found, would fall back to client_confirmation"
            )
    else:
        print(f"  ❌ Logic error: charged=False but treated as charged")

    print("\n" + "=" * 60)
    return templates


# Add this route to your app.py to run the debug:
@app.route("/debug-email-templates")
@login_required
@senior_admin_required
def debug_email_route():
    """Debug route to check email template status"""
    templates = debug_email_templates()

    output = "<h1>Email Template Debug</h1><pre>"

    # Capture the debug output
    import io
    import sys
    from contextlib import redirect_stdout

    f = io.StringIO()
    with redirect_stdout(f):
        debug_email_templates()
    output += f.getvalue()
    output += "</pre>"

    return output


# MANUAL TEST FUNCTION - Add this too:
def test_email_sending_logic():
    """Test the actual email sending with your current logic"""
    print("\n" + "=" * 60)
    print("TESTING ACTUAL EMAIL SENDING LOGIC")
    print("=" * 60)

    # Your actual student data (Bronze member)
    conn = get_db()
    bronze_student = conn.execute(
        "SELECT * FROM students WHERE membership_level = 'Bronze' LIMIT 1"
    ).fetchone()

    if not bronze_student:
        print("❌ No Bronze student found in database")
        conn.close()
        return

    print(
        f"Testing with: {bronze_student['first_name']} {bronze_student['last_name']} ({bronze_student['membership_level']})"
    )

    # Test charged cancellation
    charged_cancellation = {
        "id": 9999,
        "lesson_date": "2024-12-15",
        "lesson_time": "14:00:00",
        "charged": True,  # Should use cancellation_charged template
        "manager_notes": "Test charged cancellation",
    }

    print("\n1. Testing CHARGED cancellation:")
    print(f"   charged = {charged_cancellation['charged']}")

    # Run your actual function
    try:
        result = send_cancellation_confirmation(bronze_student, charged_cancellation)
        print(f"   Result: {result}")
    except Exception as e:
        print(f"   ❌ Error: {e}")

    # Test free cancellation
    free_cancellation = {
        "id": 9998,
        "lesson_date": "2024-12-20",
        "lesson_time": "15:00:00",
        "charged": False,  # Should use free_cancellation template
        "manager_notes": "Test free cancellation",
    }

    print("\n2. Testing FREE cancellation:")
    print(f"   charged = {free_cancellation['charged']}")

    try:
        result = send_cancellation_confirmation(bronze_student, free_cancellation)
        print(f"   Result: {result}")
    except Exception as e:
        print(f"   ❌ Error: {e}")

    conn.close()


# FIXED EMAIL SENDING FUNCTIONS FOR RIVERSIDE EQUESTRIAN
# Replace your existing functions in app.py with these:


def send_cancellation_confirmation(student, cancellation):
    """Send appropriate cancellation email to client based on charge status - FIXED"""

    # Convert sqlite3.Row to dict if needed
    if hasattr(student, "keys"):
        student_dict = dict(student)
    else:
        student_dict = student

    if hasattr(cancellation, "keys"):
        cancellation_dict = dict(cancellation)
    else:
        cancellation_dict = cancellation

    # CHOOSE THE RIGHT TEMPLATE BASED ON CHARGE STATUS
    if cancellation_dict.get("charged"):
        # Use charged template for paid cancellations
        template_id = "cancellation_charged"
        template = get_email_template("cancellation_charged")
        if not template:
            if email_config.debug_mode:
                print(
                    "Warning: No cancellation_charged template found, falling back to client_confirmation"
                )
            template_id = "client_confirmation"
            template = get_email_template("client_confirmation")
    else:
        # Use free cancellation template for free cancellations
        template_id = "free_cancellation"
        template = get_email_template("free_cancellation")
        if not template:
            if email_config.debug_mode:
                print(
                    "Warning: No free_cancellation template found, falling back to client_confirmation"
                )
            template_id = "client_confirmation"
            template = get_email_template("client_confirmation")

    if not template:
        if email_config.debug_mode:
            print("Error: No email template found at all!")
        return {"success": False, "message": "Template not found"}

    # Generate variables
    variables = get_template_variables(student_dict, cancellation_dict)

    # Add status message based on charge status
    if cancellation_dict.get("charged"):
        variables["status_message"] = (
            f"A charge will be applied to your account. Reason: {get_charge_reason(cancellation_dict)}"
        )
        variables["charge_reason"] = get_charge_reason(cancellation_dict)
    else:
        variables["status_message"] = (
            "This cancellation has been processed at no charge."
        )

    # Process template
    body, subject = process_template_variables(
        template["body"], template["subject"], variables
    )

    # Send email with correct template ID
    result = send_email(
        student_dict["email"],
        subject,
        body,
        "client",
        template_id=template_id,  # This will now be the correct template
    )

    if email_config.debug_mode:
        print(
            f"Sent {template_id} template to {student_dict['email']}, charged={cancellation_dict.get('charged', False)}"
        )

    return result


def send_appropriate_cancellation_email(student, cancellation):
    """
    IMPROVED VERSION: Send the right email based on cancellation status
    This replaces send_cancellation_confirmation with better logic
    """

    # Convert sqlite3.Row to dict if needed
    if hasattr(student, "keys"):
        student_dict = dict(student)
    else:
        student_dict = student

    if hasattr(cancellation, "keys"):
        cancellation_dict = dict(cancellation)
    else:
        cancellation_dict = cancellation

    # Debug info
    is_charged = bool(cancellation_dict.get("charged", False))
    student_name = (
        f"{student_dict.get('first_name', '')} {student_dict.get('last_name', '')}"
    )

    if email_config.debug_mode:
        print(f"DEBUG: Sending email for {student_name}")
        print(f"DEBUG: Cancellation charged status: {is_charged}")
        print(
            f"DEBUG: Will use template: {'cancellation_charged' if is_charged else 'free_cancellation'}"
        )

    # STEP 1: Choose the right template
    if is_charged:
        # This is a charged cancellation - use charged template
        template_id = "cancellation_charged"
        fallback_template_id = "client_confirmation"
        print(
            f"✅ Using CHARGED template for {student_name} (Bronze member, 4th cancellation)"
        )
    else:
        # This is a free cancellation - use free template
        template_id = "free_cancellation"
        fallback_template_id = "client_confirmation"
        print(f"✅ Using FREE template for {student_name}")

    # STEP 2: Get the template
    template = get_email_template(template_id)
    if not template:
        print(f"⚠️  Primary template '{template_id}' not found, using fallback")
        template_id = fallback_template_id
        template = get_email_template(fallback_template_id)

    if not template:
        error_msg = f"❌ No email template found (tried {template_id} and {fallback_template_id})"
        print(error_msg)
        return {"success": False, "message": error_msg}

    # STEP 3: Generate variables for the template
    variables = get_template_variables(student_dict, cancellation_dict)

    # Add charge-specific variables
    if is_charged:
        variables["charge_reason"] = get_charge_reason(cancellation_dict)
        variables["status_message"] = (
            f"A charge will be applied to your account. Reason: {variables['charge_reason']}"
        )
    else:
        variables["status_message"] = (
            "This cancellation has been processed at no charge."
        )

    # STEP 4: Process the template
    try:
        body, subject = process_template_variables(
            template["body"], template["subject"], variables
        )
    except Exception as e:
        error_msg = f"❌ Template processing failed: {str(e)}"
        print(error_msg)
        return {"success": False, "message": error_msg}

    # STEP 5: Send the email
    result = send_email(
        student_dict["email"],
        subject,
        body,
        "client",
        template_id=template_id,
    )

    # Log the result
    if result.get("success"):
        print(f"✅ SUCCESS: Sent '{template_id}' email to {student_dict['email']}")
    else:
        print(
            f"❌ FAILED: Could not send '{template_id}' email to {student_dict['email']}: {result.get('message', 'Unknown error')}"
        )

    return result


def send_manager_notification(student, cancellation):
    """Send new cancellation notification to managers - uses settings for manager emails"""
    template = get_email_template("manager_notification")
    if not template:
        return {"success": False, "message": "Template not found"}

    # Convert sqlite3.Row to dict if needed
    if hasattr(student, "keys"):
        student_dict = dict(student)
    else:
        student_dict = student

    if hasattr(cancellation, "keys"):
        cancellation_dict = dict(cancellation)
    else:
        cancellation_dict = cancellation

    variables = get_template_variables(
        student_dict,
        cancellation_dict,
        {
            "action_required": "Review cancellation and approve/charge as needed",
            "dashboard_url": f"{request.url_root if 'request' in globals() else 'http://localhost:5000/'}manager/cancellations?student={student_dict['id']}",
        },
    )

    body, subject = process_template_variables(
        template["body"], template["subject"], variables
    )

    # Get manager emails from settings instead of email_config
    manager_emails = get_manager_emails_from_settings()

    if not manager_emails:
        return {"success": False, "message": "No manager emails configured in settings"}

    # Send to all manager emails
    results = []
    for manager_email in manager_emails:
        result = send_email(
            manager_email, subject, body, "manager", template_id="manager_notification"
        )
        results.append(result)

    # Return success if any email was sent successfully (excluding skipped)
    success_count = sum(1 for r in results if r.get("success") and not r.get("skipped"))
    skipped_count = sum(1 for r in results if r.get("skipped"))

    if skipped_count > 0:
        return {
            "success": True,
            "message": f"Manager notifications disabled (would have sent to {len(manager_emails)} managers)",
            "skipped": True,
        }

    return {
        "success": success_count > 0,
        "message": f"Sent to {success_count}/{len(results)} managers",
    }


def send_override_notification_emails(
    student, cancellation, override_action, override_reason, manager_email,
    suppress_student_email=False, suppress_manager_email=False
):
    """Send emails after manager override - both to client and managers - USES TRIGGER SYSTEM"""

    # Convert data to dicts for consistency
    if hasattr(student, "keys"):
        student_dict = dict(student)
    else:
        student_dict = student

    if hasattr(cancellation, "keys"):
        cancellation_dict = dict(cancellation)
    else:
        cancellation_dict = cancellation

    results = {
        "client_email": {"success": False},
        "manager_notification": {"success": False},
    }

    # Add override-specific information
    override_info = {
        "override_action": override_action,
        "override_reason": override_reason,
        "override_by": manager_email,
        "override_date": toronto_now().strftime("%B %d, %Y at %I:%M %p"),
    }

    # STEP 1: Send updated confirmation to CLIENT using TRIGGER SYSTEM
    try:
        # Check if student email should be suppressed
        if suppress_student_email:
            print(f"🔇 Student email suppressed by manager")
            results["client_email"] = {"success": True, "message": "Suppressed by manager"}
        else:
            print(f"📧 Sending override notification to client: {student_dict['email']}")
            print(
                f"   Action: {override_action}, Charged: {cancellation_dict.get('charged', False)}"
            )

            # DETERMINE THE CORRECT TRIGGER BASED ON OVERRIDE ACTION
            if override_action == "force_free":
                trigger_id = "override_to_free_student"
            elif override_action == "force_charge" or override_action == "charge":
                trigger_id = "override_to_charged_student"
            elif override_action == "approve":
                # For approve action, use the same as force_free (no charge change, just approved)
                trigger_id = "override_to_free_student"
            else:
                # Default to free for unknown actions
                trigger_id = "override_to_free_student"

            print(f"   Using trigger: {trigger_id}")

            # USE TRIGGER SYSTEM to send email
            client_result = send_email_by_trigger(
                trigger_id,
                student_dict,
                cancellation_dict,
                recipient_type="student",
                extra_vars=override_info
            )
            results["client_email"] = client_result

            if client_result.get("success"):
                print(f"✅ Client override email sent successfully")
            else:
                print(
                    f"❌ Client override email failed: {client_result.get('message', 'Unknown error')}"
                )

    except Exception as e:
        print(f"❌ Client override email error: {str(e)}")
        results["client_email"] = {"success": False, "message": str(e)}

    # STEP 2: Send notification to MANAGERS about the override
    try:
        # Check if manager email should be suppressed
        if suppress_manager_email:
            print(f"🔇 Manager email suppressed by manager")
            results["manager_notification"] = {"success": True, "message": "Suppressed by manager"}
        else:
            # Create override notification for managers using TRIGGER SYSTEM
            
            # Determine which manager trigger to use
            if override_action == "force_free":
                manager_trigger_id = "override_to_free_manager"
            elif override_action == "force_charge" or override_action == "charge":
                manager_trigger_id = "override_to_charged_manager"
            elif override_action == "approve":
                manager_trigger_id = "override_to_free_manager"
            else:
                manager_trigger_id = "override_to_free_manager"

            print(f"   Manager trigger: {manager_trigger_id}")

            # USE TRIGGER SYSTEM to send to managers
            manager_result = send_email_by_trigger(
                manager_trigger_id,
                student_dict,
                cancellation_dict,
                recipient_type="manager",
                extra_vars=override_info
            )

            if manager_result.get("success"):
                results["manager_notification"] = {
                    "success": True,
                    "message": manager_result.get("message", "Manager notifications sent")
                }
                print(f"📧 Manager override notifications sent")
            else:
                results["manager_notification"] = {
                    "success": False,
                    "message": manager_result.get("message", "Failed to send manager notifications")
                }
                print(f"⚠️ Manager notifications failed: {manager_result.get('message')}")

    except Exception as e:
        print(f"❌ Manager override notification error: {str(e)}")
        results["manager_notification"] = {"success": False, "message": str(e)}

    return results


# ===================================
# EMAIL TESTING FUNCTION
# ===================================


def test_email_configuration():
    """Test email configuration and connectivity"""
    try:
        if not email_config.is_configured:
            return {
                "success": False,
                "message": "Email not configured - missing SMTP credentials",
                "details": {
                    "server": email_config.smtp_server,
                    "port": email_config.smtp_port,
                    "user_set": bool(email_config.smtp_user),
                    "password_set": bool(email_config.smtp_password),
                },
            }

        # Test connection
        server = email_config.get_connection()
        server.quit()

        return {
            "success": True,
            "message": "Email configuration is working correctly",
            "details": {
                "server": email_config.smtp_server,
                "port": email_config.smtp_port,
                "from_email": email_config.from_email,
                "manager_emails": get_manager_emails_from_settings(),
                "client_notifications_enabled": should_send_email("client"),
                "manager_notifications_enabled": should_send_email("manager"),
            },
        }

    except Exception as e:
        return {
            "success": False,
            "message": f"Email configuration test failed: {str(e)}",
            "details": {
                "server": email_config.smtp_server,
                "port": email_config.smtp_port,
                "error": str(e),
            },
        }


# Add these functions to your app.py file to standardize date formatting


@app.template_filter("format_date")
def format_date_filter(date_value):
    """Format date as 'August 28, 2025'"""
    return format_date_for_display(date_value)


@app.template_filter("format_datetime")
def format_datetime_filter(dt_value):
    """Format datetime as 'August 28, 2025 at 3:30 PM PST'"""
    return format_datetime_for_display(dt_value)


@app.template_filter("format_time")
def format_time_filter(time_value):
    """Format time as '3:30 PM'"""
    if time_value is None:
        return "Unknown"

    if isinstance(time_value, str):
        try:
            # Handle both HH:MM:SS and HH:MM formats
            if len(time_value.split(":")) == 3:
                time_value = datetime.strptime(time_value, "%H:%M:%S").time()
            else:
                time_value = datetime.strptime(time_value, "%H:%M").time()
        except (ValueError, TypeError):
            return str(time_value)

    if isinstance(time_value, time_type):
        return time_value.strftime("%I:%M %p")

    return str(time_value)


# ===================================
# APPLICATION INITIALIZATION
# ===================================


# ============================================
# MANAGER SUBMISSION FEATURE - NEW ROUTES
# ============================================


@app.route("/manager/api/policy-impact", methods=["GET"])
@login_required
@admin_required
def policy_impact():
    """Get policy impact for a given student and submission date"""
    try:
        student_id = request.args.get("student_id")
        submission_date = request.args.get("submission_date")
        
        if not student_id or not submission_date:
            return jsonify({"error": "Missing required parameters"}), 400
        
        conn = get_db()
        conn.row_factory = sqlite3.Row
        
        # Get student
        student = conn.execute(
            "SELECT membership_level FROM students WHERE id = ?",
            (student_id,)
        ).fetchone()
        
        if not student:
            conn.close()
            return jsonify({"error": "Student not found"}), 404
        
        # Get current month from submission date
        from datetime import datetime
        try:
            submission_dt = datetime.strptime(submission_date, "%Y-%m-%d")
        except:
            conn.close()
            return jsonify({"error": "Invalid date format"}), 400
        
        month_display = submission_dt.strftime("%B %Y")
        
        # Count free cancellations for the lesson's month (not submission month)
        # This matches how will_be_charged() counts - by lesson_date month
        month_str = submission_dt.strftime("%Y-%m")
        count = conn.execute(
            """
            SELECT COUNT(*) as count FROM cancellations
            WHERE student_id = ?
            AND strftime('%Y-%m', lesson_date) = ?
            AND excluded = 0
            AND charged = 0
        """,
            (student_id, month_str),
        ).fetchone()["count"]
        
        # Get tier info
        tier = conn.execute(
            "SELECT free_notices FROM membership_tiers WHERE level = ?",
            (student["membership_level"],)
        ).fetchone()
        
        free_notices = tier["free_notices"] if tier else 1
        
        if count >= free_notices:
            impact_text = f"This cancellation will be CHARGED (monthly limit reached: {count}/{free_notices})"
        else:
            impact_text = f"This cancellation will be FREE ({count + 1}/{free_notices} used)"
        
        conn.close()
        
        return jsonify({
            "month_display": month_display,
            "impact_text": impact_text,
            "success": True
        })
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500


@app.route("/manager/submit-cancellation", methods=["GET", "POST"])
@login_required
@admin_required
def manager_submit_cancellation():
    """
    Allow managers to submit cancellations for any student.
    GET: Display form
    POST: Process submission
    """
    if request.method == "GET":
        try:
            conn = get_db()
            students = conn.execute(
                "SELECT id, first_name, last_name, email, membership_level FROM students ORDER BY first_name, last_name"
            ).fetchall()
            conn.close()

            return render_template(
                "manager_submit_cancellation.html",
                students=students,
                manager_name=session.get("user_name", "Manager"),
            )
        except Exception as e:
            flash(f"Error loading form: {str(e)}", "error")
            return redirect(url_for("manager_dashboard"))

    elif request.method == "POST":
        try:
            student_id = request.form.get("student_id")
            lesson_date = request.form.get("lesson_date")
            lesson_time = request.form.get("lesson_time")
            submission_date_time = request.form.get(
                "submission_date_time"
            )  # Format: "2026-04-02 14:30"
            suppress_notifications = request.form.get("suppress_notifications") == "1"
            # FIX: Read granular notification suppression flags so "Don't Notify
            # Student" / "Don't Notify Managers" actually take effect (previously
            # only the "Suppress All" checkbox was honored).
            suppress_to_student = request.form.get("suppress_to_student") == "1"
            suppress_to_manager = request.form.get("suppress_to_manager") == "1"
            # Effective per-channel flags: "Suppress All" overrides individual flags
            suppress_student_email = suppress_notifications or suppress_to_student
            suppress_manager_email = suppress_notifications or suppress_to_manager
            manager_notes = request.form.get("manager_notes", "")
            
            # NEW: Handle sequential lessons and reschedule fields
            sequential_dates = request.form.getlist("sequential_dates[]")
            sequential_times = request.form.getlist("sequential_times[]")
            # FIX: Manager portal uses a single "return_date" input instead of
            # per-lesson sequential dates. We expand it into a sequential_lessons
            # list below so the dashboard and emails correctly show the cancelled
            # range and computed return date.
            return_date = request.form.get("return_date", "").strip()
            wants_reschedule = request.form.get("wants_reschedule") == "1"
            reschedule_preferences = request.form.get("reschedule_preferences", "")

            # DEBUG: Print what we received
            print(f"DEBUG: Received form data:")
            print(f"  student_id: {student_id}")
            print(f"  lesson_date: {lesson_date}")
            print(f"  lesson_time: {lesson_time}")
            print(f"  submission_date_time: {submission_date_time}")
            print(f"  manager_notes: {manager_notes}")
            print(f"  suppress_notifications: {suppress_notifications}")
            print(f"  suppress_to_student: {suppress_to_student}")
            print(f"  suppress_to_manager: {suppress_to_manager}")
            print(f"  sequential_dates: {sequential_dates}")
            print(f"  sequential_times: {sequential_times}")
            print(f"  return_date: {return_date}")
            print(f"  wants_reschedule: {wants_reschedule}")
            print(f"  reschedule_preferences: {reschedule_preferences}")

            # Validate inputs
            if not all([student_id, lesson_date, lesson_time, submission_date_time]):
                missing = []
                if not student_id:
                    missing.append("student_id")
                if not lesson_date:
                    missing.append("lesson_date")
                if not lesson_time:
                    missing.append("lesson_time")
                if not submission_date_time:
                    missing.append("submission_date_time")
                return (
                    jsonify(
                        {
                            "success": False,
                            "message": f"Missing required fields: {', '.join(missing)}",
                        }
                    ),
                    400,
                )

            conn = get_db()
            conn.row_factory = sqlite3.Row  # Enable dictionary-like access

            # Verify student exists
            student = conn.execute(
                "SELECT * FROM students WHERE id = ?", (student_id,)
            ).fetchone()
            if not student:
                conn.close()
                return jsonify({"success": False, "message": "Student not found"}), 404

            # CRITICAL FIX: Convert sqlite3.Row to dict to use .get() method
            student = dict(student)

            print(
                f"DEBUG: Found student: {student['first_name']} {student['last_name']}"
            )

            # ⭐ CRITICAL: Calculate will_be_charged immediately (just like student submission does!)
            try:
                # Parse lesson datetime
                lesson_datetime = parse_lesson_datetime(lesson_date, lesson_time)
                
                # Parse submission datetime (manager's specified date)
                submission_datetime = datetime.strptime(submission_date_time, "%Y-%m-%d %H:%M")
                pacific_tz = pytz.timezone("America/Los_Angeles")
                if submission_datetime.tzinfo is None:
                    submission_datetime = pacific_tz.localize(submission_datetime)
                
                # ⭐ THIS IS THE FIX: Call will_be_charged with manager's submission date
                will_charge, charge_reason = will_be_charged(dict(student), lesson_datetime, submission_datetime)
                
                # Calculate deadline_passed for this tier using tier-aware logic.
                # Bronze/Silver/etc. use "<time> previous day" (absolute), Gold uses
                # "X hours before lesson" (relative). The helper handles both.
                tier = get_membership_tier(student["membership_level"])
                deadline_passed = (
                    is_after_deadline(tier, lesson_datetime, submission_datetime)
                    if tier else False
                )
                
                print(f"DEBUG: will_be_charged={will_charge}, charge_reason={charge_reason}")
                print(f"DEBUG: deadline_passed={deadline_passed}")
                
            except Exception as calc_error:
                print(f"DEBUG: Error calculating charge status: {str(calc_error)}")
                will_charge = False  # Default to free if calculation fails
                deadline_passed = False
                charge_reason = "Calculation failed, defaulting to free"

            # Create cancellation
            try:
                # Prepare sequential lessons data.
                # Two input shapes are supported:
                #   (a) explicit per-lesson arrays (sequential_dates[]/sequential_times[])
                #   (b) a single return_date — the manager portal's current UI shape:
                #       every day between lesson_date+1 and return_date-1 is treated
                #       as a cancelled lesson at the same time-of-day as the original.
                # The dashboard/email "All lessons are cancelled from X until the
                # return date of Y" message is generated by format_sequential_lessons,
                # which uses the LAST entry's date + 1 day as the return date — so
                # filling in the in-between days makes that display work correctly.
                sequential_lessons = []
                if sequential_dates and sequential_times:
                    for seq_date, seq_time in zip(sequential_dates, sequential_times):
                        if seq_date and seq_time:
                            sequential_lessons.append({"date": seq_date, "time": seq_time})
                elif return_date:
                    try:
                        original_date_obj = datetime.strptime(lesson_date, "%Y-%m-%d").date()
                        return_date_obj = datetime.strptime(return_date, "%Y-%m-%d").date()
                        # Only expand if the return date is strictly after the lesson date.
                        if return_date_obj > original_date_obj:
                            # Cancelled range: day AFTER the original lesson up to the
                            # day BEFORE the return date (the return date itself is
                            # when the student resumes lessons, so it is NOT cancelled).
                            cursor_date = original_date_obj + timedelta(days=1)
                            last_cancelled_date = return_date_obj - timedelta(days=1)
                            while cursor_date <= last_cancelled_date:
                                sequential_lessons.append({
                                    "date": cursor_date.strftime("%Y-%m-%d"),
                                    "time": lesson_time,
                                })
                                cursor_date += timedelta(days=1)
                            print(
                                f"DEBUG: Expanded return_date {return_date} into "
                                f"{len(sequential_lessons)} sequential lesson(s)"
                            )
                    except Exception as range_error:
                        print(f"DEBUG: Failed to expand return_date range: {range_error}")
                
                sequential_lessons_json = json.dumps(sequential_lessons) if sequential_lessons else None

                # Get status details before inserting
                status_details = get_cancellation_status_details(int(student_id))
                
                # submission_date_time is in format "2026-04-02 14:30"
                # Store as-is in manual_submission_date (the "submission date/time" for policy)
                cursor = conn.execute(
                    """
                    INSERT INTO cancellations (
                        student_id, lesson_date, lesson_time, sequential_lessons,
                        reschedule_requested, reschedule_preferences, status, created_at,
                        submitted_by, submission_method, actual_submission_timestamp,
                        manual_submission_date, manager_submitted_by, is_manager_submission,
                        charged, deadline_passed, is_override, 
                        manager_notes, suppress_notifications, status_details
                    ) VALUES (?, ?, ?, ?, ?, ?, 'pending', CURRENT_TIMESTAMP, 'manager', 'manager_portal',
                             CURRENT_TIMESTAMP, ?, ?, 1, ?, ?, 0, ?, ?, ?)
                """,
                    (
                        student_id,
                        lesson_date,
                        lesson_time,
                        sequential_lessons_json,
                        wants_reschedule,
                        reschedule_preferences,
                        submission_date_time,
                        session.get("user_id"),
                        will_charge,           # ⭐ NOW SET CORRECTLY!
                        deadline_passed,       # ⭐ NOW SET CORRECTLY!
                        manager_notes,
                        suppress_notifications,
                        status_details,  # NEW: Add status details
                    ),
                )

                conn.commit()
                # NOTE: welcome_free_used is NOT set here even for Welcome Package free
                # cancellations. The cancellation is status='pending' and must still be
                # approved by a manager. The flag is set at approval time in
                # process_cancellation / process_all_pending, matching the student
                # submission flow. Until then, will_be_charged() uses a dynamic count of
                # charged=0 cancellations to determine lifetime-free usage correctly.

                # Get the cancellation ID for email sending
                cancellation_id = cursor.lastrowid
                
                # Send notifications based on per-channel suppression flags.
                # "Suppress All" overrides individual flags (combined into
                # suppress_student_email / suppress_manager_email above).
                # Send email to student (if not suppressed for student)
                if not suppress_student_email:
                    try:
                        cancellation_data = conn.execute(
                            """SELECT c.*, s.first_name, s.last_name, s.email
                               FROM cancellations c
                               JOIN students s ON c.student_id = s.id
                               WHERE c.id = ?""",
                            (cancellation_id,)
                        ).fetchone()
                        
                        if cancellation_data:
                            # Get full variables using get_template_variables
                            student_record = conn.execute(
                                "SELECT * FROM students WHERE id = ?", (student_id,)
                            ).fetchone()
                            
                            # Convert sqlite3.Row to dict for .get() method compatibility
                            cancellation_dict = dict(cancellation_data) if hasattr(cancellation_data, 'keys') else cancellation_data
                            
                            full_variables = get_template_variables(student_record, cancellation_dict)
                            
                            # Get email template
                            template = get_email_template("manager_submits_cancellation_student")
                            if template:
                                # Override with manager-specific variables.
                                # NOTE: sequential_lessons is intentionally NOT overridden here —
                                # the value from full_variables (set by get_template_variables /
                                # format_sequential_lessons) is the correct human-readable
                                # "All lessons are cancelled from X until the return date of Y"
                                # text. The previous override used status_details (e.g. "1 of 2
                                # used this month"), which produced wrong/missing range info in
                                # emails when a return date was provided.
                                variables = {
                                    **full_variables,
                                    "client_name": safe_value(f"{cancellation_dict.get('first_name', '')} {cancellation_dict.get('last_name', '')}".strip()),
                                    "lesson_date": safe_value(full_variables.get("lesson_date", cancellation_dict.get('lesson_date', ''))),
                                    "lesson_time": safe_value(full_variables.get("lesson_time", format_time_for_email(cancellation_dict.get('lesson_time', '')))),
                                    "submission_time": safe_value(submission_date_time),
                                    "cancellation_status": "CHARGED" if will_charge else "FREE",
                                    "status_details": safe_value(status_details),
                                    "charge_reason": safe_value(charge_reason),
                                    "membership_tier": safe_value(student.get('membership_level')),
                                    "policy_url": "https://www.riversideequestrian.ca/cancellations",
                                    "website_url": "https://www.riversideequestrian.ca",
                                    "contact_email": "stav@riversideequestrian.ca",
                                    "company_name": "Riverside Equestrian",
                                }
                                body, subject = process_template_variables(template["body"], template["subject"], variables)
                                send_email(cancellation_dict["email"], subject, body, "client")
                    except Exception as email_error:
                        print(f"Error sending student notification: {str(email_error)}")
                else:
                    print("🔇 Student notification suppressed by manager selection")
                
                # Send email to ALL managers (if not suppressed for managers)
                if not suppress_manager_email:
                    try:
                        # Get manager emails from database settings (comma-separated list)
                        manager_emails = get_manager_emails_from_settings()
                        print(f"DEBUG: Manager emails from settings: {manager_emails}")
                        
                        if manager_emails:
                            template = get_email_template("manager_submits_cancellation_manager")
                            print(f"DEBUG: Template found: {template is not None}")
                            
                            if template:
                                # Get full variables once (used for all managers)
                                student_record = conn.execute(
                                    "SELECT * FROM students WHERE id = ?", (student_id,)
                                ).fetchone()
                                
                                cancellation_record = conn.execute(
                                    "SELECT * FROM cancellations WHERE id = ?", (cancellation_id,)
                                ).fetchone()
                                
                                full_variables = get_template_variables(student_record, cancellation_record)
                                
                                # NOTE: sequential_lessons is intentionally NOT overridden —
                                # see equivalent comment in the student-email block above.
                                variables = {
                                    **full_variables,
                                    "client_name": safe_value(f"{student.get('first_name', '')} {student.get('last_name', '')}".strip()),
                                    "parent_name": safe_value(student.get('parent_name', student.get('parent_first', '') + ' ' + student.get('parent_last', '')).strip()),
                                    "client_email": safe_value(student.get('email')),
                                    "client_phone": safe_value(student.get('phone')),
                                    "membership_tier": safe_value(student.get('membership_level')),
                                    "lesson_date": safe_value(full_variables.get("lesson_date", lesson_date)),
                                    "lesson_time": safe_value(full_variables.get("lesson_time", format_time_for_email(lesson_time))),
                                    "submission_time": safe_value(submission_date_time),
                                    "cancellation_status": "CHARGED" if will_charge else "FREE",
                                    "status_details": safe_value(status_details),
                                    "charge_reason": safe_value(charge_reason),
                                    "policy_url": "https://www.riversideequestrian.ca/cancellations",
                                    "website_url": "https://www.riversideequestrian.ca",
                                    "contact_email": "stav@riversideequestrian.ca",
                                    "company_name": "Riverside Equestrian",
                                }

                                body, subject = process_template_variables(template["body"], template["subject"], variables)
                                
                                # Send to ALL managers
                                print(f"DEBUG: Sending manager notification to {len(manager_emails)} managers")
                                for manager_email in manager_emails:
                                    try:
                                        result = send_email(manager_email, subject, body, "manager")
                                        print(f"DEBUG: Manager email send result to {manager_email}: {result}")
                                        if result.get("success"):
                                            print(f"✅ Manager email sent successfully to {manager_email}")
                                        else:
                                            print(f"❌ Manager email failed for {manager_email}: {result.get('message')}")
                                    except Exception as mgr_error:
                                        print(f"❌ Error sending to manager {manager_email}: {str(mgr_error)}")
                            else:
                                print(f"DEBUG: Template not found for manager_submits_cancellation_manager")
                        else:
                            print(f"DEBUG: No manager emails configured in settings")
                    except Exception as email_error:
                        print(f"❌ Error in manager notification process: {str(email_error)}")
                        import traceback
                        traceback.print_exc()
                else:
                    print("🔇 Manager notifications suppressed by manager selection")
                
                conn.close()

                print(
                    f"DEBUG: Cancellation created successfully with submission_date_time: {submission_date_time}"
                )

                flash("Cancellation submitted successfully!", "success")

                return jsonify(
                    {"success": True, "redirect": url_for("manager_cancellations")}
                )
            except Exception as db_error:
                conn.close()
                print(f"DEBUG: Database error: {str(db_error)}")
                return (
                    jsonify(
                        {
                            "success": False,
                            "message": f"Database error: {str(db_error)}",
                        }
                    ),
                    500,
                )

        except Exception as e:
            import traceback

            print(f"DEBUG: Exception: {str(e)}")
            traceback.print_exc()
            return jsonify({"success": False, "message": str(e)}), 500


@app.route("/manager/api/student-search")
@login_required
@admin_required
def student_search():
    """AJAX endpoint for student search autocomplete"""
    query = request.args.get("q", "").lower()

    if len(query) < 2:
        return jsonify([])

    try:
        conn = get_db()
        conn.row_factory = sqlite3.Row  # Enable dictionary-like access
        students = conn.execute(
            """
            SELECT id, first_name, last_name, email, phone, parent_first, parent_last, membership_level
            FROM students
            WHERE (LOWER(first_name) LIKE ? 
                   OR LOWER(last_name) LIKE ? 
                   OR LOWER(email) LIKE ?
                   OR LOWER(parent_first) LIKE ?
                   OR LOWER(parent_last) LIKE ?
                   OR LOWER(phone) LIKE ?)
            ORDER BY first_name, last_name
            LIMIT 10
        """,
            (f"%{query}%", f"%{query}%", f"%{query}%", f"%{query}%", f"%{query}%", f"%{query}%"),
        ).fetchall()
        conn.close()

        results = []
        for s in students:
            first_name = s["first_name"] if s["first_name"] else ""
            last_name = s["last_name"] if s["last_name"] else ""
            full_name = f"{first_name} {last_name}".strip()

            results.append(
                {
                    "id": s["id"],
                    "name": full_name if full_name else "Unknown Student",
                    "email": s["email"] if s["email"] else "No email",
                    "membership": (
                        s["membership_level"] if s["membership_level"] else "Standard"
                    ),
                    "phone": s["phone"] if s["phone"] else "No phone",
                    "parent_first": s["parent_first"] if s["parent_first"] else "",
                    "parent_last": s["parent_last"] if s["parent_last"] else "",
                }
            )

        return jsonify(results)

    except Exception as e:
        import traceback

        traceback.print_exc()  # Print error to Flask console
        return jsonify({"error": str(e)}), 500



@app.route("/manager/api/student-policy-info/<int:student_id>")
@login_required
@admin_required
def student_policy_info(student_id):
    """Get current month policy info for a student"""
    try:
        conn = get_db()
        conn.row_factory = sqlite3.Row

        # Get student info
        student = conn.execute(
            "SELECT id, first_name, last_name, email, membership_level FROM students WHERE id = ?",
            (student_id,),
        ).fetchone()

        if not student:
            conn.close()
            return jsonify({"error": "Student not found"}), 404

        # Get the membership tier info (FIXED: was hardcoded to 1)
        membership = student["membership_level"] or "Standard"
        tier = get_membership_tier(membership)
        
        if not tier:
            conn.close()
            return jsonify({"error": "Membership tier not found"}), 404
        
        tier_dict = dict(tier) if hasattr(tier, "keys") else tier
        free_notices = tier_dict.get("free_notices", 1)
        deadline_display = tier_dict.get("deadline_display", "Varies by membership")

        # Get current month cancellations count (based on lesson month for this month)
        from datetime import datetime
        pacific_now = toronto_now()
        current_month = pacific_now.month
        current_year = pacific_now.year

        # Count free cancellations for the current month based on lesson_date
        student_count = conn.execute(
            """
            SELECT COUNT(*) as count FROM cancellations
            WHERE student_id = ?
            AND strftime('%m', lesson_date) = ?
            AND strftime('%Y', lesson_date) = ?
            AND charged = 0
            AND excluded = 0
        """,
            (student_id, f"{current_month:02d}", str(current_year)),
        ).fetchone()["count"]

        conn.close()

        # Get status details
        status_details = get_cancellation_status_details(student_id)

        return jsonify(
            {
                "membership_level": membership,
                "free_notices_per_month": free_notices,
                "deadline_display": deadline_display,
                "current_month_count": student_count,
                "free_remaining": max(0, free_notices - student_count),
                "status_details": status_details,
            }
        )

    except Exception as e:
        import traceback

        traceback.print_exc()
        return jsonify({"error": str(e)}), 500


if __name__ == "__main__":
    import sys

    # Check command line arguments for database operations
    if len(sys.argv) > 1:
        if sys.argv[1] == "reset-db":
            print("Resetting database...")
            if reset_database():
                print("Database reset successfully!")
            else:
                print("Database reset failed!")
            sys.exit()
        elif sys.argv[1] == "verify-db":
            print("Verifying database...")
            if verify_database():
                print("Database verification passed!")
            else:
                print("Database verification failed!")
            sys.exit()

    # Check if database exists and is properly initialized
    if not os.path.exists(app.config["DATABASE"]):
        print("Database does not exist. Initializing...")
        if init_db():
            print("Database initialized successfully!")
            print("\nDefault Login Credentials:")
            print("Senior Manager:       / admin123")
            print("Manager: manager@riversideequestrian.ca / admin123")
            print("Client: chloechow2016@gmail.com (no password needed)")
            print("-" * 50)
        else:
            print("Database initialization failed!")
            sys.exit(1)
    else:
        # Verify existing database
        print("Verifying existing database...")
        if not verify_database():
            print("Database verification failed. Attempting to reset...")
            if reset_database():
                print("Database reset and reinitialized successfully!")
            else:
                print("Database reset failed!")
                sys.exit(1)
        else:
            print("Database verification passed!")

    print("Starting Flask application...")
    app.run(debug=False, host="0.0.0.0", port=5000)